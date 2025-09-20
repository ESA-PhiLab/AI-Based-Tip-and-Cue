import os
from openpyxl import Workbook, load_workbook
import math
from datetime import timedelta
# import pandas as pd
import inspect

import os
import shutil
import atexit
import numpy as np

import sys
import time, shutil, os

import matplotlib.pyplot as plt
import pandas as pd

from matplotlib.ticker import MultipleLocator
import matplotlib
import matplotlib.pyplot as plt

plt.style.use("seaborn-v0_8-whitegrid")
matplotlib.use("TkAgg")

from .plotting.plot_pyvista import close_plotter_safely

import uuid


from simulation.plotting.plot_functions import plot_offnadir_distribution, plot_latency_distribution

def init_excel_log(path, header, sheet_name="Log"):
    if os.path.exists(path):
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws.append(header)
    wb.save(path)
    wb.close()
    return {"path": path, "sheet": sheet_name, "header": header}


def append_excel_log(writer, row):
    path = writer["path"]
    sheet = writer["sheet"]
    header = writer["header"]

    wb = load_workbook(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    if ws.max_row == 0:
        ws.append(header)
    ws.append(row)
    wb.save(path)
    wb.close()

def _to_scalar(val):
    """Convert NumPy scalars/arrays or None into plain Python types safe for Excel."""
    if val is None:
        return None
    if isinstance(val, (list, tuple, np.ndarray)):
        if len(val) == 0:
            return None
        return float(val[0])  # first element if it's an array/tuple
    try:
        return float(val)
    except Exception:
        return val

def format_hms(seconds: float) -> str:
    h, rem = divmod(int(seconds), 3600)
    m, s   = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def log_tip(writer, detection_id, target_id, tip_actor,
            tip_observation_date, tip_confirmation_date,
            tip_ai_decision, true_label, correct,
            offnadir_deg, gsd_m,
            target_lat, target_lon, target_alt,
            tip_lat, tip_lon, tip_alt,
            x, y, z, vx, vy, vz):
    """Log a TIP observation/confirmation event matching header_tip."""
    row = [
        detection_id,
        target_id,
        tip_actor,
        tip_observation_date.isoformat(timespec="seconds") + "Z" if tip_observation_date else None,
        tip_confirmation_date.isoformat(timespec="seconds") + "Z" if tip_confirmation_date else None,
        tip_ai_decision,
        true_label,
        correct,
        _to_scalar(offnadir_deg),
        _to_scalar(gsd_m),
        _to_scalar(target_lat), _to_scalar(target_lon), _to_scalar(target_alt),
        _to_scalar(tip_lat), _to_scalar(tip_lon), _to_scalar(tip_alt),
        _to_scalar(x), _to_scalar(y), _to_scalar(z),
        _to_scalar(vx), _to_scalar(vy), _to_scalar(vz)
    ]
    append_excel_log(writer, row)


def log_cue(writer, detection_id, target_id, cue_actor,
            cue_observation_date, cue_confirmation_date,
            cue_ai_decision, true_label, correct,
            offnadir_deg, gsd_m, viewing_time,
            latency_observation, latency_confirmation, slew_stab_time,
            target_lat, target_lon, target_alt,
            cue_lat, cue_lon, cue_alt,
            x, y, z, vx, vy, vz, roll, pitch, yaw):
    """Log a CUE observation/confirmation event matching header_cue."""
    row = [
        detection_id,
        target_id,
        cue_actor,
        cue_observation_date.isoformat(timespec="seconds") + "Z" if cue_observation_date else None,
        cue_confirmation_date.isoformat(timespec="seconds") + "Z" if cue_confirmation_date else None,
        cue_ai_decision,
        true_label,
        correct,
        _to_scalar(offnadir_deg),
        _to_scalar(gsd_m),
        _to_scalar(viewing_time),
        _to_scalar(latency_observation),
        _to_scalar(latency_confirmation),
        _to_scalar(slew_stab_time),
        _to_scalar(target_lat), _to_scalar(target_lon), _to_scalar(target_alt),
        _to_scalar(cue_lat), _to_scalar(cue_lon), _to_scalar(cue_alt),
        _to_scalar(x), _to_scalar(y), _to_scalar(z),
        _to_scalar(vx), _to_scalar(vy), _to_scalar(vz),
        _to_scalar(roll), _to_scalar(pitch), _to_scalar(yaw)
    ]
    append_excel_log(writer, row)



def log_combined(writer, detection_id, target_id, tip_actor, cue_actor,
                 tip_observation_date, tip_confirmation_date,
                 cue_observation_date, cue_confirmation_date,
                 tip_ai_decision, cue_ai_decision,
                 true_label, correct,
                 offnadir_deg, gsd_m, viewing_time,
                 latency_observation, latency_confirmation,
                 target_lat, target_lon, target_alt,
                 cue_lat, cue_lon, cue_alt):
    """Log a combined TIP+CUE event matching header_combined."""
    row = [
        detection_id,
        target_id,
        tip_actor,
        cue_actor,
        tip_observation_date.isoformat(timespec="seconds") + "Z" if tip_observation_date else None,
        tip_confirmation_date.isoformat(timespec="seconds") + "Z" if tip_confirmation_date else None,
        cue_observation_date.isoformat(timespec="seconds") + "Z" if cue_observation_date else None,
        cue_confirmation_date.isoformat(timespec="seconds") + "Z" if cue_confirmation_date else None,
        tip_ai_decision,
        cue_ai_decision,
        true_label,
        correct,
        _to_scalar(offnadir_deg),
        _to_scalar(gsd_m),
        _to_scalar(viewing_time),
        _to_scalar(latency_observation),
        _to_scalar(latency_confirmation),
        _to_scalar(target_lat), _to_scalar(target_lon), _to_scalar(target_alt),
        _to_scalar(cue_lat), _to_scalar(cue_lon), _to_scalar(cue_alt)
    ]
    append_excel_log(writer, row)



def log_img(writer, detection_id, cue_lat, cue_lon, cue_alt,
            tgt_lat, tgt_lon, tgt_alt, t_datetime, dem_seed):
    """Log an image generation event matching header_img_gen."""
    row = [
        detection_id,
        _to_scalar(cue_lat), _to_scalar(cue_lon), _to_scalar(cue_alt),
        _to_scalar(tgt_lat), _to_scalar(tgt_lon), _to_scalar(tgt_alt),
        t_datetime.isoformat(timespec="seconds") + "Z" if t_datetime else None,
        _to_scalar(dem_seed)
    ]
    append_excel_log(writer, row)



def gsd_offnadir(Pn_m, H_m, offnadir_deg):
    """
    Compute exact off-nadir ground sampling distance (GSD).

    Parameters
    ----------
    Pn_m : float
        Nadir ground sampling distance [m]
    H_m : float
        Sensor altitude above ground [m]
    offnadir_deg : float
        Off-nadir angle [deg]

    Returns
    -------
    Ptheta_m : float
        Off-nadir ground sampling distance [m]
    """

    theta = math.radians(offnadir_deg)

    # Step 1: pixel angular width beta from nadir GSD
    beta = 2.0 * math.atan(Pn_m / (2.0 * H_m))

    # Step 2: exact off-nadir GSD
    Ptheta_m = H_m * (math.tan(theta + beta/2.0) - math.tan(theta - beta/2.0))

    return Ptheta_m


def should_log_event(writer, whale_idx, t_datetime, min_gap_sec=600):
    import pandas as pd
    from datetime import timedelta

    path = writer["path"]
    sheet = writer["sheet"]

    if not os.path.exists(path):
        return True

    df = pd.read_excel(path, sheet_name=sheet)
    if df.empty:
        return True

    df_target = df[df["target_id"] == whale_idx]
    if df_target.empty:
        return True

    last_time = pd.to_datetime(df_target["t_observation_tip"].max())
    if (t_datetime - last_time) > timedelta(seconds=min_gap_sec):
        return True

    return False


def safe_move(src, dst, retries=5, delay=1.0):
    for i in range(retries):
        try:
            shutil.move(src, dst)
            print(f"Moved {src} to {dst.replace(os.sep, '/')}")
            return True
        except PermissionError:
            print(f"Retry {i+1}/{retries}: could not move {src}, still locked.")
            time.sleep(delay)
        except Exception as e:
            print(f"Error moving {src} to {dst}: {e}")
            return False
    print(f"Warning: could not move {src} after {retries} retries.")
    return False


def merge_tip_cue_combined(file_path: str) -> None:
    """Merge observation + confirmation rows for TIP, CUE, COMBINED sheets.
    Keeps the most recent entry if duplicates exist and overwrites the sheets in place.

    Args:
        file_path (str): Path to Excel workbook.
    """
    key_col = "detection_id"
    sheet_names = ["Tip", "Cue", "Combined"]

    all_sheets = pd.read_excel(file_path, sheet_name=sheet_names)

    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet, df in all_sheets.items():
            if key_col not in df.columns:
                raise ValueError(f"{sheet} sheet has no column '{key_col}'")

            # Reverse so last row is processed first
            df = df.iloc[::-1]

            merged_df = (
                df.groupby(key_col, dropna=False, sort=False)
                  .apply(lambda g: g.ffill().bfill().iloc[0])
                  .reset_index(drop=True)
            )

            # Restore order
            merged_df = merged_df.iloc[::-1].reset_index(drop=True)

            # Overwrite original sheet
            merged_df.to_excel(writer, sheet_name=sheet, index=False)





def at_exit(save_name, pl=None, sun_light=None, verbose_def=False, verbose_error=False):

    results_dir = os.path.join("0_results", save_name)
    os.makedirs(results_dir, exist_ok=True)

    if pl is not None:
        try:
            close_plotter_safely(pl, sun_light=sun_light)
            time.sleep(0.1)
            if verbose_def:
                print("Closed pyvista plotter")
        except Exception as e:
            if verbose_error:
                print(f"Could not close pyvista plotter: {e}")

    rename_map = {
        "sim_output.xlsx": f"results_{save_name}.xlsx",
        "simulation.mp4": f"mov_{save_name}.mov",
        "output.log": f"logs_{save_name}.log",
        "offnadir.png": f"plot_offnadir_{save_name}.png",
        "viewing_time.png": f"plot_viewing_time_{save_name}.png",
        "latency_observation.png": f"plot_latency_observation_{save_name}.png",
        "latency_confirmation.png": f"plot_latency_confirmation_{save_name}.png",
        "footprints_tip.html": f"footprints_tip_{save_name}.html",
        "footprints_cue.html": f"footprints_cue_{save_name}.html"
    }

    merged_excel_path = None

    for src, new_name in rename_map.items():
        if src == "output.log" and isinstance(sys.stdout, Logger):
            try:
                sys.stdout.close()
                sys.stdout = sys.__stdout__
                sys.stderr = sys.__stderr__
                time.sleep(0.1)
            except Exception as e:
                if verbose_error:
                    print(f"Could not close print logs: {e}")

        if os.path.exists(src):
            dst = os.path.join(results_dir, new_name)
            try:
                safe_move(src, dst)
            except:
                pass

        else:
            if verbose_error:
                print(f"Warning: {src} not found, skipping.")

    time.sleep(0.1)
    print(f"Saved results in {results_dir.replace(os.sep, '/')}")





def compute_stats(series: pd.Series):
    """Return mean, min, max, std for a pandas Series (ignoring NaN)."""
    s = series.dropna()
    if s.empty:
        return float("nan"), float("nan"), float("nan"), float("nan")
    return s.mean(), s.min(), s.max(), s.std()



class Logger:
    def __init__(self, filename):
        self.terminal = sys.__stdout__
        self.log = open(filename, "a")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        self.terminal.flush()
        self.log.flush()

    def close(self):
        try:
            self.log.close()
        except:
            pass