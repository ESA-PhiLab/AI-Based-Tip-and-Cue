import os
from openpyxl import Workbook, load_workbook
import math
from datetime import timedelta
import pandas as pd
import inspect

import os
import shutil
import atexit
import numpy as np

import sys
import time, shutil, os

import matplotlib.pyplot as plt
import pandas as pd

from matplotlib.ticker import MultipleLocator
import matplotlib
import matplotlib.pyplot as plt

plt.style.use("seaborn-v0_8-whitegrid")
matplotlib.use("TkAgg")


from simulation.plotting.plot_functions import plot_offnadir_distribution, plot_latency_distribution

def init_excel_log(path, header, sheet_name="Log"):
    if os.path.exists(path):
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws.append(header)
    wb.save(path)
    wb.close()
    return {"path": path, "sheet": sheet_name, "header": header}


def append_excel_log(writer, row):
    path = writer["path"]
    sheet = writer["sheet"]
    header = writer["header"]

    wb = load_workbook(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    if ws.max_row == 0:
        ws.append(header)
    ws.append(row)
    wb.save(path)
    wb.close()


def log_tip_observation(writer, target_id, tip_observation_date, tip_actor, offnadir_deg, gsd_m,
                      target_lat, target_lon, target_alt, tip_lat, tip_lon, tip_alt,
                      x, y, z, vx, vy, vz, tip_observation_counter):
    """Log a tip observation event matching header_tip."""

    if should_log_event(writer, target_id, tip_observation_date, min_gap_sec=1000):
        row = [target_id,
               tip_observation_date.isoformat(timespec="seconds") + "Z", tip_actor,
               offnadir_deg, gsd_m,
               target_lat, target_lon, target_alt,
               tip_lat, tip_lon, tip_alt,
               x, y, z, vx, vy, vz,
               tip_observation_counter]
        append_excel_log(writer, row)


def log_cue_observation(writer, target_id, cue_observation_date, cue_actor, offnadir_deg, gsd_m,
                      target_lat, target_lon, target_alt, cue_lat, cue_lon, cue_alt,
                      x, y, z, vx, vy, vz, roll, pitch, yaw, cue_observation_counter):
    """Log a cue observation event matching header_cue."""

    if should_log_event(writer, target_id, cue_observation_date, min_gap_sec=1000):
        row = [target_id,
               cue_observation_date.isoformat(timespec="seconds") + "Z", cue_actor,
               offnadir_deg, gsd_m,
               target_lat, target_lon, target_alt,
               cue_lat, cue_lon, cue_alt,
               x, y, z, vx, vy, vz,
               roll, pitch, yaw,
               cue_observation_counter]
        append_excel_log(writer, row)


def log_combined_observation(writer, target_id, tip_observation_date, tip_actor, cue_observation_date, cue_actor,
                           offnadir_deg, gsd_m, latency, target_lat, target_lon, target_alt,
                           cue_lat, cue_lon, cue_alt, tip_observation_counter, cue_observation_counter):
    """Log a combined tip+cue event matching header_combined."""

    if should_log_event(writer, target_id, cue_observation_date, min_gap_sec=1000):
        row = [target_id,
               tip_observation_date.isoformat(timespec="seconds") + "Z", tip_actor,
               cue_observation_date.isoformat(timespec="seconds") + "Z", cue_actor,
               offnadir_deg, gsd_m, latency,
               target_lat, target_lon, target_alt,
               cue_lat, cue_lon, cue_alt,
               tip_observation_counter, cue_observation_counter]
        append_excel_log(writer, row)



def gsd_offnadir(Pn_m, H_m, offnadir_deg):
    """
    Compute exact off-nadir ground sampling distance (GSD).

    Parameters
    ----------
    Pn_m : float
        Nadir ground sampling distance [m]
    H_m : float
        Sensor altitude above ground [m]
    offnadir_deg : float
        Off-nadir angle [deg]

    Returns
    -------
    Ptheta_m : float
        Off-nadir ground sampling distance [m]
    """

    theta = math.radians(offnadir_deg)

    # Step 1: pixel angular width beta from nadir GSD
    beta = 2.0 * math.atan(Pn_m / (2.0 * H_m))

    # Step 2: exact off-nadir GSD
    Ptheta_m = H_m * (math.tan(theta + beta/2.0) - math.tan(theta - beta/2.0))

    return Ptheta_m


def should_log_event(writer, whale_idx, t_datetime, min_gap_sec=600):
    import pandas as pd
    from datetime import timedelta

    path = writer["path"]
    sheet = writer["sheet"]

    if not os.path.exists(path):
        return True

    df = pd.read_excel(path, sheet_name=sheet)
    if df.empty:
        return True

    df_target = df[df["target_id"] == whale_idx]
    if df_target.empty:
        return True

    last_time = pd.to_datetime(df_target["date"].max())
    if (t_datetime - last_time) > timedelta(seconds=min_gap_sec):
        return True

    return False


def safe_move(src, dst, retries=5, delay=1.0):
    for i in range(retries):
        try:
            shutil.move(src, dst)
            print(f"Moved {src} to {dst.replace(os.sep, '/')}")
            return True
        except PermissionError:
            print(f"Retry {i+1}/{retries}: could not move {src}, still locked.")
            time.sleep(delay)
        except Exception as e:
            print(f"Error moving {src} to {dst}: {e}")
            return False
    print(f"Warning: could not move {src} after {retries} retries.")
    return False


def at_exit(save_name, pl=None, verbose=True):

    if verbose:
        print("Save files")
    results_dir = os.path.join("results", save_name)
    os.makedirs(results_dir, exist_ok=True)

    if pl is not None:
        try:
            pl.close()
            time.sleep(0.1)
            if verbose:
                print("Closed pyvista plotter")
        except Exception as e:
            if verbose:
                print(f"Could not close pyvista plotter: {e}")

    rename_map = {
        "sim_output.xlsx": f"results_{save_name}.xlsx",
        "simulation.mp4": f"mov_{save_name}.mov",
        "output.log": f"logs_{save_name}.log",
        f"footprints_tip.html": f"footprints_tip_{save_name}.html",
        f"footprints_cue.html": f"footprints_cue_{save_name}.html"
    }

    combined_excel = None
    for src, new_name in rename_map.items():
        if src == "output.log" and isinstance(sys.stdout, Logger):
            try:
                sys.stdout.close()
                sys.stdout = sys.__stdout__
                sys.stderr = sys.__stderr__
                time.sleep(0.1)
            except Exception as e:
                if verbose:
                    print(f"Could not close print logs: {e}")

        if os.path.exists(src):
            dst = os.path.join(results_dir, new_name)
            if safe_move(src, dst) and "results_" in new_name:
                combined_excel = dst
        else:
            if verbose:
                print(f"Warning: {src} not found, skipping.")

    time.sleep(0.1)

    if combined_excel and os.path.exists(combined_excel):
        plot_offnadir_distribution(combined_excel, results_dir, save_name, bin_size_deg=5)
        plot_latency_distribution(combined_excel, results_dir, save_name, bin_size_sec=30)
        if verbose:
            print("Created offnadir and latency distribution plots")

    print(f"Saved results in {results_dir.replace(os.sep, '/')}")

def compute_stats(series: pd.Series):
    """Return mean, min, max, std for a pandas Series (ignoring NaN)."""
    s = series.dropna()
    if s.empty:
        return float("nan"), float("nan"), float("nan"), float("nan")
    return s.mean(), s.min(), s.max(), s.std()

class Logger:
    def __init__(self, filename):
        self.terminal = sys.__stdout__
        self.log = open(filename, "a")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        self.terminal.flush()
        self.log.flush()

    def close(self):
        try:
            self.log.close()
        except:
            pass



