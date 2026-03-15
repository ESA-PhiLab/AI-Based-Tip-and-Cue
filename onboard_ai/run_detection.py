from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from detection_scripts import (
    compute_coco_ap_metrics,
    evaluate_predictions,
    extract_gt_boxes,
    filter_predictions,
    get_image_record,
    gt_details_to_rows,
    gt_sample_summary_rows,
    image_summary_row,
    load_annotations,
    load_detector,
    prediction_details_to_rows,
    run_loaded_detection_raw,
    save_detection_image,
    show_detections_for_image,
)


def _normalize_excel_value(value: Any) -> Any:
    """Convert complex values to Excel-safe values."""
    if isinstance(value, (list, dict)):
        return json.dumps(value)
    return value


def _write_sheet(workbook: Workbook, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    """Write one list of dict rows to one worksheet."""
    worksheet = workbook.create_sheet(title=sheet_name[:31])

    if not rows:
        worksheet.append(["no_data"])
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)

    for row in rows:
        worksheet.append([_normalize_excel_value(row.get(header)) for header in headers])

    for column_index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_index in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            cell_length = len(str(cell_value)) if cell_value is not None else 0
            if cell_length > max_length:
                max_length = cell_length
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)


def _save_workbook(xlsx_path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    """Write multiple sheet datasets into one XLSX workbook."""
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, rows in sheets.items():
        _write_sheet(workbook, sheet_name, rows)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(xlsx_path)


def _safe_mean(values: list[float]) -> float:
    """Return mean or 0.0 when empty."""
    return float(np.mean(values)) if values else 0.0


def _safe_std(values: list[float]) -> float:
    """Return std or 0.0 when empty."""
    return float(np.std(values)) if values else 0.0


def _compute_overall_detection_stats(
    image_summary_rows: list[dict[str, Any]],
    individual_prediction_rows: list[dict[str, Any]],
    gt_rows: list[dict[str, Any]],
    positive_sample_best_ious: list[float],
    positive_sample_best_confidences: list[float],
    all_sample_best_ious: list[float],
    all_sample_best_confidences: list[float],
) -> list[dict[str, Any]]:
    """Build aggregate detection statistics rows."""
    total_tp = int(sum(int(row["tp"]) for row in image_summary_rows))
    total_fp = int(sum(int(row["fp"]) for row in image_summary_rows))
    total_fn = int(sum(int(row["fn"]) for row in image_summary_rows))
    total_gt = int(sum(int(row["num_gt"]) for row in image_summary_rows))
    total_predictions = int(sum(int(row["num_predictions"]) for row in image_summary_rows))

    overall_precision = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 0.0
    overall_recall = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 0.0
    overall_f1 = 2.0 * overall_precision * overall_recall / (overall_precision + overall_recall) if (overall_precision + overall_recall) > 0 else 0.0

    tp_confidences = [float(row["score"]) for row in individual_prediction_rows if bool(row["is_tp"])]
    fp_confidences = [float(row["score"]) for row in individual_prediction_rows if bool(row["is_fp"])]
    all_confidences = [float(row["score"]) for row in individual_prediction_rows]
    tp_ious = [float(row["matched_iou"]) for row in individual_prediction_rows if bool(row["is_tp"]) and row["matched_iou"] is not None]
    gt_best_ious = [float(row["best_prediction_iou"]) for row in gt_rows]
    gt_best_confidences = [float(row["best_prediction_score"]) if row["best_prediction_score"] is not None else 0.0 for row in gt_rows]

    return [
        {"metric": "num_images_processed", "value": len(image_summary_rows)},
        {"metric": "num_positive_samples", "value": len(positive_sample_best_ious)},
        {"metric": "num_all_samples", "value": len(all_sample_best_ious)},
        {"metric": "total_gt_boxes", "value": total_gt},
        {"metric": "total_reported_predictions_thresholded", "value": total_predictions},
        {"metric": "total_tp", "value": total_tp},
        {"metric": "total_fp", "value": total_fp},
        {"metric": "total_fn", "value": total_fn},
        {"metric": "overall_precision", "value": overall_precision},
        {"metric": "overall_recall", "value": overall_recall},
        {"metric": "overall_f1", "value": overall_f1},
        {"metric": "avg_prediction_confidence_thresholded", "value": _safe_mean(all_confidences)},
        {"metric": "std_prediction_confidence_thresholded", "value": _safe_std(all_confidences)},
        {"metric": "avg_tp_confidence", "value": _safe_mean(tp_confidences)},
        {"metric": "std_tp_confidence", "value": _safe_std(tp_confidences)},
        {"metric": "avg_fp_confidence", "value": _safe_mean(fp_confidences)},
        {"metric": "std_fp_confidence", "value": _safe_std(fp_confidences)},
        {"metric": "avg_tp_iou", "value": _safe_mean(tp_ious)},
        {"metric": "std_tp_iou", "value": _safe_std(tp_ious)},
        {"metric": "avg_best_iou_per_gt_sample", "value": _safe_mean(gt_best_ious)},
        {"metric": "std_best_iou_per_gt_sample", "value": _safe_std(gt_best_ious)},
        {"metric": "avg_best_confidence_per_gt_sample", "value": _safe_mean(gt_best_confidences)},
        {"metric": "std_best_confidence_per_gt_sample", "value": _safe_std(gt_best_confidences)},
        {"metric": "avg_best_iou_positive_samples", "value": _safe_mean(positive_sample_best_ious)},
        {"metric": "std_best_iou_positive_samples", "value": _safe_std(positive_sample_best_ious)},
        {"metric": "avg_best_confidence_positive_samples", "value": _safe_mean(positive_sample_best_confidences)},
        {"metric": "std_best_confidence_positive_samples", "value": _safe_std(positive_sample_best_confidences)},
        {"metric": "avg_best_iou_all_samples_negative_zero", "value": _safe_mean(all_sample_best_ious)},
        {"metric": "std_best_iou_all_samples_negative_zero", "value": _safe_std(all_sample_best_ious)},
        {"metric": "avg_best_confidence_all_samples_negative_zero", "value": _safe_mean(all_sample_best_confidences)},
        {"metric": "std_best_confidence_all_samples_negative_zero", "value": _safe_std(all_sample_best_confidences)},
    ]


def main() -> None:
    """Run detections, save overlays, and write workbooks."""
    script_dir = Path(__file__).resolve().parent
    master_dir = script_dir.parent
    deimv2_repo_root = script_dir / "DEIMV2-main"

    run_name = "03_e32_flat14_noaug12_stop20_match18"
    device = "cpu"

    render_scale = 10
    line_width = 5

    results_folder_path = deimv2_repo_root / "results"
    image_folder_path = master_dir / "0_results" / "0_FINAL_RESULTS_data" / "TC_1x1sat_40deg_5min_17sd" / "satellite_images"

    best_stg_path = results_folder_path / run_name / "overview" / "best_stg2.pth"
    config_path = results_folder_path / run_name / f"{run_name}.yml"
    anns_path = image_folder_path / "annotations_postprocessed.json"

    max_images: int | None = None
    individual_score_threshold = 0.402
    individual_iou_threshold = 0.5
    ap_score_threshold = 0.0

    MODEL_LABEL_TO_CATEGORY_ID: dict[int, int] | None = None

    show_detections = False
    save_prediction_images = True
    reset_prediction_images = True

    run_output_root = image_folder_path.parent / "onboard_detection"
    prediction_images_dir = run_output_root / "prediction_images"
    results_xlsx_path = run_output_root / "onboard_detection_results.xlsx"
    per_sample_xlsx_path = run_output_root / "onboard_detection_per_sample.xlsx"

    detector = load_detector(
        pth=best_stg_path,
        config=config_path,
        device=device,
        repo_root=deimv2_repo_root,
        use_cache=True,
    )

    annotations_data = load_annotations(
        annotations=anns_path,
        use_cache=True,
    )

    annotation_category_ids = sorted(
        {
            int(ann["category_id"])
            for ann in annotations_data.get("annotations", [])
            if isinstance(ann, dict) and "category_id" in ann
        }
    )

    if not annotation_category_ids:
        raise ValueError("No annotation category ids were found in annotations_data['annotations'].")

    if len(annotation_category_ids) != 1:
        raise ValueError(
            f"Expected a single-class dataset, but found multiple annotation category ids: {annotation_category_ids}"
        )

    gt_category_id = int(annotation_category_ids[0])

    raw_categories = annotations_data.get("categories", [])
    raw_category_ids = sorted(
        int(cat["id"])
        for cat in raw_categories
        if isinstance(cat, dict) and "id" in cat
    )

    if raw_category_ids != [gt_category_id]:
        if len(raw_categories) == 1 and isinstance(raw_categories[0], dict):
            fixed_category = dict(raw_categories[0])
            fixed_category["id"] = gt_category_id
            annotations_data["categories"] = [fixed_category]
        else:
            annotations_data["categories"] = [
                {
                    "id": gt_category_id,
                    "name": "whale",
                    "supercategory": "whale",
                }
            ]

    MODEL_LABEL_TO_CATEGORY_ID = {0: gt_category_id}

    run_output_root.mkdir(parents=True, exist_ok=True)
    prediction_images_dir.mkdir(parents=True, exist_ok=True)

    if reset_prediction_images and prediction_images_dir.exists():
        for image_file in prediction_images_dir.iterdir():
            if image_file.is_file() and image_file.suffix.lower() in {".png", ".jpg", ".jpeg"}:
                image_file.unlink()

    image_paths = sorted(image_folder_path.glob("*.png"))
    if max_images is not None:
        image_paths = image_paths[:max_images]

    image_summary_rows: list[dict[str, Any]] = []
    individual_prediction_rows: list[dict[str, Any]] = []
    gt_rows: list[dict[str, Any]] = []
    gt_sample_rows: list[dict[str, Any]] = []
    ap_predictions_iou50_rows: list[dict[str, Any]] = []
    debug_model_output_rows: list[dict[str, Any]] = []

    ap_predictions_by_image: dict[str, list[dict[str, Any]]] = {}
    positive_sample_best_ious: list[float] = []
    positive_sample_best_confidences: list[float] = []
    all_sample_best_ious: list[float] = []
    all_sample_best_confidences: list[float] = []

    print(f"GT category ids in annotations: {annotation_category_ids}")
    print(f"Categories ids after normalization: {[int(cat['id']) for cat in annotations_data.get('categories', []) if isinstance(cat, dict) and 'id' in cat]}")
    print(f"MODEL_LABEL_TO_CATEGORY_ID: {MODEL_LABEL_TO_CATEGORY_ID}")

    for image_path in image_paths:
        image_record = get_image_record(annotations_data, image_path)
        image_id = int(image_record["id"])
        gt_boxes = extract_gt_boxes(
            annotations=annotations_data,
            image_path=image_path,
        )

        raw_predictions = run_loaded_detection_raw(
            detector=detector,
            image=image_path,
        )

        debug_model_output_rows.append(
            {
                "image": image_path.name,
                "image_id": image_id,
                "num_model_outputs_before_threshold": len(raw_predictions),
            }
        )

        individual_predictions = filter_predictions(
            predictions=raw_predictions,
            score_threshold=individual_score_threshold,
        )

        individual_scores = evaluate_predictions(
            predictions=individual_predictions,
            gt_boxes=gt_boxes,
            iou_threshold=individual_iou_threshold,
        )
        individual_scores["image"] = str(image_path)
        individual_scores["checkpoint"] = str(best_stg_path)
        individual_scores["config"] = str(config_path)
        individual_scores["device"] = device
        individual_scores["score_threshold"] = float(individual_score_threshold)
        individual_scores["iou_threshold"] = float(individual_iou_threshold)

        ap_predictions = filter_predictions(
            predictions=raw_predictions,
            score_threshold=ap_score_threshold,
        )
        ap_predictions_by_image[image_path.name] = ap_predictions

        ap_iou50_scores = evaluate_predictions(
            predictions=ap_predictions,
            gt_boxes=gt_boxes,
            iou_threshold=0.50,
        )

        image_summary_rows.append(
            image_summary_row(
                image_name=image_path.name,
                image_id=image_id,
                evaluation_scores=individual_scores,
                score_threshold=individual_score_threshold,
                iou_threshold=individual_iou_threshold,
            )
        )
        individual_prediction_rows.extend(
            prediction_details_to_rows(
                image_name=image_path.name,
                image_id=image_id,
                evaluation_scores=individual_scores,
            )
        )
        gt_rows.extend(
            gt_details_to_rows(
                image_name=image_path.name,
                image_id=image_id,
                evaluation_scores=individual_scores,
            )
        )
        gt_sample_rows.extend(
            gt_sample_summary_rows(
                image_name=image_path.name,
                image_id=image_id,
                evaluation_scores=individual_scores,
                score_threshold=individual_score_threshold,
                iou_threshold=individual_iou_threshold,
            )
        )
        ap_predictions_iou50_rows.extend(
            prediction_details_to_rows(
                image_name=image_path.name,
                image_id=image_id,
                evaluation_scores=ap_iou50_scores,
            )
        )

        if gt_boxes:
            gt_best_ious = [float(item["best_prediction_iou"]) for item in individual_scores["gt_details"]]
            gt_best_confidences = [float(item["best_prediction_score"]) if item["best_prediction_score"] is not None else 0.0 for item in individual_scores["gt_details"]]
            best_iou_for_sample = max(gt_best_ious) if gt_best_ious else 0.0
            best_conf_for_sample = max(gt_best_confidences) if gt_best_confidences else 0.0
            positive_sample_best_ious.append(best_iou_for_sample)
            positive_sample_best_confidences.append(best_conf_for_sample)
            all_sample_best_ious.append(best_iou_for_sample)
            all_sample_best_confidences.append(best_conf_for_sample)
        else:
            all_sample_best_ious.append(0.0)
            all_sample_best_confidences.append(0.0)

        if save_prediction_images:
            prediction_image_path = prediction_images_dir / image_path.name
            save_detection_image(
                image=image_path,
                gt_boxes=gt_boxes,
                prediction_details=individual_scores["prediction_details"],
                output_path=prediction_image_path,
                render_scale=render_scale,
                line_width=line_width,
            )

        if show_detections:
            title = (
                f"{image_path.name} | "
                f"TP={individual_scores['tp']} FP={individual_scores['fp']} FN={individual_scores['fn']} "
                f"| F1={individual_scores['f1']:.3f}"
            )
            show_detections_for_image(
                image=image_path,
                gt_boxes=gt_boxes,
                prediction_details=individual_scores["prediction_details"],
                title=title,
                render_scale=render_scale,
                line_width=line_width,
            )

        print(f"\nImage: {image_path.name}")
        print(f"Model outputs before threshold: {len(raw_predictions)}")
        print(f"Reported detections after threshold: {len(individual_predictions)}")
        print(
            f"TP={individual_scores['tp']} "
            f"FP={individual_scores['fp']} "
            f"FN={individual_scores['fn']} "
            f"Precision={individual_scores['precision']:.4f} "
            f"Recall={individual_scores['recall']:.4f} "
            f"F1={individual_scores['f1']:.4f} "
            f"mean_tp_iou={individual_scores['mean_tp_iou']:.4f} "
            f"best_tp_confidence={individual_scores['best_tp_confidence']:.4f}"
        )

    total_thresholded_predictions_individual = int(sum(int(row["num_predictions"]) for row in image_summary_rows))
    total_model_outputs_before_threshold = int(sum(int(row["num_model_outputs_before_threshold"]) for row in debug_model_output_rows))

    coco_metrics = compute_coco_ap_metrics(
        annotations_data=annotations_data,
        image_paths=image_paths,
        predictions_by_image=ap_predictions_by_image,
        label_to_category_id=MODEL_LABEL_TO_CATEGORY_ID,
    )

    ap_metrics_rows = coco_metrics["per_iou_rows"]
    ap50 = coco_metrics["ap50"]
    ap50_95 = coco_metrics["ap50_95"]
    total_gt_for_ap = coco_metrics["total_gt"]
    total_ap_predictions = coco_metrics["total_predictions"]
    gt_category_ids_for_ap = coco_metrics["gt_category_ids"]
    prediction_category_ids_for_ap = coco_metrics["prediction_category_ids"]

    overall_detection_stats_rows = _compute_overall_detection_stats(
        image_summary_rows=image_summary_rows,
        individual_prediction_rows=individual_prediction_rows,
        gt_rows=gt_rows,
        positive_sample_best_ious=positive_sample_best_ious,
        positive_sample_best_confidences=positive_sample_best_confidences,
        all_sample_best_ious=all_sample_best_ious,
        all_sample_best_confidences=all_sample_best_confidences,
    )

    run_summary_rows = [
        {
            "run_name": run_name,
            "num_images_processed": len(image_paths),
            "positive_sample_count": len(positive_sample_best_ious),
            "all_sample_count": len(all_sample_best_ious),
            "individual_score_threshold": float(individual_score_threshold),
            "individual_iou_threshold": float(individual_iou_threshold),
            "ap_score_threshold": float(ap_score_threshold),
            "annotation_category_ids": annotation_category_ids,
            "ap_gt_category_ids": gt_category_ids_for_ap,
            "ap_prediction_category_ids": prediction_category_ids_for_ap,
            "model_label_to_category_id": MODEL_LABEL_TO_CATEGORY_ID,
            "coco_ap50": ap50,
            "coco_ap50_95": ap50_95,
            "total_gt": total_gt_for_ap,
            "reported_predictions_individual_threshold": total_thresholded_predictions_individual,
            "predictions_used_for_coco_ap": total_ap_predictions,
            "debug_model_outputs_before_threshold_total": total_model_outputs_before_threshold,
            "avg_best_iou_positive_samples": _safe_mean(positive_sample_best_ious),
            "std_best_iou_positive_samples": _safe_std(positive_sample_best_ious),
            "avg_best_confidence_positive_samples": _safe_mean(positive_sample_best_confidences),
            "std_best_confidence_positive_samples": _safe_std(positive_sample_best_confidences),
            "avg_best_iou_all_samples_negative_zero": _safe_mean(all_sample_best_ious),
            "std_best_iou_all_samples_negative_zero": _safe_std(all_sample_best_ious),
            "avg_best_confidence_all_samples_negative_zero": _safe_mean(all_sample_best_confidences),
            "std_best_confidence_all_samples_negative_zero": _safe_std(all_sample_best_confidences),
            "prediction_images_dir": str(prediction_images_dir),
            "results_workbook": str(results_xlsx_path),
            "per_sample_workbook": str(per_sample_xlsx_path),
        }
    ]

    _save_workbook(
        results_xlsx_path,
        {
            "run_summary": run_summary_rows,
            "overall_detection_stats": overall_detection_stats_rows,
            "ap_metrics_by_iou": ap_metrics_rows,
            "image_summary": image_summary_rows,
            "individual_predictions": individual_prediction_rows,
            "gt_log": gt_rows,
            "AP_Predictions_IoU50": ap_predictions_iou50_rows,
            "debug_model_outputs": debug_model_output_rows,
        },
    )

    _save_workbook(
        per_sample_xlsx_path,
        {
            "run_summary": run_summary_rows,
            "gt_sample_summary": gt_sample_rows,
        },
    )

    print("\nOfficial COCO AP results:")
    print(f"AP50: {ap50:.6f}")
    print(f"AP50:95: {ap50_95:.6f}")
    print(f"Total GT: {total_gt_for_ap}")
    print(f"GT category ids used for AP: {gt_category_ids_for_ap}")
    print(f"Prediction category ids used for AP: {prediction_category_ids_for_ap}")
    print(f"Reported predictions (thresholded): {total_thresholded_predictions_individual}")
    print(f"Predictions used for COCO AP: {total_ap_predictions}")
    print(f"Debug model outputs before threshold: {total_model_outputs_before_threshold}")

    print("\nOutputs written to:")
    print(results_xlsx_path)
    print(per_sample_xlsx_path)
    print(prediction_images_dir)


if __name__ == "__main__":
    main()