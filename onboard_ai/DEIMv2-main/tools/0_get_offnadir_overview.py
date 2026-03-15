#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from PIL import Image, ImageDraw

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    from pycocotools.coco import COCO
    from pycocotools.cocoeval import COCOeval
except Exception:
    COCO = None
    COCOeval = None

# =======================
# DEFAULTS (override via CLI)
# =======================

DEFAULT_RESULTS_FOLDER = "03_e32_flat14_noaug12_stop20_match18"
DEFAULT_MODEL = "final"          # "final" or "fold1" etc
DEFAULT_SPLIT = "test"     # "validation" or "test"

SCORE_THRESHOLD = 0.3
MAX_PREDS = 50
IOU_THRESHOLD = 0.5

USE_OVERVIEW_PREDICTIONS = False
OVERVIEW_TAG = "FINAL"

COCO_AP_METRICS = [
    ("AP", 0),
    ("AP50", 1),
    ("AP75", 2),
]

COCO_AR_METRICS = [
    ("AR1", 6),
    ("AR10", 7),
    ("AR100", 8),
]

COCO_METRICS_TO_EXPORT = COCO_AP_METRICS + COCO_AR_METRICS

# =======================


def read_json(path: Path) -> Any:
    """read_json(path) -> Any: Read JSON file."""
    return json.loads(path.read_text(encoding="utf-8", errors="replace"))


def xywh_to_xyxy(b: List[float]) -> Tuple[float, float, float, float]:
    """xywh_to_xyxy(b) -> (x1,y1,x2,y2)."""
    x, y, w, h = [float(v) for v in b]
    return x, y, x + w, y + h


def iou(a: Tuple[float, float, float, float], b: Tuple[float, float, float, float]) -> float:
    """iou(a, b) -> float: IoU of two xyxy boxes."""
    x1 = max(a[0], b[0])
    y1 = max(a[1], b[1])
    x2 = min(a[2], b[2])
    y2 = min(a[3], b[3])
    iw = max(0.0, x2 - x1)
    ih = max(0.0, y2 - y1)
    inter = iw * ih
    area_a = max(0.0, a[2] - a[0]) * max(0.0, a[3] - a[1])
    area_b = max(0.0, b[2] - b[0]) * max(0.0, b[3] - b[1])
    union = area_a + area_b - inter
    return (inter / union) if union > 0 else 0.0


def draw_box(draw: ImageDraw.ImageDraw, box, color, width: int) -> None:
    """draw_box(draw, box, color, width) -> None."""
    x1, y1, x2, y2 = box
    for i in range(int(width)):
        draw.rectangle([x1 - i, y1 - i, x2 + i, y2 + i], outline=color)


def resolve_repo_root() -> Path:
    """resolve_repo_root() -> Path: Infer repo root as parent of tools/."""
    return Path(__file__).resolve().parents[1]


def resolve_results_dir(repo_root: Path, results_folder: str) -> Path:
    """resolve_results_dir(repo_root, results_folder) -> Path: results/<results_folder>."""
    p = repo_root / "results" / results_folder
    if not p.exists():
        raise RuntimeError(f"RESULTS_DIR not found: {p}")
    return p


def resolve_img_root(repo_root: Path) -> Path:
    """resolve_img_root(repo_root) -> Path: data/0_merged/... under repo root."""
    p = repo_root / "data" / "0_merged" / "reflection_offnadir_glint_255"
    if not p.exists():
        raise RuntimeError(f"IMG_ROOT not found: {p}")
    return p


def resolve_model_dir(results_dir: Path, model: str) -> Path:
    """resolve_model_dir(results_dir, model) -> Path."""
    if model.lower() == "final":
        p = results_dir / "final_location_holdout"
    else:
        p = results_dir / "cross_validation" / model
    if not p.exists():
        raise RuntimeError(f"model dir not found: {p}")
    return p


def resolve_ann_path(results_dir: Path, model_dir: Path, split: str) -> Path:
    """resolve_ann_path(results_dir, model_dir, split) -> Path."""
    if split == "validation":
        p = model_dir / "splits" / "instances_val.json"
        if p.exists():
            return p
        raise RuntimeError(f"validation annotations not found: {p}")

    if split == "test":
        p = results_dir / "test_holdout_only.json"
        if p.exists():
            return p
        raise RuntimeError(f"test annotations not found: {p}")

    raise RuntimeError("split must be validation or test")


def resolve_pred_path(results_dir: Path, model_dir: Path, split: str, model: str) -> Path:
    """resolve_pred_path(results_dir, model_dir, split, model) -> Path."""
    if USE_OVERVIEW_PREDICTIONS:
        p = results_dir / "overview" / split / "predictions" / f"{OVERVIEW_TAG}_predictions_coco.json"
        if p.exists():
            return p
        raise RuntimeError(f"overview predictions not found: {p}")

    p = model_dir / "metrics" / split / "predictions_coco.json"
    if p.exists():
        return p

    eval_dir = model_dir / ("eval_val" if split == "validation" else "eval_test") / "eval_data"
    p2 = eval_dir / "predictions_coco.json"
    if p2.exists():
        return p2

    p3 = results_dir / "overview" / split / "predictions" / f"{('FINAL' if model == 'final' else model)}_predictions_coco.json"
    if p3.exists():
        return p3

    raise RuntimeError("predictions file not found (checked metrics/, eval_data/, overview/)")


def figures_dir(repo_root: Path, results_folder: str, model: str, split: str) -> Path:
    """figures_dir(repo_root, results_folder, model, split) -> Path: figures/<results_folder>/<model>_<split>/."""
    base = repo_root / "figures"
    out = base / results_folder / f"{model}_{split}"
    out.mkdir(parents=True, exist_ok=True)
    return out


def overview_dir(repo_root: Path, results_folder: str, split: str) -> Path:
    """overview_dir(repo_root, results_folder, split) -> Path: figures/<results_folder>/overview_offnadir_<split>/."""
    out = repo_root / "figures" / results_folder / f"overview_offnadir_{split}"
    out.mkdir(parents=True, exist_ok=True)
    return out


def parse_offnadir_angle_deg(image_name: str) -> Optional[int]:
    """parse_offnadir_angle_deg(image_name) -> Optional[int]: Extract angle like '_15deg'."""
    m = re.search(r"_(\d+)\s*deg\b", image_name)
    if not m:
        m = re.search(r"_(\d+)deg\b", image_name)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def parse_fo_label(image_name: str) -> Optional[str]:
    """parse_fo_label(image_name) -> Optional[str]: Extract '_F' or '_O' token."""
    m = re.search(r"_(F|O)(?:\b|_|\.)", image_name)
    if not m:
        return None
    return m.group(1)


def autosize_ws_columns(ws) -> None:
    """autosize_ws_columns(ws) -> None: Auto-fit columns (approx)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))


def compute_avg_iou_per_image(
    gt_boxes: List[Tuple[float, float, float, float]],
    pred_boxes_scored: List[Tuple[float, Tuple[float, float, float, float]]],
) -> float:
    """compute_avg_iou_per_image(gt_boxes, pred_boxes_scored) -> float: Avg best-IoU per GT (0 if no GT)."""
    if not gt_boxes:
        return 0.0
    pred_boxes = [b for _, b in pred_boxes_scored]
    if not pred_boxes:
        return 0.0
    best_ious: List[float] = []
    for gb in gt_boxes:
        best = 0.0
        for pb in pred_boxes:
            best = max(best, iou(pb, gb))
        best_ious.append(best)
    return float(sum(best_ious) / max(1, len(best_ious)))


def filter_coco_dict_by_image_ids(coco: Dict[str, Any], image_ids: List[int]) -> Dict[str, Any]:
    """filter_coco_dict_by_image_ids(coco, image_ids) -> dict: Keep only images/annotations for given ids."""
    keep = set(int(i) for i in image_ids)
    images = [im for im in coco.get("images", []) if int(im.get("id")) in keep]
    annotations = [ann for ann in coco.get("annotations", []) if int(ann.get("image_id")) in keep]
    out = dict(coco)
    out["images"] = images
    out["annotations"] = annotations
    return out


def filter_preds_by_image_ids(preds: List[Dict[str, Any]], image_ids: List[int]) -> List[Dict[str, Any]]:
    """filter_preds_by_image_ids(preds, image_ids) -> list[dict]: Keep only preds for given ids."""
    keep = set(int(i) for i in image_ids)
    return [p for p in preds if int(p.get("image_id")) in keep]


def coco_eval_stats_for_subset(coco_gt_dict: Dict[str, Any], preds_list: List[Dict[str, Any]], image_ids: List[int]) -> Dict[str, float]:
    """coco_eval_stats_for_subset(coco_gt_dict, preds_list, image_ids) -> dict[str,float]: COCOeval stats on subset."""
    if COCO is None or COCOeval is None:
        raise RuntimeError("pycocotools is not available. Install it to compute AP/AR vs off-nadir angle.")
    if not image_ids:
        return {}

    gt_sub = filter_coco_dict_by_image_ids(coco_gt_dict, image_ids)
    preds_sub = filter_preds_by_image_ids(preds_list, image_ids)

    if not gt_sub.get("images") or not gt_sub.get("annotations"):
        return {}

    coco_gt = COCO()
    coco_gt.dataset = gt_sub
    coco_gt.createIndex()

    coco_dt = coco_gt.loadRes(preds_sub) if preds_sub else coco_gt.loadRes([])

    ev = COCOeval(coco_gt, coco_dt, iouType="bbox")
    ev.params.imgIds = [int(i) for i in image_ids]
    ev.evaluate()
    ev.accumulate()
    ev.summarize()

    stats = ev.stats.tolist() if hasattr(ev.stats, "tolist") else list(ev.stats)
    out: Dict[str, float] = {}
    for name, idx in COCO_METRICS_TO_EXPORT:
        v = float(stats[idx]) if idx < len(stats) else float("nan")
        out[name] = v
    return out


def plot_iou_vs_angle(out_path: Path, series: Dict[str, Tuple[List[int], List[float]]], title: str) -> None:
    """plot_iou_vs_angle(out_path, series, title) -> None: Plot IoU vs angle."""
    plt.figure()
    for label, (angles, vals) in series.items():
        plt.plot(angles, vals, marker="o", label=label)
    plt.xlabel("Off-nadir angle (deg)")
    plt.ylabel("Average IoU")
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_path, dpi=200)
    plt.close()


def plot_score_vs_angle(out_path: Path, angles: List[int], curves: Dict[str, List[float]], title: str, ylabel: str) -> None:
    """plot_score_vs_angle(out_path, angles, curves, title, ylabel) -> None: Plot multiple metrics vs angle."""
    plt.figure()
    for metric_name, vals in curves.items():
        plt.plot(angles, vals, marker="o", label=metric_name)
    plt.xlabel("Off-nadir angle (deg)")
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_path, dpi=200)
    plt.close()


def metrics_to_curves(stats_by_angle: Dict[int, Dict[str, float]],
                      angles_sorted: List[int],
                      metrics: List[Tuple[str, int]]) -> Dict[str, List[float]]:
    """metrics_to_curves(stats_by_angle, angles_sorted, metrics) -> dict[str,list[float]]: Curves for selected metrics."""
    curves: Dict[str, List[float]] = {}
    for metric_name, _ in metrics:
        curves[metric_name] = [float(stats_by_angle.get(a, {}).get(metric_name, float("nan"))) for a in angles_sorted]
    return curves


def plot_ap_and_ar(out_dir: Path, split: str, subset: str, stats_by_angle: Dict[int, Dict[str, float]]) -> None:
    """plot_ap_and_ar(out_dir, split, subset, stats_by_angle) -> None: Save separate AP and AR plots."""
    angs = sorted(stats_by_angle.keys())
    if not angs:
        return

    ap_curves = metrics_to_curves(stats_by_angle, angs, COCO_AP_METRICS)
    ar_curves = metrics_to_curves(stats_by_angle, angs, COCO_AR_METRICS)

    plot_score_vs_angle(
        out_dir / f"ap_vs_angle_{split}_{subset}.png",
        angs,
        ap_curves,
        title=f"COCO AP vs off-nadir angle ({subset}) ({split})",
        ylabel="AP",
    )

    plot_score_vs_angle(
        out_dir / f"ar_vs_angle_{split}_{subset}.png",
        angs,
        ar_curves,
        title=f"COCO AR vs off-nadir angle ({subset}) ({split})",
        ylabel="AR",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--results", type=str, default=DEFAULT_RESULTS_FOLDER, help="results folder, e.g. default_model_S")
    parser.add_argument("--model", type=str, default=DEFAULT_MODEL, help='model dir: "final" or "fold1" etc')
    parser.add_argument("--split", type=str, default=DEFAULT_SPLIT, choices=["validation", "test"], help="split")
    args = parser.parse_args()

    repo_root = resolve_repo_root()
    results_dir = resolve_results_dir(repo_root, args.results)
    img_root = resolve_img_root(repo_root)

    model_dir = resolve_model_dir(results_dir, args.model)
    ann_path = resolve_ann_path(results_dir, model_dir, args.split)
    pred_path = resolve_pred_path(results_dir, model_dir, args.split, args.model)

    coco = read_json(ann_path)
    preds = read_json(pred_path)
    if not isinstance(preds, list):
        raise RuntimeError(f"predictions must be a list, got {type(preds)} from {pred_path}")

    images = coco.get("images", [])
    annotations = coco.get("annotations", [])
    if not images:
        raise RuntimeError(f'No "images" in annotations file: {ann_path}')

    out_images_dir = figures_dir(repo_root, args.results, args.model, args.split)
    out_overview_dir = overview_dir(repo_root, args.results, args.split)

    print("ANN :", ann_path)
    print("PRED:", pred_path)
    print("OUT_IMAGES :", out_images_dir)
    print("OUT_OVERVIEW:", out_overview_dir)
    print(f"Processing {len(images)} images...")

    ann_by_img: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for a in annotations:
        ann_by_img[int(a.get("image_id"))].append(a)

    pred_by_img: Dict[int, List[Tuple[float, Tuple[float, float, float, float]]]] = defaultdict(list)
    for p in preds:
        try:
            image_id = int(p.get("image_id"))
            sc = float(p.get("score", 0.0))
            if sc >= float(SCORE_THRESHOLD):
                pred_by_img[image_id].append((sc, xywh_to_xyxy(p["bbox"])))
        except Exception:
            continue

    for image_id in list(pred_by_img.keys()):
        pred_by_img[image_id].sort(key=lambda t: t[0], reverse=True)
        pred_by_img[image_id] = pred_by_img[image_id][: int(MAX_PREDS)]

    per_image_rows: List[Dict[str, Any]] = []

    angle_to_img_ids_all: Dict[int, List[int]] = defaultdict(list)
    angle_to_img_ids_F: Dict[int, List[int]] = defaultdict(list)
    angle_to_img_ids_O: Dict[int, List[int]] = defaultdict(list)

    avg_iou_by_imgid: Dict[int, float] = {}

    # ---- 1) Draw images AND compute per-image IoU
    for idx, im_info in enumerate(images):
        image_id = int(im_info["id"])
        file_name = str(im_info["file_name"])
        img_path = img_root / file_name

        original_name = Path(file_name).name
        angle_deg = parse_offnadir_angle_deg(original_name)
        fo_label = parse_fo_label(original_name)

        if angle_deg is not None:
            angle_to_img_ids_all[angle_deg].append(image_id)
            if fo_label == "F":
                angle_to_img_ids_F[angle_deg].append(image_id)
            elif fo_label == "O":
                angle_to_img_ids_O[angle_deg].append(image_id)

        if not img_path.exists():
            print(f"[{idx+1}/{len(images)}] missing image file: {img_path}")
            continue

        gt_boxes = [xywh_to_xyxy(a["bbox"]) for a in ann_by_img.get(image_id, [])]
        pred_boxes_scored = pred_by_img.get(image_id, [])

        avg_iou_img = compute_avg_iou_per_image(gt_boxes, pred_boxes_scored)
        avg_iou_by_imgid[image_id] = avg_iou_img

        im = Image.open(img_path).convert("RGB")
        draw = ImageDraw.Draw(im)

        for gb in gt_boxes:
            draw_box(draw, gb, "green", 4)

        matched_gt = set()
        for sc, pb in pred_boxes_scored:
            best_i = -1
            best = 0.0
            for i, gb in enumerate(gt_boxes):
                if i in matched_gt:
                    continue
                v = iou(pb, gb)
                if v > best:
                    best = v
                    best_i = i

            if best >= float(IOU_THRESHOLD) and best_i >= 0:
                matched_gt.add(best_i)
                draw_box(draw, pb, "blue", 2)
            else:
                draw_box(draw, pb, "red", 2)

        out_path = out_images_dir / original_name
        im.save(out_path)

        per_image_rows.append(
            {
                "image_name": original_name,
                "off_nadir_deg": angle_deg if angle_deg is not None else "",
                "label_FO": fo_label if fo_label is not None else "",
                "avg_iou": float(avg_iou_img),
            }
        )

        print(f"[{idx+1}/{len(images)}] saved: {out_path.name}")

    # ---- 2) Aggregate IoU per angle for combined/F/O
    def avg_iou_per_angle(angle_to_ids: Dict[int, List[int]]) -> Tuple[List[int], List[float], Dict[int, float]]:
        angles_sorted = sorted(angle_to_ids.keys())
        vals: List[float] = []
        m: Dict[int, float] = {}
        for ang in angles_sorted:
            ids = angle_to_ids[ang]
            if not ids:
                vals.append(float("nan"))
                m[ang] = float("nan")
                continue
            v = sum(avg_iou_by_imgid.get(i, 0.0) for i in ids) / float(len(ids))
            vals.append(float(v))
            m[ang] = float(v)
        return angles_sorted, vals, m

    angles_all, iou_all, iou_all_map = avg_iou_per_angle(angle_to_img_ids_all)
    angles_F, iou_F, iou_F_map = avg_iou_per_angle(angle_to_img_ids_F)
    angles_O, iou_O, iou_O_map = avg_iou_per_angle(angle_to_img_ids_O)

    # ---- 3) COCO AP/AR metrics per angle for combined/F/O
    ap_by_angle_all: Dict[int, Dict[str, float]] = {}
    ap_by_angle_F: Dict[int, Dict[str, float]] = {}
    ap_by_angle_O: Dict[int, Dict[str, float]] = {}

    def eval_per_angle(angle_to_ids: Dict[int, List[int]]) -> Dict[int, Dict[str, float]]:
        out: Dict[int, Dict[str, float]] = {}
        for ang in sorted(angle_to_ids.keys()):
            ids = angle_to_ids[ang]
            if not ids:
                continue
            out[ang] = coco_eval_stats_for_subset(coco, preds, ids)
        return out

    try:
        if angle_to_img_ids_all:
            ap_by_angle_all = eval_per_angle(angle_to_img_ids_all)
        if angle_to_img_ids_F:
            ap_by_angle_F = eval_per_angle(angle_to_img_ids_F)
        if angle_to_img_ids_O:
            ap_by_angle_O = eval_per_angle(angle_to_img_ids_O)
    except Exception as e:
        print(f"WARNING: AP/AR-vs-angle computation skipped: {e}")
        ap_by_angle_all = {}
        ap_by_angle_F = {}
        ap_by_angle_O = {}

    # ---- 4) Write Excel (per-image + per-angle)
    xlsx_path = out_overview_dir / f"avg_iou_per_image_{args.split}.xlsx"
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "per_image"
    ws1.append(["image_name", "off_nadir_deg", "label_FO", "avg_iou"])
    for r in per_image_rows:
        ws1.append([r["image_name"], r["off_nadir_deg"], r["label_FO"], float(r["avg_iou"])])
    autosize_ws_columns(ws1)

    ws2 = wb.create_sheet("per_angle_iou")
    ws2.append(["off_nadir_deg", "avg_iou_all", "avg_iou_F", "avg_iou_O", "n_all", "n_F", "n_O"])
    all_angles_union = sorted(set(angle_to_img_ids_all.keys()) | set(angle_to_img_ids_F.keys()) | set(angle_to_img_ids_O.keys()))
    for ang in all_angles_union:
        ws2.append([
            ang,
            float(iou_all_map.get(ang, float("nan"))) if ang in iou_all_map else "",
            float(iou_F_map.get(ang, float("nan"))) if ang in iou_F_map else "",
            float(iou_O_map.get(ang, float("nan"))) if ang in iou_O_map else "",
            len(angle_to_img_ids_all.get(ang, [])),
            len(angle_to_img_ids_F.get(ang, [])),
            len(angle_to_img_ids_O.get(ang, [])),
        ])
    autosize_ws_columns(ws2)

    ws3 = wb.create_sheet("per_angle_ap_ar")
    header = ["off_nadir_deg"]
    for subset in ["all", "F", "O"]:
        for metric_name, _ in COCO_METRICS_TO_EXPORT:
            header.append(f"{subset}_{metric_name}")
    ws3.append(header)
    for ang in all_angles_union:
        row = [ang]
        stats_all = ap_by_angle_all.get(ang, {})
        stats_F = ap_by_angle_F.get(ang, {})
        stats_O = ap_by_angle_O.get(ang, {})
        for metric_name, _ in COCO_METRICS_TO_EXPORT:
            row.append(float(stats_all.get(metric_name, float("nan"))) if stats_all else "")
        for metric_name, _ in COCO_METRICS_TO_EXPORT:
            row.append(float(stats_F.get(metric_name, float("nan"))) if stats_F else "")
        for metric_name, _ in COCO_METRICS_TO_EXPORT:
            row.append(float(stats_O.get(metric_name, float("nan"))) if stats_O else "")
        ws3.append(row)
    autosize_ws_columns(ws3)

    wb.save(xlsx_path)
    print(f"Saved Excel: {xlsx_path}")

    # ---- 5) Plots (IoU + AP + AR), saved in overview folder
    plot_iou_vs_angle(
        out_overview_dir / f"iou_vs_angle_{args.split}_combined_vs_FO.png",
        {"all": (angles_all, iou_all), "F": (angles_F, iou_F), "O": (angles_O, iou_O)},
        title=f"Average IoU vs off-nadir angle ({args.split})",
    )

    if angles_all:
        plot_iou_vs_angle(
            out_overview_dir / f"iou_vs_angle_{args.split}_all.png",
            {"all": (angles_all, iou_all)},
            title=f"Average IoU vs off-nadir angle (all) ({args.split})",
        )
    if angles_F:
        plot_iou_vs_angle(
            out_overview_dir / f"iou_vs_angle_{args.split}_F.png",
            {"F": (angles_F, iou_F)},
            title=f"Average IoU vs off-nadir angle (whale F) ({args.split})",
        )
    if angles_O:
        plot_iou_vs_angle(
            out_overview_dir / f"iou_vs_angle_{args.split}_O.png",
            {"O": (angles_O, iou_O)},
            title=f"Average IoU vs off-nadir angle (ocean O) ({args.split})",
        )

    if ap_by_angle_all:
        plot_ap_and_ar(out_overview_dir, args.split, "all", ap_by_angle_all)
    if ap_by_angle_F:
        plot_ap_and_ar(out_overview_dir, args.split, "F", ap_by_angle_F)
    if ap_by_angle_O:
        plot_ap_and_ar(out_overview_dir, args.split, "O", ap_by_angle_O)

    # ---- 6) CSV overview (one column per subset)
    csv_path = out_overview_dir / f"overview_offnadir_{args.split}.csv"
    ap_metric_names = [m for m, _ in COCO_AP_METRICS]
    ar_metric_names = [m for m, _ in COCO_AR_METRICS]

    cols = ["off_nadir_deg", "n_all", "n_F", "n_O", "iou_all", "iou_F", "iou_O"]
    for mn in ap_metric_names:
        cols.append(f"all_{mn}")
    for mn in ar_metric_names:
        cols.append(f"all_{mn}")
    for mn in ap_metric_names:
        cols.append(f"F_{mn}")
    for mn in ar_metric_names:
        cols.append(f"F_{mn}")
    for mn in ap_metric_names:
        cols.append(f"O_{mn}")
    for mn in ar_metric_names:
        cols.append(f"O_{mn}")

    lines: List[str] = [",".join(cols)]

    def fmt(v: Any) -> str:
        if v is None:
            return ""
        try:
            if isinstance(v, str):
                return v
            if math.isnan(float(v)):
                return ""
            return f"{float(v):.6f}"
        except Exception:
            return ""

    for ang in all_angles_union:
        n_all = len(angle_to_img_ids_all.get(ang, []))
        n_F = len(angle_to_img_ids_F.get(ang, []))
        n_O = len(angle_to_img_ids_O.get(ang, []))

        row: List[str] = [
            str(ang),
            str(n_all),
            str(n_F),
            str(n_O),
            fmt(iou_all_map.get(ang)),
            fmt(iou_F_map.get(ang)),
            fmt(iou_O_map.get(ang)),
        ]

        s_all = ap_by_angle_all.get(ang, {})
        s_F = ap_by_angle_F.get(ang, {})
        s_O = ap_by_angle_O.get(ang, {})

        for mn in ap_metric_names:
            row.append(fmt(s_all.get(mn)) if s_all else "")
        for mn in ar_metric_names:
            row.append(fmt(s_all.get(mn)) if s_all else "")

        for mn in ap_metric_names:
            row.append(fmt(s_F.get(mn)) if s_F else "")
        for mn in ar_metric_names:
            row.append(fmt(s_F.get(mn)) if s_F else "")

        for mn in ap_metric_names:
            row.append(fmt(s_O.get(mn)) if s_O else "")
        for mn in ar_metric_names:
            row.append(fmt(s_O.get(mn)) if s_O else "")

        lines.append(",".join(row))

    csv_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Saved CSV: {csv_path}")

    print("Done.")


if __name__ == "__main__":
    main()