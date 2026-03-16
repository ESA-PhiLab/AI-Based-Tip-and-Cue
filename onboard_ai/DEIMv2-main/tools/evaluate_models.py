#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import eval_utils
import shutil

def _write_overview_mean_std_xlsx(path: Path, rows: list[dict[str, Any]], keys: list[str]) -> None:
    """_write_overview_mean_std_xlsx(path, rows, keys) -> None: mean/std xlsx (computed over provided rows)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "mean_std"
    ws.append(["metric", "mean", "std"])

    for k in keys:
        vals: list[float] = []
        for r in rows:
            v = eval_utils._safe_float(r.get(k))
            if v is not None:
                vals.append(float(v))
        if not vals:
            ws.append([k, "", ""])
            continue
        mean = sum(vals) / float(len(vals))
        var = sum((x - mean) ** 2 for x in vals) / float(len(vals))
        std = math.sqrt(var)
        ws.append([k, mean, std])

    ws.column_dimensions[get_column_letter(1)].width = 60
    ws.column_dimensions[get_column_letter(2)].width = 18
    ws.column_dimensions[get_column_letter(3)].width = 18
    wb.save(str(path))

def _write_fold_split_overview_xlsx(path: Path, folds: list[Path]) -> None:
    """_write_fold_split_overview_xlsx(path, folds) -> None: Write fold -> train/val/test locations overview xlsx."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "fold_splits"

    header = ["fold", "train_locations", "val_locations", "test_locations"]
    ws.append(header)

    for fold_dir in folds:
        meta_path = fold_dir / "fold_meta.json"
        if not meta_path.exists():
            ws.append([fold_dir.name, "", "", ""])
            continue

        try:
            meta = read_json(meta_path)
        except Exception:
            ws.append([fold_dir.name, "", "", ""])
            continue

        train_locs = meta.get("train_locations", []) or []
        val_locs = meta.get("val_locations", []) or []
        holdout_test_locs = meta.get("holdout_test_locations", []) or []

        ws.append(
            [
                fold_dir.name,
                ", ".join(map(str, train_locs)),
                ", ".join(map(str, val_locs)),
                ", ".join(map(str, holdout_test_locs)),
            ]
        )

    for ci, c in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(80, max(14, len(str(c)) + 2))

    wb.save(str(path))

def read_json(path: Path) -> Any:
    """read_json(path) -> Any: Read JSON."""
    return json.loads(path.read_text(encoding="utf-8", errors="replace"))


def write_json(path: Path, obj: Any) -> None:
    """write_json(path, obj) -> None: Write JSON with mkdir."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")


def safe_int(s: Any, default: int = 0) -> int:
    """safe_int(s, default=0) -> int: Parse int safely."""
    try:
        return int(str(s).strip())
    except Exception:
        return int(default)


def find_folds(results_dir: Path) -> list[Path]:
    """find_folds(results_dir) -> list[Path]: Find fold dirs under cross_validation."""
    cv = results_dir / "cross_validation"
    if not cv.exists():
        return []
    folds = [p for p in cv.iterdir() if p.is_dir() and p.name.startswith("fold")]

    def key(p: Path) -> int:
        m = re.match(r"fold(\d+)$", p.name)
        return int(m.group(1)) if m else 10**9

    return sorted(folds, key=key)


def resolve_train_log(fold_dir: Path) -> Path | None:
    """resolve_train_log(fold_dir) -> Path|None: Find the training stdout log for a fold."""
    cand = [
        fold_dir / "logs" / "train_stdout.log",
        fold_dir / "logs" / "train.log",
        fold_dir / "train_stdout.log",
        fold_dir / "train.log",
        # launcher logs (main_local.sh)
        fold_dir.parent / "final_train_stdout.log" if fold_dir.name == "final_location_holdout" else None,
        fold_dir.parent / "cv_train_stdout.log" if fold_dir.name.startswith("fold") else None,
    ]
    for p in cand:
        if p is None:
            continue
        if p.exists():
            return p
    return None


def _plot_all_ap_ar_from_coco_rows(fold_dir: Path, coco_rows: list[dict[str, Any]], run_name: str) -> None:
    """_plot_all_ap_ar_from_coco_rows(fold_dir, coco_rows, run_name) -> None: Plot all AP/AR metrics per epoch."""
    if not coco_rows:
        return

    keys: list[str] = []
    keys_seen: set[str] = set()
    for r in coco_rows:
        for k in r.keys():
            if k not in keys_seen:
                keys_seen.add(k)
                keys.append(k)

    ap_keys = [k for k in keys if k.startswith("AP_") or k.startswith("AP_precision")]
    ar_keys = [k for k in keys if k.startswith("AR_") or k.startswith("AR_recall")]

    plots_dir = fold_dir / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    # CLEANUP old AP/AR epoch plots
    eval_utils._cleanup_epoch_plots(
        plots_dir,
        prefixes=[
            "AP_",
            "AR_",
            "AP_precision",
            "AR_recall",
        ],
    )

    fold_tag = "final" if fold_dir.name == "final_location_holdout" else fold_dir.name
    rn = eval_utils.safe_slug(run_name) if run_name else "run"

    fold_tag = "final" if fold_dir.name == "final_location_holdout" else fold_dir.name
    rn = eval_utils.safe_slug(run_name) if run_name else "run"

    for k in ap_keys + ar_keys:
        out = plots_dir / f"{eval_utils.safe_slug(k)}_over_epoch_{rn}_{fold_tag}.png"
        eval_utils.plot_metric_over_epoch(out, f"{fold_dir.name}: {k} over epoch", coco_rows, k)




def _plot_train_val_loss_from_log(fold_dir: Path, loss_rows: list[dict[str, Any]], run_name: str) -> None:
    """_plot_train_val_loss_from_log(fold_dir, loss_rows, run_name) -> None: Save train/val loss curve plot."""
    if not loss_rows:
        return
    plots_dir = fold_dir / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)
    metrics_dir = fold_dir / "metrics"
    metrics_dir.mkdir(parents=True, exist_ok=True)

    eval_utils._cleanup_epoch_plots(
        plots_dir,
        prefixes=[
            "train_vs_val_loss_over_epoch",
            "train_vs_val_loss_shared_over_epoch",
            "train_real_vs_val_shared_loss_over_epoch",
            "train_minus_val_loss_over_epoch",
            "train_val_components_over_epoch",
        ],
    )

    eval_utils._cleanup_epoch_plots(
        metrics_dir,
        prefixes=[
            "loss_curves",
            "loss_curves_shared",
            "loss_curves_real_vs_val",
            "loss_gap",
            "loss_components",
        ],
    )

    fold_tag = "final" if fold_dir.name == "final_location_holdout" else fold_dir.name
    rn = eval_utils.safe_slug(run_name) if run_name else "run"

    # (1) As-is: 3 curves
    eval_utils.plot_train_val_loss(
        plots_dir / f"train_vs_val_loss_over_epoch_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train/val loss vs epoch",
        loss_rows,
        mode="all",
    )
    eval_utils.plot_train_val_loss(
        metrics_dir / f"loss_curves_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train/val loss vs epoch",
        loss_rows,
        mode="all",
    )

    # (2) Only train+val shared
    eval_utils.plot_train_val_loss(
        plots_dir / f"train_vs_val_loss_shared_over_epoch_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train/val shared loss vs epoch",
        loss_rows,
        mode="shared",
    )
    eval_utils.plot_train_val_loss(
        metrics_dir / f"loss_curves_shared_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train/val shared loss vs epoch",
        loss_rows,
        mode="shared",
    )

    # (3) Train real vs val shared
    eval_utils.plot_train_val_loss(
        plots_dir / f"train_real_vs_val_shared_loss_over_epoch_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train real vs val shared loss vs epoch",
        loss_rows,
        mode="real_vs_val",
    )
    eval_utils.plot_train_val_loss(
        metrics_dir / f"loss_curves_real_vs_val_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train real vs val shared loss vs epoch",
        loss_rows,
        mode="real_vs_val",
    )

    eval_utils.plot_train_val_loss_gap(
        plots_dir / f"train_minus_val_loss_over_epoch_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train - val loss vs epoch",
        loss_rows,
    )
    eval_utils.plot_train_val_loss_gap(
        metrics_dir / f"loss_gap_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train - val loss vs epoch",
        loss_rows,
    )

    # NEW: component plot (orange/blue/green + dashed val)
    eval_utils.plot_train_val_components_from_rows(
        plots_dir / f"train_val_components_over_epoch_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train/val component losses vs epoch",
        loss_rows,
    )
    eval_utils.plot_train_val_components_from_rows(
        metrics_dir / f"loss_components_{rn}_{fold_tag}.png",
        f"{fold_dir.name}: train/val component losses vs epoch",
        loss_rows,
    )


def _write_fold_train_outputs(fold_dir: Path, run_name: str) -> dict[str, Any]:
    """_write_fold_train_outputs(fold_dir, run_name) -> dict[str,Any]: Parse training logs to produce XLSX + plots."""
    log_path = resolve_train_log(fold_dir)
    if log_path is None:
        return {"ok": False, "reason": "missing_train_log"}

    coco_rows = eval_utils.parse_train_coco_metrics_from_log(log_path)
    loss_rows = eval_utils.parse_loss_curves_from_log(log_path)

    _plot_all_ap_ar_from_coco_rows(fold_dir, coco_rows, run_name)

    if coco_rows:
        eval_utils.write_train_metrics_xlsx(fold_dir, coco_rows)
    if loss_rows:
        eval_utils.write_loss_xlsx(fold_dir, loss_rows)

    _plot_train_val_loss_from_log(fold_dir, loss_rows, run_name)

    return {
        "ok": True,
        "num_coco_rows": int(len(coco_rows)),
        "num_loss_rows": int(len(loss_rows)),
        "train_log": str(log_path),
    }

def _read_extra_metrics(src_dir: Path) -> dict[str, Any]:
    """_read_extra_metrics(src_dir) -> dict[str,Any]: Read metrics_extra.json if present and flatten keys."""
    p = src_dir / "metrics_extra.json"
    if not p.exists():
        return {}
    try:
        j = eval_utils.read_json(p)
    except Exception:
        return {}

    out: dict[str, Any] = {}
    if not isinstance(j, dict):
        return out

    # Confusion matrix
    if isinstance(j.get("image_level_confusion"), dict):
        cm = j["image_level_confusion"]
        out["image_level_TP"] = eval_utils._safe_float(cm.get("TP"))
        out["image_level_FP"] = eval_utils._safe_float(cm.get("FP"))
        out["image_level_FN"] = eval_utils._safe_float(cm.get("FN"))
        out["image_level_TN"] = eval_utils._safe_float(cm.get("TN"))

    # Existing extra metrics
    for k in [
        "image_level_f1",
        "image_level_precision",
        "image_level_recall",
        "pr_ap_global",
        "roc_auc_image_level",
    ]:
        if k in j:
            out[k] = eval_utils._safe_float(j.get(k))

    # ---- NEW: expose the image-level score threshold used as an overview column ----
    # Your extra_detection_metrics.py writes "score_thr_used" (float). We map it into a key that
    # passes _keep_root_key() (starts with "image_level_").
    if "score_thr_used" in j:
        out["image_level_score_thr_used"] = eval_utils._safe_float(j.get("score_thr_used"))

    # Optional: if present, also keep the F1-optimum threshold and stats (these are already numeric)
    # and will appear as columns.
    for src_key, dst_key in [
        ("best_thr_f1", "image_level_best_thr_f1"),
        ("best_f1", "image_level_best_f1"),
        ("best_precision", "image_level_best_precision"),
        ("best_recall", "image_level_best_recall"),
    ]:
        if src_key in j:
            out[dst_key] = eval_utils._safe_float(j.get(src_key))

    return out


def _collect_eval_metrics_for_fold(fold_dir: Path, split: str, eval_name: str) -> dict[str, Any]:
    """_collect_eval_metrics_for_fold(fold_dir, split, eval_name) -> dict[str,Any]: Read eval metrics (+extra) if present."""
    src_dir = fold_dir / "metrics" / split
    if not src_dir.exists():
        if split == "validation":
            src_dir = fold_dir / "eval_val" / eval_name
        else:
            src_dir = fold_dir / "eval_test" / eval_name

    if not src_dir.exists():
        return {}

    m = eval_utils.collect_eval_metrics(src_dir)
    m.update(_read_extra_metrics(src_dir))
    return m


def _collect_eval_metrics_for_final(results_dir: Path, split: str, eval_name: str) -> dict[str, Any]:
    """_collect_eval_metrics_for_final(results_dir, split, eval_name) -> dict[str,Any]: Read final metrics (+extra) if present."""
    final_dir = results_dir / "final_location_holdout"
    if not final_dir.exists():
        return {}

    src_dir = final_dir / "metrics" / split
    if not src_dir.exists():
        if split == "validation":
            src_dir = final_dir / "eval_val" / eval_name
        else:
            src_dir = final_dir / "eval_test" / eval_name

    if not src_dir.exists():
        return {}

    m = eval_utils.collect_eval_metrics(src_dir)
    m.update(_read_extra_metrics(src_dir))
    return m


def _mean_row(rows: list[dict[str, Any]], keys: list[str]) -> dict[str, Any]:
    """_mean_row(rows, keys) -> dict[str,Any]: Compute MEAN row over provided rows."""
    mean_row: dict[str, Any] = {"fold": "MEAN"}
    for k in keys:
        vals: list[float] = []
        for r in rows:
            v = eval_utils._safe_float(r.get(k))
            if v is not None and not math.isnan(float(v)):
                vals.append(float(v))
        mean_row[k] = (sum(vals) / float(len(vals))) if vals else ""
    return mean_row


def _write_overview_xlsx_with_final(path: Path, fold_rows: list[dict[str, Any]], keys: list[str], final_row: dict[str, Any] | None) -> None:
    """_write_overview_xlsx_with_final(path, fold_rows, keys, final_row) -> None: Write folds + MEAN + blank + FINAL (FINAL excluded from mean)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "overview"

    header = ["fold"] + keys
    ws.append(header)

    for r in fold_rows:
        ws.append([r.get("fold", "")] + [r.get(k, "") for k in keys])

    ws.append([_mean_row(fold_rows, keys).get("fold", "")] + [_mean_row(fold_rows, keys).get(k, "") for k in keys])

    if final_row is not None:
        ws.append([""] + [""] * len(keys))
        ws.append([final_row.get("fold", "")] + [final_row.get(k, "") for k in keys])

    for ci, c in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(60, max(10, len(str(c)) + 2))

    wb.save(str(path))


def _write_overview_mean_std_xlsx(path: Path, rows: list[dict[str, Any]], keys: list[str]) -> None:
    """_write_overview_mean_std_xlsx(path, rows, keys) -> None: mean/std xlsx (computed over provided rows)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "mean_std"
    ws.append(["metric", "mean", "std"])

    for k in keys:
        vals: list[float] = []
        for r in rows:
            v = eval_utils._safe_float(r.get(k))
            if v is not None:
                vals.append(float(v))
        if not vals:
            ws.append([k, "", ""])
            continue
        mean = sum(vals) / float(len(vals))
        var = sum((x - mean) ** 2 for x in vals) / float(len(vals))
        ws.append([k, mean, var ** 0.5])

    ws.column_dimensions[get_column_letter(1)].width = 60
    ws.column_dimensions[get_column_letter(2)].width = 18
    ws.column_dimensions[get_column_letter(3)].width = 18
    wb.save(str(path))


def _plot_overview_bar_and_box(out_plots: Path, rows: list[dict[str, Any]], keys: list[str], run_name: str) -> None:
    """_plot_overview_bar_and_box(out_plots, rows, keys, run_name) -> None: Bar plots for AP/AR/latency + extra metrics."""
    out_plots.mkdir(parents=True, exist_ok=True)
    rn = eval_utils.safe_slug(run_name) if run_name else "run"

    def _keep(k: str) -> bool:
        return (
                k.startswith("AP_")
                or k.startswith("AR_")
                or k.startswith("AP_precision")
                or k.startswith("AR_recall")
                or k.startswith("latency_")
                or k.startswith("image_level_")
                or k in (
                    "precision", "recall", "f1", "tp", "fp", "fn",
                    "pr_ap_global", "roc_auc_image_level",
                    "image_level_TP", "image_level_FP", "image_level_FN", "image_level_TN",
                )
        )

    for k in [kk for kk in keys if _keep(kk)]:
        eval_utils.plot_barplot(
            out_plots / f"{eval_utils.safe_slug(k)}_bar_{rn}.png",
            f"{k} per fold",
            rows,
            k,
        )

def _copy_extra_plots_to_overview(overview_split_dir: Path, folds: list[Path], results_dir: Path, split: str, run_name: str) -> None:
    """_copy_extra_plots_to_overview(overview_split_dir, folds, results_dir, split, run_name) -> None: Copy per-fold extra plots into overview."""
    import shutil

    dst_root = overview_split_dir / "extra_plots"
    dst_root.mkdir(parents=True, exist_ok=True)
    rn = eval_utils.safe_slug(run_name) if run_name else "run"

    def _copy_from(src_dir: Path, tag: str) -> None:
        plots = src_dir / "plots"
        if not plots.exists():
            return
        for p in plots.glob("*.png"):
            shutil.copyfile(p, dst_root / f"{p.stem}_{rn}_{tag}{p.suffix}")

    for f in folds:
        mdir = f / "metrics" / split
        if mdir.exists():
            _copy_from(mdir, f.name)

    final_mdir = results_dir / "final_location_holdout" / "metrics" / split
    if final_mdir.exists():
        _copy_from(final_mdir, "final")

def _write_combined_overview_xlsx(path: Path, sections: list[tuple[str, list[dict[str, Any]], list[str], dict[str, Any] | None]]) -> None:
    """_write_combined_overview_xlsx(path, sections) -> None: One sheet with multiple sections stacked."""
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "overview"

    for si, (title, fold_rows, keys, final_row) in enumerate(sections):
        if si != 0:
            ws.append([])
            ws.append([])

        ws.append([title])
        ws.append(["fold"] + keys)

        for r in fold_rows:
            ws.append([r.get("fold", "")] + [r.get(k, "") for k in keys])

        mean_r = _mean_row(fold_rows, keys)
        ws.append([mean_r.get("fold", "")] + [mean_r.get(k, "") for k in keys])

        if final_row is not None:
            ws.append([""] + [""] * len(keys))
            ws.append([final_row.get("fold", "")] + [final_row.get(k, "") for k in keys])

    wb.save(str(path))

def _copy_metrics_extra_to_overview(overview_split_dir: Path, folds: list[Path], results_dir: Path, split: str) -> None:
    """_copy_metrics_extra_to_overview(overview_split_dir, folds, results_dir, split) -> None: Copy metrics_extra.json into overview/{split}/extra_metrics."""
    import shutil

    dst_dir = overview_split_dir / "extra_metrics"
    dst_dir.mkdir(parents=True, exist_ok=True)

    def _copy_one(src_dir: Path, tag: str) -> None:
        src = src_dir / "metrics_extra.json"
        if src.exists():
            shutil.copyfile(src, dst_dir / f"{tag}_metrics_extra.json")

        # Also copy best-threshold txt (if present)
        thr_files = sorted(src_dir.glob("best_image_level_f1_threshold_*.txt"))
        if thr_files:
            shutil.copyfile(thr_files[0], dst_dir / f"{tag}_best_image_level_f1_threshold.txt")

    for f in folds:
        mdir = f / "metrics" / split
        if mdir.exists():
            _copy_one(mdir, f.name)

    final_mdir = results_dir / "final_location_holdout" / "metrics" / split
    if final_mdir.exists():
        _copy_one(final_mdir, "FINAL")


def _copy_predictions_to_overview(overview_split_dir: Path, folds: list[Path], results_dir: Path, split: str) -> None:
    """_copy_predictions_to_overview(overview_split_dir, folds, results_dir, split) -> None: Copy predictions (merged + ranks) into overview/{split}/predictions."""
    import shutil

    dst_dir = overview_split_dir / "predictions"
    dst_dir.mkdir(parents=True, exist_ok=True)

    def _copy_one(src_dir: Path, tag: str) -> None:
        merged = src_dir / "predictions_coco.json"
        rank0 = src_dir / "predictions_coco.json.rank0.json"
        ranks = sorted(src_dir.glob("predictions_coco.json.rank*.json"))

        # Prefer merged; otherwise fall back to rank0; otherwise first available rank.
        src_main: Path | None = None
        if merged.exists():
            src_main = merged
        elif rank0.exists():
            src_main = rank0
        elif ranks:
            src_main = ranks[0]

        if src_main is not None:
            shutil.copyfile(src_main, dst_dir / f"{tag}_predictions_coco.json")

        # Always also copy all rank files (if any) for debugging/traceability
        for rp in ranks:
            shutil.copyfile(rp, dst_dir / f"{tag}_{rp.name}")

    for f in folds:
        mdir = f / "metrics" / split
        if mdir.exists():
            _copy_one(mdir, f.name)

    final_mdir = results_dir / "final_location_holdout" / "metrics" / split
    if final_mdir.exists():
        _copy_one(final_mdir, "FINAL")

def _copy_final_checkpoint(results_dir: Path, overview_dir: Path) -> None:
    """_copy_final_checkpoint(results_dir, overview_dir) -> None: Copy final_location_holdout checkpoint into overview."""
    import shutil

    final_dir = results_dir / "final_location_holdout"
    if not final_dir.exists():
        return

    ckpt: Path | None = None
    for name in ["best_stg2.pth", "best_stg1.pth", "last.pth"]:
        p = final_dir / name
        if p.exists():
            ckpt = p
            break
    if ckpt is None:
        return

    dst = overview_dir / ckpt.name
    shutil.copyfile(ckpt, dst)


def _plot_suffix(run_name: str, tag: str) -> str:
    """_plot_suffix(run_name, tag) -> str: Build _RUN_TAG suffix (empty if missing)."""
    run = (run_name or "").strip()
    t = (tag or "").strip()
    if run and t:
        return f"_{run}_{t}"
    if run:
        return f"_{run}"
    if t:
        return f"_{t}"
    return ""




def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo_root", default=".")
    ap.add_argument("--results_dir", required=True)
    ap.add_argument("--eval_name", default="eval_data")
    ap.add_argument("--select_metric", default="AP_precision_iou_0.50_area_all_maxdets_100")
    ap.add_argument("--per_fold_only", default="0")
    ap.add_argument("--overwrite", default="0")
    ap.add_argument("--run_name", default="")
    args = ap.parse_args()

    results_dir = Path(args.results_dir).resolve()
    run_name = str(getattr(args, "run_name", "") or "").strip()

    # DELETE all old plots folders under this run before re-generating anything

    for p in results_dir.rglob("plots"):
        if p.is_dir():
            shutil.rmtree(p, ignore_errors=True)

    folds = find_folds(results_dir)

    final_dir = results_dir / "final_location_holdout"
    has_final_dir = final_dir.exists() and final_dir.is_dir()

    if not folds and not has_final_dir:
        print(
            f"[evaluate_models.py] ERROR: no folds found under {results_dir / 'cross_validation'} "
            f"and no final_location_holdout found under {final_dir}"
        )
        return 2

    if not folds and has_final_dir:
        print(
            "[evaluate_models.py] INFO: no cross-validation folds found; "
            "continuing with final_location_holdout only.",
            flush=True,
        )

    fold_train_reports: list[dict[str, Any]] = []
    for f in folds:
        rep = _write_fold_train_outputs(f, run_name)
        fold_train_reports.append({"fold": f.name, **rep})

    final_train_report: dict[str, Any] | None = None
    if has_final_dir:
        rep = _write_fold_train_outputs(final_dir, run_name)
        final_train_report = {"folder": final_dir.name, **rep}

    overview_dir = results_dir / "overview"
    overview_dir.mkdir(parents=True, exist_ok=True)

    final_val = _collect_eval_metrics_for_final(results_dir, "validation", args.eval_name)
    final_test = _collect_eval_metrics_for_final(results_dir, "test", args.eval_name)

    # Collect VAL rows/keys
    val_rows: list[dict[str, Any]] = []
    val_keys: set[str] = set()
    for f in folds:
        m = _collect_eval_metrics_for_fold(f, "validation", args.eval_name)
        row = {"fold": f.name}
        row.update(m)
        for k in m.keys():
            if eval_utils._safe_float(m.get(k)) is not None:
                val_keys.add(k)
        val_rows.append(row)
    for k in final_val.keys():
        if eval_utils._safe_float(final_val.get(k)) is not None:
            val_keys.add(k)
    val_keys_sorted = sorted(val_keys)

    # Collect TEST rows/keys
    test_rows: list[dict[str, Any]] = []
    test_keys: set[str] = set()
    for f in folds:
        m = _collect_eval_metrics_for_fold(f, "test", args.eval_name)
        row = {"fold": f.name}
        row.update(m)
        for k in m.keys():
            if eval_utils._safe_float(m.get(k)) is not None:
                test_keys.add(k)
        test_rows.append(row)
    for k in final_test.keys():
        if eval_utils._safe_float(final_test.get(k)) is not None:
            test_keys.add(k)
    test_keys_sorted = sorted(test_keys)

    final_val_row = {"fold": "FINAL", **final_val} if final_val else None
    final_test_row = {"fold": "FINAL", **final_test} if final_test else None

    val_mean_rows = list(val_rows)
    if not val_mean_rows and final_val_row is not None:
        val_mean_rows = [final_val_row]

    test_mean_rows = list(test_rows)
    if not test_mean_rows and final_test_row is not None:
        test_mean_rows = [final_test_row]

    val_plot_rows = list(val_rows)
    if not val_plot_rows and final_val_row is not None:
        val_plot_rows = [final_val_row]

    test_plot_rows = list(test_rows)
    if not test_plot_rows and final_test_row is not None:
        test_plot_rows = [final_test_row]

    # Subfolder overviews
    val_dir = overview_dir / "validation"
    val_dir.mkdir(parents=True, exist_ok=True)
    _write_overview_xlsx_with_final(val_dir / f"overview_{run_name}.xlsx", val_rows, val_keys_sorted, final_val_row)
    _write_overview_mean_std_xlsx(val_dir / f"overview_mean_std_{run_name}.xlsx", val_mean_rows, val_keys_sorted)
    if val_keys_sorted and val_plot_rows:
        _plot_overview_bar_and_box(val_dir / "plots", val_plot_rows, val_keys_sorted, run_name)

    test_dir = overview_dir / "test"
    test_dir.mkdir(parents=True, exist_ok=True)
    _write_overview_xlsx_with_final(test_dir / f"overview_{run_name}.xlsx", test_rows, test_keys_sorted, final_test_row)
    _write_overview_mean_std_xlsx(test_dir / f"overview_mean_std_{run_name}.xlsx", test_mean_rows, test_keys_sorted)
    if test_keys_sorted and test_plot_rows:
        _plot_overview_bar_and_box(test_dir / "plots", test_plot_rows, test_keys_sorted, run_name)

    _copy_predictions_to_overview(val_dir, folds, results_dir, "validation")
    _copy_metrics_extra_to_overview(val_dir, folds, results_dir, "validation")

    _copy_predictions_to_overview(test_dir, folds, results_dir, "test")
    _copy_metrics_extra_to_overview(test_dir, folds, results_dir, "test")

    _copy_extra_plots_to_overview(val_dir, folds, results_dir, "validation", run_name)
    _copy_extra_plots_to_overview(test_dir, folds, results_dir, "test", run_name)

    _copy_final_checkpoint(results_dir, overview_dir)

    # Create overview/training and copy AP + train/val loss plots into it
    training_dir = overview_dir / "training"
    training_dir.mkdir(parents=True, exist_ok=True)

    def _is_training_plot(p: Path) -> bool:
        n = p.name
        return (
                "_over_epoch_" in n
                or n.startswith("train_vs_val_loss_over_epoch_")
                or n.startswith("train_minus_val_loss_over_epoch_")
        )

    # Copy from fold/final training plots folders
    for src in results_dir.rglob("plots"):
        if not src.is_dir():
            continue
        for png in src.glob("*.png"):
            if _is_training_plot(png):
                shutil.copy2(png, training_dir / png.name)

    # Top-level overview.xlsx: stack VALIDATION section then TEST section (each has its own MEAN and FINAL)
    def _keep_root_key(k: str) -> bool:
        return (
                k.startswith("AP_")
                or k.startswith("AR_")
                or k.startswith("AP_precision")
                or k.startswith("AR_recall")
                or k.startswith("latency_")
                or k.startswith("image_level_")
                or k in (
                    "precision", "recall", "f1", "tp", "fp", "fn",
                    "pr_ap_global", "roc_auc_image_level",
                    "image_level_TP", "image_level_FP", "image_level_FN", "image_level_TN",
                )
        )

    val_keys_root = [k for k in val_keys_sorted if _keep_root_key(k)]
    test_keys_root = [k for k in test_keys_sorted if _keep_root_key(k)]

    # Filter rows to only those keys for cleaner sheet (optional, but matches your example intent)
    def _filter_rows(rows: list[dict[str, Any]], keys: list[str]) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        for r in rows:
            rr: dict[str, Any] = {"fold": r.get("fold", "")}
            for k in keys:
                rr[k] = r.get(k, "")
            out.append(rr)
        return out

    val_keys_root = [k for k in val_keys_sorted if _keep_root_key(k)]
    test_keys_root = [k for k in test_keys_sorted if _keep_root_key(k)]

    def _filter_rows(rows: list[dict[str, Any]], keys: list[str]) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        for r in rows:
            rr: dict[str, Any] = {"fold": r.get("fold", "")}
            for k in keys:
                rr[k] = r.get(k, "")
            out.append(rr)
        return out

    val_rows_root = _filter_rows(val_rows, val_keys_root)
    test_rows_root = _filter_rows(test_rows, test_keys_root)

    final_val_root = None
    if final_val_row is not None:
        final_val_root = {"fold": "FINAL", **{k: final_val_row.get(k, "") for k in val_keys_root}}

    final_test_root = None
    if final_test_row is not None:
        final_test_root = {"fold": "FINAL", **{k: final_test_row.get(k, "") for k in test_keys_root}}

    _write_combined_overview_xlsx(
        overview_dir / f"overview_ALL_summary_{run_name}.xlsx",
        [
            ("VALIDATION", val_rows_root, val_keys_root, final_val_root),
            ("TEST", test_rows_root, test_keys_root, final_test_root),
        ],
    )

    val_stats_root_rows = list(val_rows_root)
    if not val_stats_root_rows and final_val_root is not None:
        val_stats_root_rows = [final_val_root]

    _write_overview_mean_std_xlsx(
        overview_dir / f"overview_ALL_mean_std_{run_name}.xlsx",
        val_stats_root_rows,
        val_keys_root,
    )
    _write_fold_split_overview_xlsx(overview_dir / f"overview_ALL_fold_split_{run_name}.xlsx", folds)

    write_json(
        overview_dir / "summary.json",
        {
            "results_dir": str(results_dir),
            "num_folds": len(folds),
            "folds": [f.name for f in folds],
            "fold_train_reports": fold_train_reports,
            "final_train_report": final_train_report,
            "val_metric_keys": val_keys_sorted,
            "test_metric_keys": test_keys_sorted,
            "root_overview_sections": ["validation", "test"],
            "final_validation_included": bool(final_val_row is not None),
            "final_test_included": bool(final_test_row is not None),
            "note": f"overview/overview_{run_name}.xlsx stacks VALIDATION then TEST, each with its own MEAN and FINAL (FINAL excluded from mean). Subfolder overview_{run_name}.xlsx files also include FINAL similarly. mean/std files exclude FINAL.",
        },
    )



    print(f"[evaluate_models.py] Wrote overview to: {overview_dir}", flush=True)
    return 0



if __name__ == "__main__":
    raise SystemExit(main())