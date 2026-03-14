#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


PLOT_FONT_SIZE_TITLE = 18
PLOT_FONT_SIZE_AXIS = 18
PLOT_FONT_SIZE_TICKS = 16
PLOT_FONT_SIZE_LEGEND = 16
PLOT_FONT_SIZE_LEGEND_LARGE = 18

# Manual y-limit overrides. Set to None to use default behavior.
YLIM_METRIC_MIN = 0.0
YLIM_METRIC_MAX = 1.0

YLIM_LOSS_SHARED_MIN = 0.0
YLIM_LOSS_SHARED_MAX = 10.0

YLIM_LOSS_TOTAL_MIN = 0.0
YLIM_LOSS_TOTAL_MAX = 60.0

YLIM_LOSS_GAP_MIN: float | None = None
YLIM_LOSS_GAP_MAX: float | None = None


DEFAULT_LOSS_WEIGHT_DICT: dict[str, float] = {
    "loss_mal": 1.0,
    "loss_bbox": 5.0,
    "loss_giou": 2.0,
    "loss_fgl": 0.0,
    "loss_ddf": 0.0,
}

# DEFAULT_LOSS_WEIGHT_DICT: dict[str, float] = {
#     "loss_mal": 1.0,
#     "loss_bbox": 1.0,
#     "loss_giou": 1.0,
#     "loss_fgl": 1.0,
#     "loss_ddf": 1.0,
# }

_AVG_RE = re.compile(r"\bAveraged stats:\s.*?\bloss:\s[-+0-9.eE]+\s\(([-+0-9.eE]+)\)")
_EPOCH_RE = re.compile(r"^Epoch:\s*\[(\d+)\]")
_LOSS_KV_RE = re.compile(r"\b(loss(?:_[A-Za-z0-9]+)*):\s*[-+0-9.eE]+\s*\(([-+0-9.eE]+)\)")
_STDOUT_EPOCH_RE = re.compile(r"^Epoch:\s*\[(\d+)\]")
_STDOUT_AVG_RE = re.compile(r"^Averaged stats:\s*(.*)$")
_STDOUT_KV_RE = re.compile(r"([A-Za-z0-9_]+):\s*([+-]?\d+(?:\.\d+)?)(?:\s*\(([+-]?\d+(?:\.\d+)?)\))?")



def _clean_plot_title(title: str, key: str | None = None) -> str:
    """_clean_plot_title(title, key=None) -> str: Replace long metric keys in titles with short readable labels."""
    t = str(title or "")

    replacements = [
        ("AP_precision_iou_0.50:0.95_area_all_maxdets_100", "AP"),
        ("AP_precision_iou_0.50_area_all_maxdets_100", "AP50"),
        ("AP_precision_iou_0.75_area_all_maxdets_100", "AP75"),
        ("AR_recall_iou_0.50:0.95_area_all_maxdets_1", "AR@1"),
        ("AR_recall_iou_0.50:0.95_area_all_maxdets_10", "AR@10"),
        ("AR_recall_iou_0.50:0.95_area_all_maxdets_100", "AR"),
        ("AR_recall_iou_0.50:0.95_area_small_maxdets_100", "AR small"),
        ("AR_recall_iou_0.50:0.95_area_medium_maxdets_100", "AR medium"),
        ("AR_recall_iou_0.50:0.95_area_large_maxdets_100", "AR large"),
        ("AP_precision_iou_0.50:0.95_area_small_maxdets_100", "AP small"),
        ("AP_precision_iou_0.50:0.95_area_medium_maxdets_100", "AP medium"),
        ("AP_precision_iou_0.50:0.95_area_large_maxdets_100", "AP large"),
    ]

    for old, new in replacements:
        t = t.replace(old, new)

    if key:
        t = t.replace(str(key), _clean_metric_label(str(key)))

    t = re.sub(r"\s+", " ", t).strip()
    return t



def _is_metric_key(key: str | None) -> bool:
    """_is_metric_key(key) -> bool: True for AP/AR/mAP-style metric keys."""
    k = str(key or "").lower()
    return (
        "map" in k
        or k.startswith("ap_")
        or k.startswith("ar_")
        or "ap_precision" in k
        or "ar_recall" in k
        or re.search(r"\bap\b", k) is not None
        or re.search(r"\bar\b", k) is not None
    )

def _clean_metric_label(key: str) -> str:
    """_clean_metric_label(key) -> str: Convert long COCO metric key to readable axis label."""
    k = str(key or "").lower()

    # ---- AP metrics ----
    if "ap_precision_iou_0.50:0.95_area_all_maxdets_100" in k:
        return "AP"
    if "ap_precision_iou_0.50_area_all_maxdets_100" in k:
        return "AP50"
    if "ap_precision_iou_0.75_area_all_maxdets_100" in k:
        return "AP75"

    if "ap_precision_iou_0.50:0.95_area_small" in k:
        return "AP small"
    if "ap_precision_iou_0.50:0.95_area_medium" in k:
        return "AP medium"
    if "ap_precision_iou_0.50:0.95_area_large" in k:
        return "AP large"

    # ---- AR metrics ----
    if "ar_recall_iou_0.50:0.95_area_all_maxdets_1" in k:
        return "AR@1"
    if "ar_recall_iou_0.50:0.95_area_all_maxdets_10" in k:
        return "AR@10"
    if "ar_recall_iou_0.50:0.95_area_all_maxdets_100" in k:
        return "AR@100"

    if "ar_recall_iou_0.50:0.95_area_small_maxdets_100" in k:
        return "AR small"
    if "ar_recall_iou_0.50:0.95_area_medium_maxdets_100" in k:
        return "AR medium"
    if "ar_recall_iou_0.50:0.95_area_large_maxdets_100" in k:
        return "AR large"

    return key

def _apply_title(title: str) -> None:
    """_apply_title(title) -> None: Apply consistent title font size."""
    plt.title(title, fontsize=PLOT_FONT_SIZE_TITLE)


def _apply_axis_fonts(xlabel: str, ylabel: str) -> None:
    """_apply_axis_fonts(xlabel, ylabel) -> None: Apply axis label/tick font sizes."""
    plt.xlabel(xlabel, fontsize=PLOT_FONT_SIZE_AXIS)
    plt.ylabel(ylabel, fontsize=PLOT_FONT_SIZE_AXIS)
    plt.tick_params(axis="both", labelsize=PLOT_FONT_SIZE_TICKS)


def _apply_metric_axis_ticks(key: str | None) -> None:
    """_apply_metric_axis_ticks(key) -> None: Apply 0..1 metric ticks, respecting manual overrides."""
    if not _is_metric_key(key):
        return

    ymin = 0.0 if YLIM_METRIC_MIN is None else float(YLIM_METRIC_MIN)
    ymax = 1.0 if YLIM_METRIC_MAX is None else float(YLIM_METRIC_MAX)

    plt.ylim(ymin, ymax)

    if abs(ymin - 0.0) < 1e-12 and abs(ymax - 1.0) < 1e-12:
        ticks = [i / 10.0 for i in range(11)]
        plt.yticks(ticks, [f"{t:.1f}" for t in ticks])


def _apply_shared_loss_axis() -> None:
    """_apply_shared_loss_axis() -> None: Apply shared-loss y-axis limits and ticks."""
    ymin = 0.0 if YLIM_LOSS_SHARED_MIN is None else float(YLIM_LOSS_SHARED_MIN)
    ymax = 8.0 if YLIM_LOSS_SHARED_MAX is None else float(YLIM_LOSS_SHARED_MAX)
    plt.ylim(ymin, ymax)

    if float(ymin).is_integer() and float(ymax).is_integer():
        a = int(round(ymin))
        b = int(round(ymax))
        if b >= a and (b - a) <= 25:
            plt.yticks(range(a, b + 1, 1))


def _apply_total_loss_axis() -> None:
    """_apply_total_loss_axis() -> None: Apply total-loss y-axis limits and ticks."""
    ymin = 0.0 if YLIM_LOSS_TOTAL_MIN is None else float(YLIM_LOSS_TOTAL_MIN)
    ymax = 100.0 if YLIM_LOSS_TOTAL_MAX is None else float(YLIM_LOSS_TOTAL_MAX)
    plt.ylim(ymin, ymax)

    if abs(ymin - round(ymin)) < 1e-12 and abs(ymax - round(ymax)) < 1e-12:
        a = int(round(ymin))
        b = int(round(ymax))
        if b >= a:
            step = 10 if (b - a) >= 20 else 1
            plt.yticks(range(a, b + 1, step))


def _apply_loss_gap_axis() -> None:
    """_apply_loss_gap_axis() -> None: Apply optional manual y-axis limits for loss-gap plots."""
    if YLIM_LOSS_GAP_MIN is None and YLIM_LOSS_GAP_MAX is None:
        return

    cur_lo, cur_hi = plt.ylim()
    ymin = cur_lo if YLIM_LOSS_GAP_MIN is None else float(YLIM_LOSS_GAP_MIN)
    ymax = cur_hi if YLIM_LOSS_GAP_MAX is None else float(YLIM_LOSS_GAP_MAX)
    plt.ylim(ymin, ymax)


def read_json(path: str) -> Any:
    """read_json(path) -> Any: Read JSON."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: str, obj: Any) -> None:
    """write_json(path, obj) -> None: Write JSON with mkdir."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")


def safe_float(x: Any) -> float | None:
    """safe_float(x) -> float|None: Parse float safely."""
    try:
        if x is None:
            return None
        if isinstance(x, bool):
            return None
        return float(x)
    except Exception:
        return None


def _safe_float(x: Any) -> float | None:
    """_safe_float(x) -> float|None: Backward-compatible alias for safe_float."""
    return safe_float(x)


def mean_std(vals: list[float]) -> tuple[float | None, float | None]:
    """mean_std(vals) -> (mean,std): Mean/std for list; returns (None,None) if empty."""
    if not vals:
        return None, None
    if len(vals) == 1:
        return float(vals[0]), 0.0
    m = sum(vals) / float(len(vals))
    v = sum((x - m) ** 2 for x in vals) / float(max(1, len(vals) - 1))
    return float(m), float(math.sqrt(v))


def ensure_dir(p: Path) -> None:
    """ensure_dir(p) -> None: mkdir(parents=True, exist_ok=True)."""
    p.mkdir(parents=True, exist_ok=True)


def _first_existing(paths: list[Path]) -> Path | None:
    """_first_existing(paths) -> Path|None: First existing path."""
    for p in paths:
        try:
            if p.exists():
                return p
        except Exception:
            continue
    return None


def strip_numbered_suffix(p: Path) -> Path:
    """strip_numbered_suffix(p) -> Path: Convert DEIMxxx_4 to DEIMxxx if present."""
    m = re.match(r"^(.*)_\d+$", p.name)
    if not m:
        return p
    return p.with_name(m.group(1))


def resolve_results_dir(results_dir: Path) -> Path:
    """resolve_results_dir(results_dir) -> Path: Handle training-machine vs eval-machine naming."""
    if results_dir.exists():
        return results_dir
    alt = strip_numbered_suffix(results_dir)
    if alt.exists():
        return alt
    return results_dir


def read_meta(fold_dir: Path) -> dict[str, Any]:
    """read_meta(fold_dir) -> dict[str,Any]: Read fold_meta.json or final_meta.json if present."""
    cand = fold_dir / "fold_meta.json"
    if cand.exists():
        try:
            j = read_json(str(cand))
            if isinstance(j, dict):
                return j
        except Exception:
            return {}
    cand2 = fold_dir / "final_meta.json"
    if cand2.exists():
        try:
            j = read_json(str(cand2))
            if isinstance(j, dict):
                return j
        except Exception:
            return {}
    return {}


def resolve_val_ann_path(*args) -> Path:
    """resolve_val_ann_path(results_dir?, fold_dir, meta) -> Path: Find validation annotations (supports old/new call signatures)."""
    if len(args) == 2:
        fold_dir, meta = args
    elif len(args) == 3:
        _, fold_dir, meta = args
    else:
        raise TypeError(f"resolve_val_ann_path() expected 2 or 3 args, got {len(args)}")

    fold_dir = Path(fold_dir)
    meta = meta if isinstance(meta, dict) else {}

    for k in ["val_ann", "coco_val", "ann_val"]:
        v = str(meta.get(k) or "").strip()
        if v:
            p = Path(v).expanduser()
            if p.exists():
                return p
            p2 = fold_dir / p.name
            if p2.exists():
                return p2

    cands = [
        fold_dir / "splits" / "instances_val.json",
        fold_dir / "instances_val.json",
        fold_dir / "val.json",
        fold_dir / "annotations" / "instances_val.json",
        fold_dir / "annotations" / "val.json",
    ]
    p = _first_existing(cands)
    if p is None:
        raise FileNotFoundError(f"Could not resolve val annotation json under {fold_dir}")
    return p


def resolve_test_ann_path(*args) -> Path:
    """resolve_test_ann_path(results_dir?, fold_dir, meta) -> Path: Find test annotations (supports old/new call signatures)."""
    if len(args) == 2:
        fold_dir, meta = args
    elif len(args) == 3:
        _, fold_dir, meta = args
    else:
        raise TypeError(f"resolve_test_ann_path() expected 2 or 3 args, got {len(args)}")

    fold_dir = Path(fold_dir)
    meta = meta if isinstance(meta, dict) else {}

    for k in ["test_ann", "coco_test", "ann_test"]:
        v = str(meta.get(k) or "").strip()
        if v:
            p = Path(v).expanduser()
            if p.exists():
                return p
            p2 = fold_dir / p.name
            if p2.exists():
                return p2

    cands = [
        fold_dir / "splits" / "instances_test.json",
        fold_dir / "instances_test.json",
        fold_dir / "test.json",
        fold_dir / "annotations" / "instances_test.json",
        fold_dir / "annotations" / "test.json",
    ]
    p = _first_existing(cands)
    if p is None:
        raise FileNotFoundError(f"Could not resolve test annotation json under {fold_dir}")
    return p


def resolve_img_root(meta: dict[str, Any], default: Path) -> Path:
    """resolve_img_root(meta, default) -> Path: Resolve img_root with fallback."""
    v = str(meta.get("img_root") or meta.get("image_root") or "").strip()
    if v:
        p = Path(v).expanduser()
        if p.exists():
            return p
    return default


def resolve_img_root_test(meta: dict[str, Any], default: Path) -> Path:
    """resolve_img_root_test(meta, default) -> Path: Resolve img_root_test with fallback."""
    v = str(meta.get("img_root_test") or meta.get("image_root_test") or "").strip()
    if v:
        p = Path(v).expanduser()
        if p.exists():
            return p
    return default


def resolve_checkpoint(meta: dict[str, Any], fold_dir: Path) -> Path | None:
    """resolve_checkpoint(meta, fold_dir) -> Path|None: Resolve checkpoint path."""
    for k in ["best_checkpoint", "checkpoint", "final_checkpoint", "ckpt"]:
        v = str(meta.get(k) or "").strip()
        if v:
            p = Path(v).expanduser()
            if p.exists():
                return p
            p2 = fold_dir / p.name
            if p2.exists():
                return p2
            p3 = fold_dir / "checkpoints" / p.name
            if p3.exists():
                return p3

    for name in ["best_stg2.pth", "best_stg1.pth", "best.pth", "last.pth"]:
        p = fold_dir / name
        if p.exists():
            return p
        p2 = fold_dir / "checkpoints" / name
        if p2.exists():
            return p2

    return None


def _make_xlsx(path: Path, header: list[str], rows: list[list[Any]]) -> None:
    """_make_xlsx(path, header, rows) -> None: Write simple XLSX."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(header)
    for r in rows:
        ws.append(list(r))

    for col in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(45, len(str(header[col - 1])) + 2))
    wb.save(str(path))


def collect_eval_metrics(eval_dir: Path) -> dict[str, Any]:
    """collect_eval_metrics(eval_dir) -> dict[str,Any]: Load metrics(+extra) and latency if present."""
    out: dict[str, Any] = {}
    m = eval_dir / "metrics.json"
    mx = eval_dir / "metrics_extra.json"
    lat = eval_dir / "latency.json"

    if m.exists():
        try:
            mj = read_json(str(m))
            if isinstance(mj, dict):
                out.update(mj)
        except Exception:
            pass

    if mx.exists():
        try:
            mxj = read_json(str(mx))
            if isinstance(mxj, dict):
                for k, v in mxj.items():
                    out[f"extra_{k}"] = v
        except Exception:
            pass

    if lat.exists():
        try:
            lj = read_json(str(lat))
            if isinstance(lj, dict):
                if "ms_per_image" in lj:
                    out["latency_ms_per_image"] = lj.get("ms_per_image")
                if "total_ms" in lj:
                    out["latency_total_ms"] = lj.get("total_ms")
                if "num_images" in lj:
                    out["latency_num_images"] = lj.get("num_images")
        except Exception:
            pass

    return out


def _cleanup_epoch_plots(plots_dir: Path, prefixes: list[str]) -> None:
    """_cleanup_epoch_plots(plots_dir, prefixes) -> None: Delete old epoch plots before re-plotting."""
    plots_dir.mkdir(parents=True, exist_ok=True)
    for pref in prefixes:
        for p in plots_dir.glob(f"{pref}*.png"):
            try:
                p.unlink()
            except Exception:
                pass


def plot_barplot(out_path: Path, title: str, rows: list[dict[str, Any]], key: str) -> None:
    """plot_barplot(out_path, title, rows, key) -> None: Bar plot per fold for key."""
    labels: list[str] = []
    ys: list[float] = []
    for r in rows:
        labels.append(str(r.get("fold", "")))
        v = safe_float(r.get(key))
        ys.append(float(v) if v is not None else float("nan"))

    if not labels:
        return

    out_path.parent.mkdir(parents=True, exist_ok=True)
    plt.figure(figsize=(max(6, len(labels) * 0.6), 4))
    plt.bar(labels, ys)

    _apply_title(_clean_plot_title(title, key))
    _apply_axis_fonts("Run", _clean_metric_label(key))
    plt.xticks(rotation=45, ha="right")

    _apply_metric_axis_ticks(key)

    plt.tight_layout()
    plt.savefig(str(out_path))
    plt.close()


def plot_boxplot(out_path: Path, title: str, rows: list[dict[str, Any]], key: str) -> None:
    """plot_boxplot(out_path, title, rows, key) -> None: Box plot per fold for key."""
    labels: list[str] = []
    vals: list[list[float]] = []
    for r in rows:
        labels.append(str(r.get("fold", "")))
        v = safe_float(r.get(key))
        if v is None:
            vals.append([float("nan")])
        else:
            vals.append([float(v)])

    if not labels:
        return

    out_path.parent.mkdir(parents=True, exist_ok=True)
    plt.figure(figsize=(max(6, len(labels) * 0.4), 4))
    plt.boxplot(vals, labels=labels, vert=True)

    _apply_title(_clean_plot_title(title, key))
    _apply_axis_fonts("Run", _clean_metric_label(key))
    plt.xticks(rotation=45, ha="right")

    _apply_metric_axis_ticks(key)

    plt.tight_layout()
    plt.savefig(str(out_path))
    plt.close()


def copy_eval_artifacts_to_metrics_folder(eval_dir: Path, fold_dir: Path, split_name: str, overwrite: str = "0") -> dict[str, Any]:
    """copy_eval_artifacts_to_metrics_folder(eval_dir, fold_dir, split_name, overwrite) -> dict[str,Any]: Copy key eval files into fold_dir/metrics/{split_name}."""
    dst = fold_dir / "metrics" / split_name
    dst.mkdir(parents=True, exist_ok=True)

    copied: list[str] = []
    for fn in [
        "metrics.json",
        "metrics_detailed.json",
        "metrics_extra.json",
        "eval_stdout.log",
        "latency.json",
        "metrics.xlsx",
        "predictions_coco.json",
    ]:
        src = eval_dir / fn
        if src.exists():
            d = dst / fn
            if overwrite == "1" or not d.exists():
                try:
                    shutil.copyfile(src, d)
                    copied.append(fn)
                except Exception:
                    pass

    for rp in sorted(eval_dir.glob("predictions_coco.json.rank*.json")):
        d = dst / rp.name
        if overwrite == "1" or not d.exists():
            try:
                shutil.copyfile(rp, d)
                copied.append(rp.name)
            except Exception:
                pass

    if fold_dir.name == "final_location_holdout":
        src_pred = eval_dir / "predictions_coco.json"
        if src_pred.exists():
            if split_name == "validation":
                dst_pred = fold_dir / "final_predictions_val.json"
                if overwrite == "1" or not dst_pred.exists():
                    try:
                        shutil.copyfile(src_pred, dst_pred)
                        copied.append(dst_pred.name)
                    except Exception:
                        pass

            if split_name == "test":
                dst_pred = fold_dir / "final_predictions_test.json"
                if overwrite == "1" or not dst_pred.exists():
                    try:
                        shutil.copyfile(src_pred, dst_pred)
                        copied.append(dst_pred.name)
                    except Exception:
                        pass

            test_p = fold_dir / "final_predictions_test.json"
            val_p = fold_dir / "final_predictions_val.json"

            if test_p.exists():
                alias_src = test_p
            elif val_p.exists():
                alias_src = val_p
            else:
                alias_src = src_pred

            alias = fold_dir / "final_predictions.json"
            if overwrite == "1" or not alias.exists():
                try:
                    shutil.copyfile(alias_src, alias)
                    copied.append(alias.name)
                except Exception:
                    pass

    plots_src = eval_dir / "plots"
    if plots_src.exists():
        plots_dst = dst / "plots"
        plots_dst.mkdir(parents=True, exist_ok=True)
        for p in plots_src.glob("*.png"):
            d = plots_dst / p.name
            if overwrite == "1" or not d.exists():
                try:
                    shutil.copyfile(p, d)
                    copied.append(f"plots/{p.name}")
                except Exception:
                    pass

    return {"ok": True, "dst": str(dst), "copied": copied}


def resolve_checkpoint_path(results_dir: Path, fold_dir: Path, ckpt_str: str) -> Path | None:
    """resolve_checkpoint_path(results_dir, fold_dir, ckpt_str) -> Path|None: Backward-compatible checkpoint resolver."""
    s = str(ckpt_str or "").strip()
    if not s:
        return None

    p = Path(s).expanduser()
    if p.exists():
        return p

    cand = fold_dir / p.name
    if cand.exists():
        return cand

    cand2 = fold_dir / "checkpoints" / p.name
    if cand2.exists():
        return cand2

    cand3 = results_dir / "final_location_holdout" / p.name
    if cand3.exists():
        return cand3

    cand4 = results_dir / "final_location_holdout" / "checkpoints" / p.name
    if cand4.exists():
        return cand4

    return None


def safe_slug(s: str) -> str:
    """safe_slug(s) -> str: Filesystem-safe slug."""
    t = re.sub(r"[^A-Za-z0-9._-]+", "_", str(s))
    t = re.sub(r"_+", "_", t).strip("_")
    return t or "metric"


def _try_parse_json_line(line: str) -> dict[str, Any] | None:
    """_try_parse_json_line(line) -> dict|None: Parse JSON dict from a single line if possible."""
    line = line.strip()
    if not line or not line.startswith("{") or not line.endswith("}"):
        return None
    try:
        obj = json.loads(line)
    except Exception:
        return None
    return obj if isinstance(obj, dict) else None


def _coco_stats_to_named_metrics(stats: list[float]) -> dict[str, float]:
    """_coco_stats_to_named_metrics(stats) -> dict[str,float]: Convert COCOeval stats[0..11] into named AP/AR keys."""
    if not isinstance(stats, (list, tuple)) or len(stats) < 12:
        return {}
    s = [float(x) for x in stats[:12]]
    return {
        "AP_precision_iou_0.50:0.95_area_all_maxdets_100": s[0],
        "AP_precision_iou_0.50_area_all_maxdets_100": s[1],
        "AP_precision_iou_0.75_area_all_maxdets_100": s[2],
        "AP_precision_iou_0.50:0.95_area_small_maxdets_100": s[3],
        "AP_precision_iou_0.50:0.95_area_medium_maxdets_100": s[4],
        "AP_precision_iou_0.50:0.95_area_large_maxdets_100": s[5],
        "AR_recall_iou_0.50:0.95_area_all_maxdets_1": s[6],
        "AR_recall_iou_0.50:0.95_area_all_maxdets_10": s[7],
        "AR_recall_iou_0.50:0.95_area_all_maxdets_100": s[8],
        "AR_recall_iou_0.50:0.95_area_small_maxdets_100": s[9],
        "AR_recall_iou_0.50:0.95_area_medium_maxdets_100": s[10],
        "AR_recall_iou_0.50:0.95_area_large_maxdets_100": s[11],
    }


def parse_train_coco_metrics_from_log(log_path: Path) -> list[dict[str, Any]]:
    """parse_train_coco_metrics_from_log(log_path) -> list[dict]: Extract per-epoch COCO AP/AR rows from log.txt or stdout log."""
    log_path = Path(log_path)
    fold_dir = log_path.parent.parent if log_path.parent.name == "logs" else log_path.parent
    log_txt = fold_dir / "log.txt"
    sources: list[Path] = [log_txt] if log_txt.exists() else [log_path]

    rows: dict[int, dict[str, Any]] = {}
    for src in sources:
        for line in src.read_text(encoding="utf-8", errors="ignore").splitlines():
            obj = _try_parse_json_line(line)
            if not obj or "epoch" not in obj:
                continue
            try:
                ep = int(obj.get("epoch"))
            except Exception:
                continue

            r = rows.setdefault(ep, {"epoch": ep})

            for key in ["test_coco_eval_bbox", "val_coco_eval_bbox", "coco_eval_bbox"]:
                stats = obj.get(key)
                if isinstance(stats, (list, tuple)) and len(stats) >= 12:
                    r.update(_coco_stats_to_named_metrics(list(stats)))
                    break

            for k, v in obj.items():
                if k.startswith("AP_") or k.startswith("AR_") or k.startswith("AP_precision") or k.startswith("AR_recall"):
                    fv = safe_float(v)
                    if fv is not None:
                        r[k] = fv

    return [rows[k] for k in sorted(rows.keys())]


def _discover_shared_loss_keys(rows: list[dict[str, Any]]) -> list[str]:
    """_discover_shared_loss_keys(rows) -> list[str]: Loss keys that appear in BOTH train_ and val_ across rows."""
    train_keys: set[str] = set()
    val_keys: set[str] = set()

    for r in rows:
        for k in r.keys():
            if not isinstance(k, str):
                continue
            if k.startswith("train_loss") and k != "train_loss":
                train_keys.add(k[len("train_"):])
            if k.startswith("val_loss") and k != "val_loss":
                val_keys.add(k[len("val_"):])

    shared = sorted(train_keys & val_keys)
    if not shared:
        shared = ["loss_mal", "loss_bbox", "loss_giou"]
    return shared


def _compute_weighted_total(row: dict[str, Any], prefix: str, allowed_loss_keys: list[str], weight_dict: dict[str, float]) -> float | None:
    """_compute_weighted_total(row, prefix, allowed_loss_keys, weight_dict) -> float|None: Sum weight*component over allowed keys."""
    total = 0.0
    used = 0
    for lk in allowed_loss_keys:
        v = safe_float(row.get(f"{prefix}_{lk}"))
        if v is None:
            continue
        w = float(weight_dict.get(lk, 0.0))
        total += w * float(v)
        used += 1
    return total if used > 0 else None


def _attach_shared_weighted_totals(rows: list[dict[str, Any]], weight_dict: dict[str, float]) -> list[dict[str, Any]]:
    """_attach_shared_weighted_totals(rows, weight_dict) -> rows: Add train/val weighted totals over shared keys only."""
    allowed = _discover_shared_loss_keys(rows)
    for r in rows:
        tr = _compute_weighted_total(r, "train", allowed, weight_dict)
        va = _compute_weighted_total(r, "val", allowed, weight_dict)
        if tr is not None:
            r["train_loss_weighted"] = tr
        if va is not None:
            r["val_loss_weighted"] = va
    return rows


def _is_main_loss_key(k: str) -> bool:
    """_is_main_loss_key(k) -> bool: True for main loss keys (exclude aux/dn/enc/pre variants)."""
    if not k.startswith("loss"):
        return False
    if any(x in k for x in ["dn_", "aux_", "enc_", "_pre", "pre_", "detr"]):
        return False
    return True


def parse_loss_curves_from_stdout_log(log_path: Path) -> list[dict[str, Any]]:
    """parse_loss_curves_from_stdout_log(log_path) -> list[dict]: Parse per-epoch train/val main losses, plus full train losses from stdout."""
    log_path = Path(log_path)
    if log_path.is_dir():
        cand = log_path / "train_stdout.log"
        if cand.exists():
            log_path = cand

    if not log_path.exists():
        return []

    rows_by_ep: dict[int, dict[str, Any]] = {}
    cur_epoch: int | None = None
    mode: str = "train"

    lines = log_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    for line in lines:
        line = line.strip()

        m_ep = _STDOUT_EPOCH_RE.match(line)
        if m_ep:
            cur_epoch = int(m_ep.group(1))
            mode = "train"
            rows_by_ep.setdefault(cur_epoch, {"epoch": cur_epoch})
            continue

        if line.startswith("Test:"):
            mode = "val"
            continue

        m_avg = _STDOUT_AVG_RE.match(line)
        if not m_avg or cur_epoch is None:
            continue

        payload = m_avg.group(1)
        r = rows_by_ep.setdefault(cur_epoch, {"epoch": cur_epoch})

        for km in _STDOUT_KV_RE.finditer(payload):
            k = km.group(1)
            v = safe_float(km.group(3) if km.group(3) is not None else km.group(2))
            if v is None or not k.startswith("loss"):
                continue

            if mode == "train":
                r[f"trainfull_{k}"] = float(v)
                if _is_main_loss_key(k):
                    r[f"train_{k}"] = float(v)
            else:
                if _is_main_loss_key(k):
                    r[f"val_{k}"] = float(v)

    return [rows_by_ep[k] for k in sorted(rows_by_ep.keys())]


def plot_train_val_components(epochs, train_mal, val_mal, train_bbox, val_bbox, train_giou, val_giou, out_path=None):
    """plot_train_val_components(epochs, train_mal, val_mal, train_bbox, val_bbox, train_giou, val_giou, out_path=None) -> None: Plot train vs val component losses."""
    import numpy as np

    epochs = np.asarray(epochs)

    plt.figure(figsize=(10, 5))
    plt.plot(epochs, train_mal, color="orange", linestyle="-", label="mal train")
    plt.plot(epochs, val_mal, color="orange", linestyle="--", label="mal val")
    plt.plot(epochs, train_bbox, color="blue", linestyle="-", label="bbox train")
    plt.plot(epochs, val_bbox, color="blue", linestyle="--", label="bbox val")
    plt.plot(epochs, train_giou, color="green", linestyle="-", label="giou train")
    plt.plot(epochs, val_giou, color="green", linestyle="--", label="giou val")

    _apply_title("Train vs Val component losses")
    _apply_axis_fonts("Epoch", "Loss")
    _apply_shared_loss_axis()

    plt.grid(True, alpha=0.3)
    plt.legend(ncols=3, fontsize=PLOT_FONT_SIZE_LEGEND_LARGE)
    plt.tight_layout()

    if out_path is not None:
        plt.savefig(out_path, dpi=200)
    else:
        plt.show()

    return None


def parse_loss_curves_from_jsonl_logtxt(log_path: Path) -> list[dict[str, Any]]:
    """parse_loss_curves_from_jsonl_logtxt(log_path) -> list[dict]: Parse per-epoch train/val component losses from JSONL log.txt."""
    log_path = Path(log_path)

    if log_path.is_dir():
        cand = log_path / "log.txt"
        if cand.exists():
            log_path = cand

    if log_path.name != "log.txt":
        cand = log_path.parent / "log.txt"
        if cand.exists():
            log_path = cand

    if not log_path.exists():
        return []

    rows_by_ep: dict[int, dict[str, Any]] = {}
    txt = log_path.read_text(encoding="utf-8", errors="ignore").splitlines()

    def is_main_loss_key(k: str) -> bool:
        if not k.startswith("loss"):
            return False
        if any(x in k for x in ["dn_", "aux_", "enc_", "_pre", "pre_", "detr"]):
            return False
        return True

    def is_any_loss_key(k: str) -> bool:
        return k.startswith("loss")

    for line in txt:
        obj = _try_parse_json_line(line)
        if not obj or "epoch" not in obj:
            continue

        ep = safe_float(obj.get("epoch"))
        if ep is None:
            continue
        ep_i = int(ep)
        r = rows_by_ep.setdefault(ep_i, {"epoch": ep_i})

        is_val = any(k.startswith("test_") or k.startswith("val_") for k in obj.keys())

        if is_val:
            for k, v in obj.items():
                if not isinstance(k, str):
                    continue
                if k.startswith("test_") and is_main_loss_key(k[len("test_"):]):
                    fv = safe_float(v)
                    if fv is not None:
                        r[f"val_{k[len('test_'):]}"] = float(fv)
                elif k.startswith("val_") and is_main_loss_key(k[len("val_"):]):
                    fv = safe_float(v)
                    if fv is not None:
                        r[f"val_{k[len('val_'):]}"] = float(fv)
        else:
            for k, v in obj.items():
                if not isinstance(k, str):
                    continue
                if is_any_loss_key(k):
                    fv = safe_float(v)
                    if fv is not None:
                        r[f"trainfull_{k}"] = float(fv)
                        if is_main_loss_key(k):
                            r[f"train_{k}"] = float(fv)

    return [rows_by_ep[k] for k in sorted(rows_by_ep.keys())]


def _base_loss_key(loss_key: str, weight_dict: dict[str, float]) -> str | None:
    """_base_loss_key(loss_key, weight_dict) -> str|None: Map variant loss key to base key."""
    for base in weight_dict.keys():
        if loss_key == base or loss_key.startswith(base + "_"):
            return base
    return None


def _attach_real_train_weighted_totals(rows: list[dict[str, Any]], weight_dict: dict[str, float]) -> list[dict[str, Any]]:
    """_attach_real_train_weighted_totals(rows, weight_dict) -> rows: Add train_loss_real_weighted from all trainfull_loss_* keys."""
    for r in rows:
        total = 0.0
        used = 0
        for k, v in r.items():
            if not isinstance(k, str) or not k.startswith("trainfull_loss"):
                continue
            loss_key = k[len("trainfull_"):]
            base = _base_loss_key(loss_key, weight_dict)
            if base is None:
                continue
            fv = safe_float(v)
            if fv is None:
                continue
            total += float(weight_dict[base]) * float(fv)
            used += 1

        if used > 0:
            r["train_loss_real_weighted"] = float(total)

    return rows


def parse_loss_curves_from_log(log_path: Path) -> list[dict[str, Any]]:
    """parse_loss_curves_from_log(log_path) -> list[dict]: Parse loss curves and add shared + real train weighted totals."""
    log_path = Path(log_path)

    rows = parse_loss_curves_from_jsonl_logtxt(log_path)
    if not rows:
        rows = parse_loss_curves_from_stdout_log(log_path)
    if not rows:
        return []

    rows = _attach_shared_weighted_totals(rows, DEFAULT_LOSS_WEIGHT_DICT)
    rows = _attach_real_train_weighted_totals(rows, DEFAULT_LOSS_WEIGHT_DICT)
    return rows


def _loss_component_suffixes(row: dict[str, Any], prefix: str) -> set[str]:
    """_loss_component_suffixes(row, prefix) -> set[str]: Component-loss suffixes for a prefix (excludes total)."""
    total_key = f"{prefix}loss"
    suff: set[str] = set()
    for k in row.keys():
        if k.startswith(prefix) and "loss" in k and k != total_key:
            suff.add(k[len(prefix):])
    return suff


def _compute_total_loss(row: dict[str, Any], prefix: str, allowed_suffixes: set[str] | None = None) -> float | None:
    """_compute_total_loss(row, prefix, allowed_suffixes=None) -> float|None: Sum component losses (ignores total key)."""
    total_key = f"{prefix}loss"
    comps: list[float] = []
    for k, v in row.items():
        if not (k.startswith(prefix) and "loss" in k and k != total_key):
            continue
        suffix = k[len(prefix):]
        if allowed_suffixes is not None and suffix not in allowed_suffixes:
            continue
        fv = safe_float(v)
        if fv is not None:
            comps.append(float(fv))
    if comps:
        return float(sum(comps))
    return None


def plot_metric_over_epoch(out_path: Path, title: str, rows: list[dict[str, Any]], key: str) -> None:
    """plot_metric_over_epoch(out_path, title, rows, key) -> None: Line plot for a metric vs epoch."""
    xs: list[int] = []
    ys: list[float] = []
    for r in rows:
        ep = r.get("epoch")
        v = r.get(key)
        if ep is None:
            continue
        fv = safe_float(v)
        if fv is None:
            continue
        xs.append(int(ep))
        ys.append(float(fv))

    if not xs:
        return

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    plt.figure(figsize=(7, 4))
    plt.plot(xs, ys)

    _apply_title(f"{_clean_metric_label(key)} vs Epoch")
    _apply_axis_fonts("Epoch", _clean_metric_label(key))
    _apply_metric_axis_ticks(key)

    plt.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(str(out_path))
    plt.close()


def plot_train_val_loss(out_path: Path, title: str, rows: list[dict[str, Any]], mode: str = "all") -> None:
    """plot_train_val_loss(out_path, title, rows, mode='all') -> None: Plot loss curves with fixed colors."""
    xs: list[int] = []
    tr_shared: list[float] = []
    va_shared: list[float] = []
    tr_real: list[float] = []

    for r in rows:
        ep = r.get("epoch")
        if ep is None:
            continue

        tw = safe_float(r.get("train_loss_weighted"))
        vw = safe_float(r.get("val_loss_weighted"))
        rw = safe_float(r.get("train_loss_real_weighted"))

        if tw is None or vw is None:
            continue

        xs.append(int(ep))
        tr_shared.append(float(tw))
        va_shared.append(float(vw))
        tr_real.append(float(rw) if rw is not None else float("nan"))

    if not xs:
        return

    mode = str(mode or "all").strip().lower()
    if mode not in {"all", "shared", "real_vs_val"}:
        mode = "all"

    L_TRAIN_SHARED = "train (shared, weighted)"
    L_VAL_SHARED = "val (shared, weighted)"
    L_TRAIN_REAL = "train (real, weighted)"

    C_TRAIN_SHARED = "tab:blue"
    C_VAL_SHARED = "tab:orange"
    C_TRAIN_REAL = "tab:green"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    plt.figure(figsize=(7, 4))

    if mode == "shared":
        plt.plot(xs, tr_shared, label=L_TRAIN_SHARED, color=C_TRAIN_SHARED)
        plt.plot(xs, va_shared, label=L_VAL_SHARED, color=C_VAL_SHARED)
        ylabel = "Loss"
        _apply_shared_loss_axis()
    elif mode == "real_vs_val":
        plt.plot(xs, va_shared, label=L_VAL_SHARED, color=C_VAL_SHARED)
        plt.plot(xs, tr_real, label=L_TRAIN_REAL, color=C_TRAIN_REAL)
        ylabel = "Loss"
        _apply_total_loss_axis()
    else:
        plt.plot(xs, tr_shared, label=L_TRAIN_SHARED, color=C_TRAIN_SHARED)
        plt.plot(xs, va_shared, label=L_VAL_SHARED, color=C_VAL_SHARED)
        plt.plot(xs, tr_real, label=L_TRAIN_REAL, color=C_TRAIN_REAL)
        ylabel = "Loss"
        _apply_total_loss_axis()

    _apply_title(title)
    _apply_axis_fonts("Epoch", ylabel)
    plt.legend(fontsize=PLOT_FONT_SIZE_LEGEND_LARGE)
    plt.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(str(out_path))
    plt.close()


def write_train_metrics_xlsx(fold_dir: Path, coco_rows: list[dict[str, Any]]) -> None:
    """write_train_metrics_xlsx(fold_dir, coco_rows) -> None: Write train metrics to fold_dir/metrics/train_metrics.xlsx."""
    fold_dir = Path(fold_dir)
    out = fold_dir / "metrics" / "train_metrics.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    keys: list[str] = []
    seen: set[str] = set()
    for r in coco_rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    keys = ["epoch"] + [k for k in keys if k != "epoch"]

    wb = Workbook()
    ws = wb.active
    ws.title = "train_metrics"
    ws.append(keys)
    for r in coco_rows:
        ws.append([r.get(k, "") for k in keys])

    for ci, c in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(60, max(10, len(str(c)) + 2))
    wb.save(str(out))


def write_loss_xlsx(fold_dir: Path, loss_rows: list[dict[str, Any]]) -> None:
    """write_loss_xlsx(fold_dir, loss_rows) -> None: Write loss curves to fold_dir/metrics/loss_curves.xlsx."""
    fold_dir = Path(fold_dir)
    out = fold_dir / "metrics" / "loss_curves.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    keys: list[str] = []
    seen: set[str] = set()
    for r in loss_rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    keys = ["epoch"] + [k for k in keys if k != "epoch"]

    wb = Workbook()
    ws = wb.active
    ws.title = "loss_curves"
    ws.append(keys)
    for r in loss_rows:
        ws.append([r.get(k, "") for k in keys])

    for ci, c in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(60, max(10, len(str(c)) + 2))
    wb.save(str(out))


def plot_train_val_loss_gap(out_path: Path, title: str, rows: list[dict[str, Any]]) -> None:
    """plot_train_val_loss_gap(out_path, title, rows) -> None: Plot (train_weighted - val_weighted) loss gap vs epoch."""
    xs: list[int] = []
    gap: list[float] = []

    for r in rows:
        ep = r.get("epoch")
        if ep is None:
            continue

        tw = safe_float(r.get("train_loss_weighted"))
        vw = safe_float(r.get("val_loss_weighted"))
        if tw is None or vw is None:
            continue

        xs.append(int(ep))
        gap.append(float(tw) - float(vw))

    if not xs:
        return

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    plt.figure(figsize=(7, 4))
    plt.plot(xs, gap, label="train_shared - val_shared")
    plt.axhline(0.0, linewidth=1.0)

    _apply_title(title)
    _apply_axis_fonts("Epoch", "Loss")
    _apply_loss_gap_axis()

    plt.legend(fontsize=PLOT_FONT_SIZE_LEGEND)
    plt.tight_layout()

    plt.grid(True, alpha=0.3)
    plt.savefig(str(out_path))
    plt.close()


def _series(rows: list[dict[str, Any]], key: str) -> list[tuple[int, float]]:
    """_series(rows, key) -> list[(epoch,val)]: Extract sorted (epoch,val) for a key."""
    out: list[tuple[int, float]] = []
    for r in rows:
        ep = r.get("epoch")
        v = safe_float(r.get(key))
        if ep is None or v is None:
            continue
        out.append((int(ep), float(v)))
    out.sort(key=lambda t: t[0])
    return out


def plot_train_val_components_from_rows(out_path: Path, title: str, rows: list[dict[str, Any]]) -> None:
    """plot_train_val_components_from_rows(out_path, title, rows) -> None: Plot train/val MAL,BBOX,GIoU."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    mal_tr = _series(rows, "train_loss_mal")
    mal_va = _series(rows, "val_loss_mal")
    bb_tr = _series(rows, "train_loss_bbox")
    bb_va = _series(rows, "val_loss_bbox")
    gi_tr = _series(rows, "train_loss_giou")
    gi_va = _series(rows, "val_loss_giou")

    if not (mal_tr or mal_va or bb_tr or bb_va or gi_tr or gi_va):
        return

    plt.figure(figsize=(10, 5))

    if mal_tr:
        xs, ys = zip(*mal_tr)
        plt.plot(xs, ys, color="orange", linestyle="-", label="mal train")
    if mal_va:
        xs, ys = zip(*mal_va)
        plt.plot(xs, ys, color="orange", linestyle="--", label="mal val")

    if bb_tr:
        xs, ys = zip(*bb_tr)
        plt.plot(xs, ys, color="blue", linestyle="-", label="bbox train")
    if bb_va:
        xs, ys = zip(*bb_va)
        plt.plot(xs, ys, color="blue", linestyle="--", label="bbox val")

    if gi_tr:
        xs, ys = zip(*gi_tr)
        plt.plot(xs, ys, color="green", linestyle="-", label="giou train")
    if gi_va:
        xs, ys = zip(*gi_va)
        plt.plot(xs, ys, color="green", linestyle="--", label="giou val")

    _apply_title(title)
    _apply_axis_fonts("Epoch", "Loss")
    _apply_shared_loss_axis()

    plt.grid(True, alpha=0.3)
    plt.legend(ncols=3, fontsize=PLOT_FONT_SIZE_LEGEND_LARGE)
    plt.tight_layout()
    plt.savefig(str(out_path), dpi=200)
    plt.close()