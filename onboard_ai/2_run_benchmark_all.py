from __future__ import annotations

import math
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _normalize_name(text: str) -> str:
    """Normalize sheet/column names for robust matching."""
    return "".join(ch.lower() for ch in str(text) if ch.isalnum())


def _to_bool(value: object) -> bool:
    """Convert common Excel/string truthy values to bool."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"true", "1", "yes", "y", "whale", "positive"}


def _coerce_bool_series(df: pd.DataFrame, column_name: str, default: bool = False) -> pd.Series:
    """Return a boolean Series from a mixed or missing dataframe column."""
    if column_name not in df.columns:
        return pd.Series(default, index=df.index, dtype=bool)
    return df[column_name].apply(lambda value: default if pd.isna(value) else _to_bool(value)).astype(bool)


def _find_case_directories(master_results_root: Path) -> list[Path]:
    """Return all case directories inside the master results folder."""
    return sorted(path for path in master_results_root.iterdir() if path.is_dir())


def _resolve_variant_jobs(case_dir: Path, mode: str, distinct_locations: list[str]) -> list[dict[str, str | Path]]:
    """Resolve benchmark variants per case based on mode."""
    if mode not in {"random", "distinct", "all"}:
        raise ValueError("mode must be one of: 'random', 'distinct', 'all'")

    jobs: list[dict[str, str | Path]] = []
    case_name = case_dir.name

    if mode in {"random", "all"}:
        jobs.append(
            {
                "label": "default",
                "suffix": "",
                "benchmark_filename": f"benchmark_{case_name}.xlsx",
                "detection_dir": case_dir / "onboard_detection",
            }
        )

    if mode in {"distinct", "all"}:
        for location in distinct_locations:
            jobs.append(
                {
                    "label": location,
                    "suffix": f"_{location}",
                    "benchmark_filename": f"benchmark_{location}_{case_name}.xlsx",
                    "detection_dir": case_dir / f"onboard_detection_{location}",
                }
            )

    return jobs


def _find_mission_results_workbook(case_dir: Path) -> Path:
    """Find the main simulation results workbook inside one case directory."""
    candidates = sorted(
        path for path in case_dir.glob("results_*.xlsx")
        if path.is_file() and not path.name.lower().startswith("benchmark")
    )

    if not candidates:
        raise FileNotFoundError(f"No mission results workbook found in {case_dir}")

    return candidates[0]


def _find_required_workbook(path: Path) -> Path:
    """Validate workbook existence."""
    if not path.exists():
        raise FileNotFoundError(f"Required workbook does not exist: {path}")
    return path


def _has_existing_benchmark(benchmark_path: Path) -> bool:
    """Return True when the benchmark workbook already exists."""
    return benchmark_path.exists() and benchmark_path.is_file()


def _read_sheet_with_aliases(xlsx_path: Path, aliases: list[str]) -> pd.DataFrame:
    """Read a sheet by trying multiple normalized alias names."""
    excel_file = pd.ExcelFile(xlsx_path)
    normalized_to_original = {_normalize_name(name): name for name in excel_file.sheet_names}

    for alias in aliases:
        key = _normalize_name(alias)
        if key in normalized_to_original:
            return pd.read_excel(xlsx_path, sheet_name=normalized_to_original[key])

    raise KeyError(f"None of the sheet aliases {aliases} were found in {xlsx_path}")


def _extract_overview_value(overview_df: pd.DataFrame, metric_name: str) -> float | int | str | None:
    """Extract one metric value from the Overview sheet."""
    if "Metric" not in overview_df.columns or "Value" not in overview_df.columns:
        return None

    metric_series = overview_df["Metric"].astype(str).fillna("")
    match = overview_df.loc[metric_series == metric_name, "Value"]
    if match.empty:
        return None
    return match.iloc[0]


def _extract_metric_from_long_table(df: pd.DataFrame, metric_name: str) -> float | int | str | None:
    """Extract a metric from a long two-column metric/value table."""
    if "metric" not in df.columns or "value" not in df.columns:
        return None

    metric_series = df["metric"].astype(str).fillna("")
    match = df.loc[metric_series == metric_name, "value"]
    if match.empty:
        return None
    return match.iloc[0]


def _to_float(value: object) -> float | None:
    """Convert to float when possible."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _extract_total_satellites_from_case_name(case_name: str) -> int | None:
    """Extract total satellite count from case name, counting Tip and Cue separately for TC cases."""
    case_name = str(case_name)

    tc_match = re.match(r"^TC_(\d+)x(\d+)sat(?:_|$)", case_name)
    if tc_match:
        n_orbits = int(tc_match.group(1))
        n_pairs_per_orbit = int(tc_match.group(2))
        return n_orbits * n_pairs_per_orbit * 2

    c_match = re.match(r"^C_(\d+)x(\d+)sat(?:_|$)", case_name)
    if c_match:
        n_orbits = int(c_match.group(1))
        n_sats_per_orbit = int(c_match.group(2))
        return n_orbits * n_sats_per_orbit

    return None


def _parse_datetime_column(df: pd.DataFrame, column_name: str) -> pd.Series:
    """Parse datetime values using fixed formats and numeric Excel timestamps only."""
    if column_name not in df.columns:
        return pd.Series([pd.NaT] * len(df), index=df.index)

    series = df[column_name]

    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce")

    parsed = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")

    numeric_series = pd.to_numeric(series, errors="coerce")
    numeric_mask = numeric_series.notna()
    if numeric_mask.any():
        parsed.loc[numeric_mask] = pd.to_datetime(
            numeric_series[numeric_mask],
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )

    text_series = series.astype(str).str.strip()
    known_formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S.%f",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S.%f",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ]

    for fmt in known_formats:
        remaining_mask = parsed.isna()
        if not remaining_mask.any():
            break
        parsed_part = pd.to_datetime(text_series[remaining_mask], format=fmt, errors="coerce")
        parsed.loc[remaining_mask] = parsed_part

    return parsed


def _count_unique_non_null(df: pd.DataFrame, column_name: str) -> int:
    """Count unique non-null values in one column."""
    if column_name not in df.columns:
        return 0
    return int(df[column_name].dropna().nunique())


def _count_unique_combinations(df: pd.DataFrame, columns: list[str]) -> int:
    """Count unique non-null combinations across columns."""
    existing_columns = [column for column in columns if column in df.columns]
    if not existing_columns:
        return 0

    if len(existing_columns) == 1:
        return _count_unique_non_null(df, existing_columns[0])

    work_df = df[existing_columns].dropna(subset=existing_columns).drop_duplicates(subset=existing_columns)
    return int(len(work_df))


def _build_task_base_df(combined_df: pd.DataFrame, dataset_df: pd.DataFrame) -> pd.DataFrame:
    """Build one row per Cue task with the linked saved image."""
    combined = combined_df.copy()
    dataset = dataset_df.copy()

    if "detection_id" not in combined.columns:
        raise KeyError("Combined sheet must contain 'detection_id'")
    if "detection_id" not in dataset.columns or "saved_image" not in dataset.columns:
        raise KeyError("dataset_generaton sheet must contain 'detection_id' and 'saved_image'")

    combined["cue_confirmation_time_dt"] = _parse_datetime_column(combined, "cue_confirmation_time")
    combined["cue_observation_time_dt"] = _parse_datetime_column(combined, "cue_observation_time")
    combined["tip_observation_time_dt"] = _parse_datetime_column(combined, "tip_observation_time")

    dataset_link_df = (
        dataset[["detection_id", "saved_image"]]
        .dropna(subset=["detection_id"])
        .drop_duplicates(subset=["detection_id"], keep="first")
        .copy()
    )

    task_base_df = combined.merge(
        dataset_link_df,
        on="detection_id",
        how="left",
        validate="one_to_one",
    )

    task_base_df["latency_for_benchmark_s"] = pd.to_numeric(task_base_df.get("latency_confirmation"), errors="coerce")
    if "latency_confirmation" not in task_base_df.columns:
        task_base_df["latency_for_benchmark_s"] = pd.to_numeric(task_base_df.get("latency_observation"), errors="coerce")

    task_base_df["viewing_time_s"] = pd.to_numeric(task_base_df.get("viewing_time"), errors="coerce")
    task_base_df["offnadir_deg_num"] = pd.to_numeric(task_base_df.get("offnadir_deg"), errors="coerce")

    return task_base_df


def _aggregate_detection_per_image(gt_sample_df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate image-level task success using any valid retained whale localization."""
    if gt_sample_df.empty:
        return pd.DataFrame(
            columns=[
                "image",
                "candidate_score",
                "candidate_iou",
                "candidate_confidence",
                "has_valid_localization",
                "num_gt_rows",
                "num_valid_localized_whales",
            ]
        )

    work_df = gt_sample_df.copy()

    work_df["has_valid_localization"] = work_df.get("has_valid_localization", False).apply(_to_bool)
    work_df["valid_localization_iou_num"] = pd.to_numeric(work_df.get("valid_localization_iou"), errors="coerce")
    work_df["valid_localization_score_num"] = pd.to_numeric(work_df.get("valid_localization_score"), errors="coerce")
    work_df["best_prediction_iou_num"] = pd.to_numeric(work_df.get("best_prediction_iou"), errors="coerce")
    work_df["best_prediction_score_num"] = pd.to_numeric(work_df.get("best_prediction_score"), errors="coerce")

    aggregated_rows: list[dict[str, object]] = []

    for image_name, group in work_df.groupby("image", dropna=False):
        group = group.copy()
        valid_group = group.loc[group["has_valid_localization"]].copy()

        has_valid_localization = not valid_group.empty
        num_valid_localized_whales = int(valid_group.shape[0])

        if has_valid_localization:
            valid_group["iou_sort"] = valid_group["valid_localization_iou_num"].fillna(-1.0)
            valid_group["score_sort"] = valid_group["valid_localization_score_num"].fillna(-1.0)
            best_row = valid_group.sort_values(["iou_sort", "score_sort"], ascending=[False, False]).iloc[0]
            candidate_iou = _to_float(best_row["valid_localization_iou_num"])
            candidate_confidence = _to_float(best_row["valid_localization_score_num"])
        else:
            fallback_group = group.copy()
            fallback_group["iou_sort"] = fallback_group["best_prediction_iou_num"].fillna(-1.0)
            fallback_group["score_sort"] = fallback_group["best_prediction_score_num"].fillna(-1.0)
            best_row = fallback_group.sort_values(["iou_sort", "score_sort"], ascending=[False, False]).iloc[0]
            candidate_iou = _to_float(best_row["best_prediction_iou_num"])
            candidate_confidence = _to_float(best_row["best_prediction_score_num"])

        aggregated_rows.append(
            {
                "image": image_name,
                "candidate_score": candidate_confidence,
                "candidate_confidence": candidate_confidence,
                "candidate_iou": candidate_iou,
                "has_valid_localization": has_valid_localization,
                "num_gt_rows": int(group.shape[0]),
                "num_valid_localized_whales": num_valid_localized_whales,
            }
        )

    return pd.DataFrame(aggregated_rows)


def _build_task_event_table(task_base_df: pd.DataFrame, detection_image_df: pd.DataFrame, tau_max_seconds: float | None) -> pd.DataFrame:
    """Build one row per task with task-level success based on image-level whale localization."""
    task_df = task_base_df.copy()
    detection_df = detection_image_df.copy()

    task_df = task_df.merge(
        detection_df,
        left_on="saved_image",
        right_on="image",
        how="left",
        validate="many_to_one",
    )

    task_df["candidate_confidence_num"] = pd.to_numeric(task_df.get("candidate_confidence"), errors="coerce")
    task_df["candidate_iou_num"] = pd.to_numeric(task_df.get("candidate_iou"), errors="coerce")
    task_df["has_valid_localization"] = _coerce_bool_series(task_df, "has_valid_localization", default=False)
    task_df["has_detection_candidate"] = task_df["candidate_confidence_num"].notna()

    if tau_max_seconds is None:
        task_df["passes_latency"] = True
    else:
        task_df["passes_latency"] = task_df["latency_for_benchmark_s"].notna() & (task_df["latency_for_benchmark_s"] <= float(tau_max_seconds))

    task_df["successful_detection_task"] = task_df["has_valid_localization"] & task_df["passes_latency"]
    task_df["successful_detection_task_without_latency"] = task_df["has_valid_localization"]

    preferred_columns = [
        "detection_id",
        "target_id",
        "saved_image",
        "tip_actor",
        "cue_actor",
        "tip_observation_time",
        "cue_observation_time",
        "cue_confirmation_time",
        "latency_observation",
        "latency_confirmation",
        "latency_for_benchmark_s",
        "viewing_time",
        "viewing_time_s",
        "offnadir_deg",
        "offnadir_deg_num",
        "gsd_m",
        "candidate_confidence",
        "candidate_confidence_num",
        "candidate_iou",
        "candidate_iou_num",
        "has_detection_candidate",
        "has_valid_localization",
        "num_gt_rows",
        "num_valid_localized_whales",
        "passes_latency",
        "successful_detection_task_without_latency",
        "successful_detection_task",
    ]

    existing_columns = [column for column in preferred_columns if column in task_df.columns]
    remaining_columns = [column for column in task_df.columns if column not in existing_columns]

    return task_df[existing_columns + remaining_columns].copy()


def _build_whale_event_table(task_base_df: pd.DataFrame, gt_sample_df: pd.DataFrame, tau_max_seconds: float | None) -> pd.DataFrame:
    """Build one row per GT whale inside each task image."""
    if gt_sample_df.empty:
        return pd.DataFrame(
            columns=[
                "detection_id",
                "target_id",
                "saved_image",
                "gt_ann_id",
                "has_valid_localization",
                "successful_detection_whale",
            ]
        )

    task_base = task_base_df.copy()
    gt_df = gt_sample_df.copy()

    gt_df["has_valid_localization"] = _coerce_bool_series(gt_df, "has_valid_localization", default=False)
    gt_df["valid_localization_iou_num"] = pd.to_numeric(gt_df.get("valid_localization_iou"), errors="coerce")
    gt_df["valid_localization_confidence_num"] = pd.to_numeric(gt_df.get("valid_localization_score"), errors="coerce")
    gt_df["best_prediction_iou_num"] = pd.to_numeric(gt_df.get("best_prediction_iou"), errors="coerce")
    gt_df["best_prediction_confidence_num"] = pd.to_numeric(gt_df.get("best_prediction_score"), errors="coerce")

    whale_df = task_base.merge(
        gt_df,
        left_on="saved_image",
        right_on="image",
        how="left",
    )

    whale_df = whale_df.loc[whale_df["gt_ann_id"].notna()].copy()

    if tau_max_seconds is None:
        whale_df["passes_latency"] = True
    else:
        whale_df["passes_latency"] = whale_df["latency_for_benchmark_s"].notna() & (whale_df["latency_for_benchmark_s"] <= float(tau_max_seconds))

    whale_df["successful_detection_whale"] = whale_df["has_valid_localization"] & whale_df["passes_latency"]
    whale_df["successful_detection_whale_without_latency"] = whale_df["has_valid_localization"]

    preferred_columns = [
        "detection_id",
        "target_id",
        "saved_image",
        "gt_ann_id",
        "gt_index",
        "sample_status",
        "is_fn",
        "best_prediction_iou",
        "best_prediction_score",
        "valid_localization_iou",
        "valid_localization_score",
        "valid_localization_iou_num",
        "valid_localization_confidence_num",
        "has_valid_localization",
        "latency_for_benchmark_s",
        "viewing_time_s",
        "offnadir_deg_num",
        "passes_latency",
        "successful_detection_whale_without_latency",
        "successful_detection_whale",
    ]

    existing_columns = [column for column in preferred_columns if column in whale_df.columns]
    remaining_columns = [column for column in whale_df.columns if column not in existing_columns]

    return whale_df[existing_columns + remaining_columns].copy()


def _collapse_task_event_df_to_unique_events(task_event_df: pd.DataFrame) -> pd.DataFrame:
    """Return one row per detection_id for task-based evaluation."""
    if "detection_id" not in task_event_df.columns:
        return task_event_df.copy()

    work_df = task_event_df.copy()
    work_df["candidate_confidence_num"] = pd.to_numeric(work_df.get("candidate_confidence_num"), errors="coerce")
    work_df["candidate_iou_num"] = pd.to_numeric(work_df.get("candidate_iou_num"), errors="coerce")
    work_df["successful_detection_task"] = _coerce_bool_series(work_df, "successful_detection_task", default=False)
    work_df["successful_detection_task_without_latency"] = _coerce_bool_series(
        work_df,
        "successful_detection_task_without_latency",
        default=False,
    )
    work_df["score_sort"] = work_df["candidate_confidence_num"].fillna(-1.0)
    work_df["iou_sort"] = work_df["candidate_iou_num"].fillna(-1.0)
    work_df["succ_sort"] = work_df["successful_detection_task"].astype(int)
    work_df["succ_no_latency_sort"] = work_df["successful_detection_task_without_latency"].astype(int)

    collapsed = (
        work_df.sort_values(
            ["detection_id", "succ_sort", "succ_no_latency_sort", "iou_sort", "score_sort"],
            ascending=[True, False, False, False, False],
        )
        .drop_duplicates(subset=["detection_id"], keep="first")
        .drop(columns=["score_sort", "iou_sort", "succ_sort", "succ_no_latency_sort"], errors="ignore")
        .reset_index(drop=True)
    )

    return collapsed


def _collapse_whale_event_df_to_unique_whales(whale_event_df: pd.DataFrame) -> pd.DataFrame:
    """Return one row per task-whale pair."""
    if "detection_id" not in whale_event_df.columns or "gt_ann_id" not in whale_event_df.columns:
        return whale_event_df.copy()

    work_df = whale_event_df.copy()
    work_df["valid_localization_confidence_num"] = pd.to_numeric(work_df.get("valid_localization_confidence_num"), errors="coerce")
    work_df["valid_localization_iou_num"] = pd.to_numeric(work_df.get("valid_localization_iou_num"), errors="coerce")
    work_df["successful_detection_whale"] = _coerce_bool_series(work_df, "successful_detection_whale", default=False)
    work_df["successful_detection_whale_without_latency"] = _coerce_bool_series(
        work_df,
        "successful_detection_whale_without_latency",
        default=False,
    )
    work_df["score_sort"] = work_df["valid_localization_confidence_num"].fillna(-1.0)
    work_df["iou_sort"] = work_df["valid_localization_iou_num"].fillna(-1.0)
    work_df["succ_sort"] = work_df["successful_detection_whale"].astype(int)
    work_df["succ_no_latency_sort"] = work_df["successful_detection_whale_without_latency"].astype(int)

    collapsed = (
        work_df.sort_values(
            ["detection_id", "gt_ann_id", "succ_sort", "succ_no_latency_sort", "iou_sort", "score_sort"],
            ascending=[True, True, False, False, False, False],
        )
        .drop_duplicates(subset=["detection_id", "gt_ann_id"], keep="first")
        .drop(columns=["score_sort", "iou_sort", "succ_sort", "succ_no_latency_sort"], errors="ignore")
        .reset_index(drop=True)
    )

    return collapsed


def _extract_run_summary_value(run_summary_df: pd.DataFrame, *aliases: str) -> object:
    """Extract one run_summary value by trying multiple aliases."""
    if run_summary_df.empty:
        return None

    row = run_summary_df.iloc[0]
    normalized_columns = {_normalize_name(column): column for column in run_summary_df.columns}

    for alias in aliases:
        key = _normalize_name(alias)
        if key in normalized_columns:
            return row[normalized_columns[key]]

    return None


def _extract_detection_metric(overall_detection_stats_df: pd.DataFrame, *metric_names: str) -> float | None:
    """Extract one metric from overall_detection_stats using several possible names."""
    for metric_name in metric_names:
        value = _extract_metric_from_long_table(overall_detection_stats_df, metric_name)
        if value is not None:
            return _to_float(value)
    return None


def _count_cue_tasks_received(task_event_unique_df: pd.DataFrame, fallback_count: int) -> int:
    """Count Cue tasks received from task table."""
    if "detection_id" not in task_event_unique_df.columns:
        return fallback_count

    if "cue_actor" in task_event_unique_df.columns:
        received_df = task_event_unique_df.loc[task_event_unique_df["cue_actor"].notna()].copy()
        return int(received_df["detection_id"].dropna().nunique())

    return fallback_count


def _count_cue_tasks_handled(task_event_unique_df: pd.DataFrame, fallback_count: int) -> int:
    """Count Cue tasks handled from task table."""
    if "detection_id" not in task_event_unique_df.columns:
        return fallback_count

    handled_mask = pd.Series(False, index=task_event_unique_df.index)

    for column_name in ["saved_image", "cue_observation_time", "cue_confirmation_time"]:
        if column_name in task_event_unique_df.columns:
            handled_mask = handled_mask | task_event_unique_df[column_name].notna()

    if handled_mask.any():
        handled_df = task_event_unique_df.loc[handled_mask].copy()
        return int(handled_df["detection_id"].dropna().nunique())

    return fallback_count


def _build_benchmark_overview(
    case_name: str,
    variant_label: str,
    mission_workbook: Path,
    detection_results_workbook: Path,
    detection_per_sample_workbook: Path,
    overview_df: pd.DataFrame,
    combined_df: pd.DataFrame,
    run_summary_df: pd.DataFrame,
    overall_detection_stats_df: pd.DataFrame,
    task_event_df: pd.DataFrame,
    whale_event_df: pd.DataFrame,
    tau_max_seconds: float | None,
) -> pd.DataFrame:
    """Build tidy benchmark overview rows with mission, task, whale, and geometry metrics."""
    n_sat_from_name = _extract_total_satellites_from_case_name(case_name)
    n_sat_from_overview = _to_float(_extract_overview_value(overview_df, "Number of satellites"))
    n_sat = float(n_sat_from_name) if n_sat_from_name is not None else n_sat_from_overview
    t_sim_hours = _to_float(_extract_overview_value(overview_df, "Simulation time (h)"))
    n_gt_overview = _to_float(_extract_overview_value(overview_df, "Total targets"))

    task_event_unique_df = _collapse_task_event_df_to_unique_events(task_event_df)
    whale_event_unique_df = _collapse_whale_event_df_to_unique_whales(whale_event_df)

    n_mission_task = _count_unique_non_null(combined_df, "detection_id")
    if "detection_id" in combined_df.columns and "target_id" in combined_df.columns:
        n_mission = _count_unique_combinations(combined_df, ["detection_id", "target_id"])
    elif "detection_id" in combined_df.columns:
        n_mission = _count_unique_non_null(combined_df, "detection_id")
    else:
        n_mission = int(len(combined_df))

    n_cue_task_received = _count_cue_tasks_received(task_event_unique_df, fallback_count=n_mission_task)
    n_cue_task_handled = _count_cue_tasks_handled(task_event_unique_df, fallback_count=n_mission_task)

    n_task = int(task_event_unique_df["detection_id"].nunique()) if "detection_id" in task_event_unique_df.columns else int(len(task_event_unique_df))
    n_succ_task = int(task_event_unique_df["successful_detection_task"].fillna(False).sum()) if "successful_detection_task" in task_event_unique_df.columns else 0
    n_succ_task_no_latency = int(task_event_unique_df["successful_detection_task_without_latency"].fillna(False).sum()) if "successful_detection_task_without_latency" in task_event_unique_df.columns else 0

    n_gt = int(len(whale_event_unique_df))
    n_succ = int(whale_event_unique_df["successful_detection_whale"].fillna(False).sum()) if "successful_detection_whale" in whale_event_unique_df.columns else 0
    n_succ_no_latency = int(whale_event_unique_df["successful_detection_whale_without_latency"].fillna(False).sum()) if "successful_detection_whale_without_latency" in whale_event_unique_df.columns else 0

    c_mission_task = None
    c_mission = None
    c_cue_task_received = None
    c_cue_task_handled = None
    c_succ = None
    c_succ_task = None

    if n_sat is not None and t_sim_hours is not None and n_sat > 0:
        t_sim_days = t_sim_hours / 24.0
        if t_sim_days > 0:
            c_mission_task = n_mission_task / (n_sat * t_sim_days)
            c_mission = n_mission / (n_sat * t_sim_days)
            c_cue_task_received = n_cue_task_received / (n_sat * t_sim_days)
            c_cue_task_handled = n_cue_task_handled / (n_sat * t_sim_days)
            c_succ = n_succ / (n_sat * t_sim_days)
            c_succ_task = n_succ_task / (n_sat * t_sim_days)

    offnadir_observed_mean_deg = _to_float(task_event_unique_df["offnadir_deg_num"].mean()) if "offnadir_deg_num" in task_event_unique_df.columns else None

    l_mean_success = _to_float(
        whale_event_unique_df.loc[whale_event_unique_df["successful_detection_whale"], "latency_for_benchmark_s"].mean()
    ) if "successful_detection_whale" in whale_event_unique_df.columns else None
    v_mean_success = _to_float(
        whale_event_unique_df.loc[whale_event_unique_df["successful_detection_whale"], "viewing_time_s"].mean()
    ) if "successful_detection_whale" in whale_event_unique_df.columns else None
    iou_mean_success = _to_float(
        whale_event_unique_df.loc[whale_event_unique_df["successful_detection_whale"], "valid_localization_iou_num"].mean()
    ) if "successful_detection_whale" in whale_event_unique_df.columns else None
    q_mean_success = _to_float(
        whale_event_unique_df.loc[whale_event_unique_df["successful_detection_whale"], "valid_localization_confidence_num"].mean()
    ) if "successful_detection_whale" in whale_event_unique_df.columns else None

    l_mean_task_success = _to_float(
        task_event_unique_df.loc[task_event_unique_df["successful_detection_task"], "latency_for_benchmark_s"].mean()
    ) if "successful_detection_task" in task_event_unique_df.columns else None
    v_mean_task_success = _to_float(
        task_event_unique_df.loc[task_event_unique_df["successful_detection_task"], "viewing_time_s"].mean()
    ) if "successful_detection_task" in task_event_unique_df.columns else None
    iou_mean_task_success = _to_float(
        task_event_unique_df.loc[task_event_unique_df["successful_detection_task"], "candidate_iou_num"].mean()
    ) if "successful_detection_task" in task_event_unique_df.columns else None
    q_mean_task_success = _to_float(
        task_event_unique_df.loc[task_event_unique_df["successful_detection_task"], "candidate_confidence_num"].mean()
    ) if "successful_detection_task" in task_event_unique_df.columns else None

    coco_ap50 = _to_float(_extract_run_summary_value(run_summary_df, "coco_ap50", "coco ap50"))
    coco_ap50_95 = _to_float(_extract_run_summary_value(run_summary_df, "coco_ap50_95", "coco ap50 95", "coco_ap50:95"))
    detector_score_threshold = _to_float(_extract_run_summary_value(run_summary_df, "individual_score_threshold", "individual score threshold"))
    detector_iou_threshold = _to_float(_extract_run_summary_value(run_summary_df, "individual_iou_threshold", "individual iou threshold"))
    avg_best_iou = _to_float(
        _extract_run_summary_value(
            run_summary_df,
            "avg_best_iou",
            "avg best iou",
            "average_best_iou",
            "average best iou",
            "avg_best_iou_positive_samples",
            "avg best iou positive samples",
            "avg_best_iou_all_samples_negative_zero",
            "avg best iou all samples negative zero",
        )
    )
    avg_best_confidence = _to_float(
        _extract_run_summary_value(
            run_summary_df,
            "avg_best_confidence",
            "avg best confidence",
            "average_best_confidence",
            "average best confidence",
            "avg_best_confidence_positive_samples",
            "avg best confidence positive samples",
            "avg_best_confidence_all_samples_negative_zero",
            "avg best confidence all samples negative zero",
        )
    )

    detector_precision = _extract_detection_metric(overall_detection_stats_df, "overall_precision", "precision")
    detector_recall = _extract_detection_metric(overall_detection_stats_df, "overall_recall", "recall")
    detector_f1 = _extract_detection_metric(overall_detection_stats_df, "overall_f1", "f1")

    rows = [
        {"metric": "case_name", "value": case_name, "comment": None},
        {"metric": "variant_label", "value": variant_label, "comment": "default / location-specific"},
        {"metric": "mission_results_workbook", "value": str(mission_workbook), "comment": None},
        {"metric": "detection_results_workbook", "value": str(detection_results_workbook), "comment": None},
        {"metric": "detection_per_sample_workbook", "value": str(detection_per_sample_workbook), "comment": None},
        {"metric": "tau_max_seconds", "value": tau_max_seconds, "comment": "None means no latency hard-filter"},
        {"metric": "detector_score_threshold", "value": detector_score_threshold, "comment": "Fixed deployed detector threshold"},
        {"metric": "detector_iou_threshold", "value": detector_iou_threshold, "comment": "IoU threshold used for valid localization"},
        {"metric": "N_sat", "value": n_sat, "comment": "Number of satellites"},
        {"metric": "T_sim_hours", "value": t_sim_hours, "comment": "Simulation time in hours"},
        {"metric": "N_gt_overview", "value": n_gt_overview, "comment": "Total targets from mission overview"},
        {"metric": "N_mission_task", "value": n_mission_task, "comment": "Total unique mission task events from simulation results"},
        {"metric": "C_mission_task", "value": c_mission_task, "comment": "Mission task events per satellite per simulated day"},
        {"metric": "N_mission", "value": n_mission, "comment": "Total mission whale detections from simulation results"},
        {"metric": "C_mission", "value": c_mission, "comment": "Mission whale detections per satellite per simulated day"},
        {"metric": "N_cue_task_received", "value": n_cue_task_received, "comment": "Cue tasks received by Cue satellites"},
        {"metric": "C_cue_task_received", "value": c_cue_task_received, "comment": "Cue tasks received per satellite per simulated day"},
        {"metric": "N_cue_task_handled", "value": n_cue_task_handled, "comment": "Cue tasks handled by Cue satellites before correctness filtering"},
        {"metric": "C_cue_task_handled", "value": c_cue_task_handled, "comment": "Cue tasks handled per satellite per simulated day"},
        {"metric": "N_gt", "value": n_gt, "comment": "Total GT whales present in all evaluated Cue images"},
        {"metric": "N_succ_detection_only", "value": n_succ_no_latency, "comment": "GT whales with at least one valid retained localization, ignoring latency"},
        {"metric": "N_succ", "value": n_succ, "comment": "GT whales with at least one valid retained localization and latency filter passes"},
        {"metric": "C_succ", "value": c_succ, "comment": "Successful whale detections per satellite per simulated day"},
        {"metric": "L_mean_success_s", "value": l_mean_success, "comment": "Mean latency over successful whale detections"},
        {"metric": "V_mean_success_s", "value": v_mean_success, "comment": "Mean viewing time over successful whale detections"},
        {"metric": "IoU_mean_success", "value": iou_mean_success, "comment": "Mean IoU over successful whale detections"},
        {"metric": "Q_mean_success", "value": q_mean_success, "comment": "Mean confidence over successful whale detections"},
        {"metric": "N_task", "value": n_task, "comment": "Number of Cue tasks represented in the task_event table"},
        {"metric": "offnadir_observed_mean_deg", "value": offnadir_observed_mean_deg, "comment": "Mean observed off-nadir angle over unique Cue task events"},
        {"metric": "N_succ_task_detection_only", "value": n_succ_task_no_latency, "comment": "Tasks with at least one valid retained whale localization in the task image, ignoring latency"},
        {"metric": "N_succ_task", "value": n_succ_task, "comment": "Tasks with at least one valid retained whale localization in the task image and latency filter passes"},
        {"metric": "C_succ_task", "value": c_succ_task, "comment": "Successful task confirmations per satellite per simulated day"},
        {"metric": "L_mean_task_success_s", "value": l_mean_task_success, "comment": "Mean latency over successful tasks"},
        {"metric": "V_mean_task_success_s", "value": v_mean_task_success, "comment": "Mean viewing time over successful tasks"},
        {"metric": "IoU_mean_task_success", "value": iou_mean_task_success, "comment": "Mean representative IoU over successful tasks"},
        {"metric": "Q_mean_task_success", "value": q_mean_task_success, "comment": "Mean representative confidence over successful tasks"},
        {"metric": "coco_ap50", "value": coco_ap50, "comment": "Detector metric at IoU 0.50"},
        {"metric": "coco_ap50_95", "value": coco_ap50_95, "comment": "Detector metric across IoU thresholds 0.50:0.95"},
        {"metric": "detector_precision", "value": detector_precision, "comment": "Detector precision at the deployed operating threshold"},
        {"metric": "detector_recall", "value": detector_recall, "comment": "Detector recall at the deployed operating threshold"},
        {"metric": "detector_f1", "value": detector_f1, "comment": "Detector F1 at the deployed operating threshold"},
        {"metric": "avg_best_iou", "value": avg_best_iou, "comment": "Detection workbook run_summary average best IoU"},
        {"metric": "avg_best_confidence", "value": avg_best_confidence, "comment": "Detection workbook run_summary average best confidence"},
    ]

    return pd.DataFrame(rows)


def _normalize_excel_value(value: object) -> object:
    """Convert values to Excel-safe entries."""
    if isinstance(value, (list, dict, tuple, set)):
        return str(value)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if pd.isna(value):
        return None
    return value


def _write_dataframe_sheet(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    """Write one dataframe to one worksheet."""
    worksheet = workbook.create_sheet(title=sheet_name[:31])

    if dataframe.empty:
        worksheet.append(["no_data"])
        return

    headers = list(dataframe.columns)
    worksheet.append(headers)

    for _, row in dataframe.iterrows():
        worksheet.append([_normalize_excel_value(row.get(header)) for header in headers])

    for column_index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_index in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            cell_length = len(str(cell_value)) if cell_value is not None else 0
            if cell_length > max_length:
                max_length = cell_length
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)


def _write_benchmark_workbook(
    benchmark_path: Path,
    benchmark_overview_df: pd.DataFrame,
    whale_event_df: pd.DataFrame,
    whale_event_unique_df: pd.DataFrame,
    task_event_df: pd.DataFrame,
    task_event_unique_df: pd.DataFrame,
    mission_overview_df: pd.DataFrame,
    combined_df: pd.DataFrame,
    dataset_df: pd.DataFrame,
    detection_run_summary_df: pd.DataFrame,
    overall_detection_stats_df: pd.DataFrame,
    detection_image_summary_df: pd.DataFrame | None,
    gt_sample_df: pd.DataFrame,
    overwrite_results: bool,
) -> bool:
    """Write benchmark workbook with overview and source sheets."""
    if benchmark_path.exists() and not overwrite_results:
        print(f"[SKIP] Benchmark already exists and overwrite_results=False: {benchmark_path}")
        return False

    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_dataframe_sheet(workbook, "benchmark_overview", benchmark_overview_df)
    _write_dataframe_sheet(workbook, "whale_event_details", whale_event_df)
    _write_dataframe_sheet(workbook, "whale_event_details_unique", whale_event_unique_df)
    _write_dataframe_sheet(workbook, "task_event_details", task_event_df)
    _write_dataframe_sheet(workbook, "task_event_details_unique", task_event_unique_df)
    _write_dataframe_sheet(workbook, "mission_overview_src", mission_overview_df)
    _write_dataframe_sheet(workbook, "mission_combined_src", combined_df)
    _write_dataframe_sheet(workbook, "dataset_generation_src", dataset_df)
    _write_dataframe_sheet(workbook, "detection_run_summary", detection_run_summary_df)
    _write_dataframe_sheet(workbook, "overall_detection_stats", overall_detection_stats_df)

    if detection_image_summary_df is not None:
        _write_dataframe_sheet(workbook, "image_summary_src", detection_image_summary_df)

    _write_dataframe_sheet(workbook, "gt_sample_summary_src", gt_sample_df)

    benchmark_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(benchmark_path)
    return True


def process_case_variant(
    case_dir: Path,
    variant_label: str,
    detection_dir: Path,
    benchmark_filename: str,
    overwrite_results: bool,
    skip_existing_results: bool,
    tau_max_seconds: float | None,
) -> bool:
    """Process one case and one variant into one benchmark workbook."""
    benchmark_path = case_dir / benchmark_filename

    if skip_existing_results and _has_existing_benchmark(benchmark_path):
        print(f"[SKIP] Existing benchmark found: {benchmark_path}")
        return False

    mission_workbook = _find_mission_results_workbook(case_dir)
    detection_results_workbook = _find_required_workbook(detection_dir / "onboard_detection_results.xlsx")
    detection_per_sample_workbook = _find_required_workbook(detection_dir / "onboard_detection_per_sample.xlsx")

    overview_df = _read_sheet_with_aliases(mission_workbook, ["Overview"])
    combined_df = _read_sheet_with_aliases(mission_workbook, ["Combined"])
    dataset_df = _read_sheet_with_aliases(mission_workbook, ["dataset_generaton", "dataset_generation"])

    detection_run_summary_df = _read_sheet_with_aliases(detection_results_workbook, ["run_summary"])
    overall_detection_stats_df = _read_sheet_with_aliases(detection_results_workbook, ["overall_detection_stats"])
    gt_sample_df = _read_sheet_with_aliases(detection_per_sample_workbook, ["gt_sample_summary"])

    detection_image_summary_df: pd.DataFrame | None
    try:
        detection_image_summary_df = _read_sheet_with_aliases(detection_results_workbook, ["image_summary"])
    except Exception:
        detection_image_summary_df = None

    task_base_df = _build_task_base_df(combined_df=combined_df, dataset_df=dataset_df)
    detection_image_df = _aggregate_detection_per_image(gt_sample_df)

    task_event_df = _build_task_event_table(
        task_base_df=task_base_df,
        detection_image_df=detection_image_df,
        tau_max_seconds=tau_max_seconds,
    )
    task_event_unique_df = _collapse_task_event_df_to_unique_events(task_event_df)

    whale_event_df = _build_whale_event_table(
        task_base_df=task_base_df,
        gt_sample_df=gt_sample_df,
        tau_max_seconds=tau_max_seconds,
    )
    whale_event_unique_df = _collapse_whale_event_df_to_unique_whales(whale_event_df)

    benchmark_overview_df = _build_benchmark_overview(
        case_name=case_dir.name,
        variant_label=variant_label,
        mission_workbook=mission_workbook,
        detection_results_workbook=detection_results_workbook,
        detection_per_sample_workbook=detection_per_sample_workbook,
        overview_df=overview_df,
        combined_df=combined_df,
        run_summary_df=detection_run_summary_df,
        overall_detection_stats_df=overall_detection_stats_df,
        task_event_df=task_event_df,
        whale_event_df=whale_event_df,
        tau_max_seconds=tau_max_seconds,
    )

    written = _write_benchmark_workbook(
        benchmark_path=benchmark_path,
        benchmark_overview_df=benchmark_overview_df,
        whale_event_df=whale_event_df,
        whale_event_unique_df=whale_event_unique_df,
        task_event_df=task_event_df,
        task_event_unique_df=task_event_unique_df,
        mission_overview_df=overview_df,
        combined_df=combined_df,
        dataset_df=dataset_df,
        detection_run_summary_df=detection_run_summary_df,
        overall_detection_stats_df=overall_detection_stats_df,
        detection_image_summary_df=detection_image_summary_df,
        gt_sample_df=gt_sample_df,
        overwrite_results=overwrite_results,
    )

    if written:
        print(f"[OK] Wrote benchmark: {benchmark_path}")

    return written


def main() -> None:
    """Process benchmark workbooks across all cases using selected mode."""
    script_dir = Path(__file__).resolve().parent
    master_dir = script_dir.parent

    master_results_list = [
        "reflection_offnadir_glint_255",
        "reflection_nadir_glint_255",
        "texture_offnadir_255",
        "texture_nadir_255",
    ]
    mode = "all"

    distinct_locations = ["Auckland2006", "Pelagos2016"]
    overwrite_results = True
    skip_existing_results = True
    tau_max_seconds: float | None = None

    for master_results in master_results_list:
        print(f"\n=============== Start processing {master_results} ===============")

        master_results = "EXPERIMENTS/" + master_results
        master_results_root = master_dir / "0_results" / master_results

        if not master_results_root.exists():
            raise FileNotFoundError(f"Master results root does not exist: {master_results_root}")

        case_dirs = _find_case_directories(master_results_root)
        if not case_dirs:
            raise FileNotFoundError(f"No case directories found under {master_results_root}")

        all_jobs: list[dict[str, str | Path]] = []
        for case_dir in case_dirs:
            for job in _resolve_variant_jobs(case_dir=case_dir, mode=mode, distinct_locations=distinct_locations):
                all_jobs.append(
                    {
                        "case_dir": case_dir,
                        "variant_label": str(job["label"]),
                        "detection_dir": Path(job["detection_dir"]),
                        "benchmark_filename": str(job["benchmark_filename"]),
                    }
                )

        if not all_jobs:
            raise ValueError("No jobs were resolved. Check mode and distinct_locations.")

        print(f"Found {len(case_dirs)} case folders.")
        print(f"Resolved {len(all_jobs)} benchmark jobs.")
        print(f"Mode: {mode}")
        print(f"overwrite_results: {overwrite_results}")
        print(f"skip_existing_results: {skip_existing_results}")
        print(f"tau_max_seconds: {tau_max_seconds}")
        print()

        processed = 0
        skipped = 0
        failed = 0

        for index, job in enumerate(all_jobs, start=1):
            case_dir = Path(job["case_dir"])
            variant_label = str(job["variant_label"])
            detection_dir = Path(job["detection_dir"])
            benchmark_filename = str(job["benchmark_filename"])

            print(f"\n[{index}/{len(all_jobs)}] Case: {case_dir.name} | Variant: {variant_label}")

            try:
                was_written = process_case_variant(
                    case_dir=case_dir,
                    variant_label=variant_label,
                    detection_dir=detection_dir,
                    benchmark_filename=benchmark_filename,
                    overwrite_results=overwrite_results,
                    skip_existing_results=skip_existing_results,
                    tau_max_seconds=tau_max_seconds,
                )
                if was_written:
                    processed += 1
                else:
                    skipped += 1
            except FileNotFoundError as exc:
                skipped += 1
                print(f"[SKIP] {case_dir.name} | {variant_label}: {exc}")
            except Exception as exc:
                failed += 1
                print(f"[FAIL] {case_dir.name} | {variant_label}: {exc}")

        print("\n" + "=" * 100)
        print("Finished processing benchmark workbooks.")
        print(f"Processed successfully: {processed}")
        print(f"Skipped: {skipped}")
        print(f"Failed: {failed}")
        print("=" * 100)


if __name__ == "__main__":
    main()