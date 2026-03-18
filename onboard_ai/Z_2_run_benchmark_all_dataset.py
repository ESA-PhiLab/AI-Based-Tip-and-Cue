from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


CASE_COLUMN = "case_name"

EXTRACTED_COLUMNS = [
    "N_sats_total",
    "offnadir_angle_deg",
    "time_delay_min",
]

METRIC_COLUMNS_FIRST = [
    "C_succ",
    "E_e2e",
    "L_mean_success_s",
    "V_mean_success_s",
    "Q_mean_success",
    "theta_mean_success_deg",
    "overall_f1_detection",
    "coco_ap50",
    "coco_ap50_95",
]

METRIC_COLUMNS_SECOND = [
    "N_sat",
    "T_sim_hours",
    "N_gt",
    "N_obs",
    "N_positive_obs",
    "N_succ_detection_only",
    "N_succ",
]

ALL_METRIC_COLUMNS = METRIC_COLUMNS_FIRST + METRIC_COLUMNS_SECOND

FINAL_METRIC_COLUMNS_FIRST = [
    "C_succ_sat",
    "C_succ_overall",
    "E_e2e",
    "L_mean_success_s",
    "V_mean_success_s",
    "Q_mean_success",
    "theta_mean_success_deg",
    "overall_f1_detection",
    "coco_ap50",
    "coco_ap50_95",
]

FINAL_METRIC_COLUMNS_SECOND = [
    "N_sat",
    "T_sim_hours",
    "N_gt",
    "N_obs",
    "N_positive_obs",
    "N_succ_detection_only",
    "N_succ",
]

FINAL_OUTPUT_METRIC_COLUMNS = FINAL_METRIC_COLUMNS_FIRST + FINAL_METRIC_COLUMNS_SECOND

ALL_CASES_OUTPUT_COLUMNS = (
    [CASE_COLUMN]
    + EXTRACTED_COLUMNS
    + FINAL_OUTPUT_METRIC_COLUMNS
    + ["n_images", "source_file"]
)


def _find_case_directories(final_results_root: Path) -> list[Path]:
    """Return all case directories inside the final results folder."""
    return sorted(path for path in final_results_root.iterdir() if path.is_dir())


def _resolve_overview_jobs(mode: str, distinct_locations: list[str]) -> list[dict[str, str]]:
    """Resolve which overview files should be written."""
    if mode not in {"random", "distinct", "all"}:
        raise ValueError("mode must be one of: 'random', 'distinct', 'all'")

    jobs: list[dict[str, str]] = []

    if mode in {"random", "all"}:
        jobs.append(
            {
                "label": "default",
                "output_filename": "overview_random.xlsx",
            }
        )

    if mode in {"distinct", "all"}:
        for location in distinct_locations:
            jobs.append(
                {
                    "label": location,
                    "output_filename": f"overview_{location}.xlsx",
                }
            )

    return jobs


def _find_matching_benchmark_workbooks(case_dir: Path, label: str) -> list[Path]:
    """Find benchmark workbooks matching one overview mode label."""
    all_benchmarks = sorted(
        path for path in case_dir.glob("benchmark*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    )

    if label == "default":
        return [
            path for path in all_benchmarks
            if not path.name.startswith("benchmark_Auckland2006_")
            and not path.name.startswith("benchmark_Pelagos2016_")
        ]

    return [path for path in all_benchmarks if path.name.startswith(f"benchmark_{label}_")]


def _read_benchmark_overview_sheet(xlsx_path: Path) -> pd.DataFrame:
    """Read benchmark_overview sheet and convert metric/value rows into one row."""
    df = pd.read_excel(xlsx_path, sheet_name="benchmark_overview")

    required_columns = {"metric", "value"}
    missing = required_columns - set(df.columns)
    if missing:
        raise KeyError(f"{xlsx_path} benchmark_overview is missing columns: {sorted(missing)}")

    row_dict: dict[str, object] = {"source_file": str(xlsx_path)}
    for _, row in df.iterrows():
        metric_name = str(row["metric"])
        metric_value = row["value"]
        row_dict[metric_name] = metric_value

    return pd.DataFrame([row_dict])


def _extract_seed_group(case_name: str) -> str:
    """Remove trailing seed suffix like _1sd, _17sd, _42sd from a case name."""
    return re.sub(r"_\d+sd$", "", str(case_name))


def _extract_case_parameters(case_name: str) -> dict[str, object]:
    """Extract satellite count, off-nadir angle, and delay from grouped case name."""
    case_name = str(case_name)

    tc_match = re.match(r"^TC_(\d+)x(\d+)sat_(\d+)deg_(\d+)min$", case_name)
    if tc_match:
        n_orbits = int(tc_match.group(1))
        n_sats_per_orbit = int(tc_match.group(2))
        offnadir_angle_deg = int(tc_match.group(3))
        time_delay_min = int(tc_match.group(4))
        return {
            "N_sats_total": n_orbits * n_sats_per_orbit,
            "offnadir_angle_deg": offnadir_angle_deg,
            "time_delay_min": time_delay_min,
        }

    c_match = re.match(r"^C_(\d+)x(\d+)sat$", case_name)
    if c_match:
        n_orbits = int(c_match.group(1))
        n_sats_per_orbit = int(c_match.group(2))
        return {
            "N_sats_total": n_orbits * n_sats_per_orbit,
            "offnadir_angle_deg": None,
            "time_delay_min": None,
        }

    return {
        "N_sats_total": None,
        "offnadir_angle_deg": None,
        "time_delay_min": None,
    }


def _count_images_in_dir(image_dir: Path) -> int | None:
    """Count images in one image folder."""
    if not image_dir.exists():
        return None

    return sum(
        1
        for f in image_dir.iterdir()
        if f.is_file() and f.suffix.lower() in {".png", ".jpg", ".jpeg", ".tif", ".tiff"}
    )


def _build_case_group_to_image_count(case_dirs: list[Path], label: str) -> dict[str, int | None]:
    """Build mapping from grouped case name to image count for the selected label."""
    image_counts: dict[str, int | None] = {}

    for case_dir in case_dirs:
        case_group = _extract_seed_group(case_dir.name)

        if label == "default":
            image_dir = case_dir / "satellite_images"
        else:
            image_dir = case_dir / f"satellite_images_{label}"

        image_count = _count_images_in_dir(image_dir)

        if case_group not in image_counts:
            image_counts[case_group] = image_count
        else:
            existing = image_counts[case_group]
            if existing is None and image_count is not None:
                image_counts[case_group] = image_count

    return image_counts


def _add_extracted_columns(df: pd.DataFrame, source_case_column: str, image_count_map: dict[str, int | None]) -> pd.DataFrame:
    """Add extracted numerical columns next to the case column."""
    output = df.copy()
    extracted_series = output[source_case_column].astype(str).apply(_extract_case_parameters)
    extracted_df = pd.DataFrame(list(extracted_series), index=output.index)

    for column in EXTRACTED_COLUMNS:
        output[column] = extracted_df[column] if column in extracted_df.columns else None

    output["n_images"] = output[source_case_column].astype(str).map(image_count_map)

    return output


def _safe_numeric_convert(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Convert selected columns to numeric where possible."""
    output = df.copy()
    for column in columns:
        if column in output.columns:
            output[column] = pd.to_numeric(output[column], errors="coerce")
    return output


def _add_capacity_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Split C_succ into per-satellite and overall mission capacity."""
    output = df.copy()

    if "C_succ" not in output.columns:
        output["C_succ"] = pd.NA
    if "N_sats_total" not in output.columns:
        output["N_sats_total"] = pd.NA

    output["C_succ_sat"] = pd.to_numeric(output["C_succ"], errors="coerce")
    output["C_succ_overall"] = (
        pd.to_numeric(output["C_succ_sat"], errors="coerce")
        * pd.to_numeric(output["N_sats_total"], errors="coerce")
    )

    return output


def _ensure_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Add missing columns with None and reorder."""
    output = df.copy()
    for column in columns:
        if column not in output.columns:
            output[column] = None
    return output[columns]


def _excel_safe_value(value: object) -> object:
    """Convert pandas/openpyxl-unfriendly missing values to None."""
    if value is None:
        return None
    if pd.isna(value):
        return None
    return value


def _write_dataframe_sheet(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    """Write one dataframe to one worksheet."""
    worksheet = workbook.create_sheet(title=sheet_name[:31])

    if dataframe.empty:
        worksheet.append(["no_data"])
        return

    headers = list(dataframe.columns)
    worksheet.append(headers)

    for _, row in dataframe.iterrows():
        worksheet.append([_excel_safe_value(row.get(header)) for header in headers])

    for column_index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_index in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            cell_length = len(str(cell_value)) if cell_value is not None else 0
            if cell_length > max_length:
                max_length = cell_length
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)


def _write_overview_workbook(output_path: Path, all_cases_df: pd.DataFrame, grouped_mean_df: pd.DataFrame, overwrite_results: bool) -> None:
    """Write overview workbook with all cases and grouped means."""
    if output_path.exists() and not overwrite_results:
        print(f"[SKIP] Overview workbook already exists and overwrite_results=False: {output_path}")
        return

    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_dataframe_sheet(workbook, "all_cases", all_cases_df)
    _write_dataframe_sheet(workbook, "grouped_mean", grouped_mean_df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _build_overview_tables(row_dfs: list[pd.DataFrame], image_count_map: dict[str, int | None]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build all_cases and grouped_mean tables."""
    if not row_dfs:
        empty_all = pd.DataFrame(columns=["case_group"] + ALL_CASES_OUTPUT_COLUMNS)
        empty_grouped = pd.DataFrame(columns=[CASE_COLUMN] + EXTRACTED_COLUMNS + FINAL_OUTPUT_METRIC_COLUMNS + ["n_images"])
        return empty_all, empty_grouped

    all_cases_df = pd.concat(row_dfs, ignore_index=True)

    if CASE_COLUMN not in all_cases_df.columns:
        raise KeyError(f"Required column '{CASE_COLUMN}' was not found in benchmark_overview data.")

    all_cases_df = _safe_numeric_convert(all_cases_df, ALL_METRIC_COLUMNS)
    all_cases_df["case_group"] = all_cases_df[CASE_COLUMN].astype(str).apply(_extract_seed_group)
    all_cases_df = _add_extracted_columns(all_cases_df, source_case_column="case_group", image_count_map=image_count_map)
    all_cases_df = _add_capacity_columns(all_cases_df)

    all_cases_output_columns = ["case_group"] + ALL_CASES_OUTPUT_COLUMNS
    all_cases_df = _ensure_columns(all_cases_df, all_cases_output_columns)
    all_cases_df = all_cases_df.sort_values(["case_group", CASE_COLUMN]).reset_index(drop=True)

    grouped_mean_source_columns = [
        "N_sats_total",
        "offnadir_angle_deg",
        "time_delay_min",
        "n_images",
        "C_succ_sat",
        "C_succ_overall",
        "E_e2e",
        "L_mean_success_s",
        "V_mean_success_s",
        "Q_mean_success",
        "theta_mean_success_deg",
        "overall_f1_detection",
        "coco_ap50",
        "coco_ap50_95",
        "N_sat",
        "T_sim_hours",
        "N_gt",
        "N_obs",
        "N_positive_obs",
        "N_succ_detection_only",
        "N_succ",
    ]

    grouped_mean_df = (
        all_cases_df.groupby("case_group", dropna=False)[grouped_mean_source_columns]
        .mean(numeric_only=True)
        .reset_index()
    )
    grouped_mean_df = grouped_mean_df.rename(columns={"case_group": CASE_COLUMN})
    grouped_mean_df = _ensure_columns(
        grouped_mean_df,
        [CASE_COLUMN] + EXTRACTED_COLUMNS + FINAL_OUTPUT_METRIC_COLUMNS + ["n_images"],
    )
    grouped_mean_df = grouped_mean_df.sort_values([CASE_COLUMN]).reset_index(drop=True)

    return all_cases_df, grouped_mean_df


def main() -> None:
    """Create separate overview workbooks according to mode."""
    script_dir = Path(__file__).resolve().parent
    master_dir = script_dir.parent

    final_results_folder_name = "EXPERIMENTS/texture_offnadir_255"
    mode = "all"
    distinct_locations = ["Auckland2006", "Pelagos2016"]
    overwrite_results = True

    final_results_root = master_dir / "0_results" / final_results_folder_name

    if not final_results_root.exists():
        raise FileNotFoundError(f"Final results root does not exist: {final_results_root}")

    case_dirs = _find_case_directories(final_results_root)
    if not case_dirs:
        raise FileNotFoundError(f"No case directories found under {final_results_root}")

    overview_jobs = _resolve_overview_jobs(mode=mode, distinct_locations=distinct_locations)

    if not overview_jobs:
        raise ValueError("No overview jobs were resolved. Check mode and distinct_locations.")

    print(f"Found {len(case_dirs)} case directories.")
    print(f"Resolved {len(overview_jobs)} overview output job(s).")
    print()

    for job in overview_jobs:
        label = job["label"]
        output_filename = job["output_filename"]

        print(f"Building overview for label: {label}")

        row_dfs: list[pd.DataFrame] = []
        image_count_map = _build_case_group_to_image_count(case_dirs=case_dirs, label=label)

        for case_dir in case_dirs:
            benchmark_files = _find_matching_benchmark_workbooks(case_dir=case_dir, label=label)

            if not benchmark_files:
                print(f"[SKIP] No matching benchmark files for {label} in {case_dir.name}")
                continue

            for benchmark_file in benchmark_files:
                try:
                    row_df = _read_benchmark_overview_sheet(benchmark_file)
                    row_dfs.append(row_df)
                except Exception as exc:
                    print(f"[FAIL] Could not read {benchmark_file}: {exc}")

        all_cases_df, grouped_mean_df = _build_overview_tables(row_dfs=row_dfs, image_count_map=image_count_map)

        output_path = final_results_root / output_filename
        _write_overview_workbook(
            output_path=output_path,
            all_cases_df=all_cases_df,
            grouped_mean_df=grouped_mean_df,
            overwrite_results=overwrite_results,
        )

        print(f"[OK] Wrote overview workbook: {output_path}\n")

    print("Done.")


if __name__ == "__main__":
    main()