#!/usr/bin/env python3
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional, Tuple, List

from openpyxl import load_workbook, Workbook


# =========================
# CONFIG
# =========================

FINAL_DIRNAME = "0_FINAL"
OUTPUT_XLSX_NAME = "summary_results.xlsx"

# Keys are your output columns. Values must match the exact label text in the run Excel (col A).
METRIC_LABELS: Dict[str, str] = {
    "avg_off_nadir_deg": "Average off-nadir angle (deg)",
    "avg_gsd_m": "Average GSD (m)",
    "avg_viewing_time_s": "Average viewing time (s)",
    "avg_latency_observation_s": "Average latency, observation (s)",
    "avg_latency_confirmation_s": "Average latency, confirmation (s)",
    "confirmations_per_day_all_sats": "Confirmations per day (all sats)",
    "confirmations_per_satellite_per_day": "Confirmations per satellite per day",
}

# =========================


@dataclass(frozen=True)
class MissionMeta:
    mission_name: str
    orbits: Optional[int]
    sats_per_orbit: Optional[int]
    off_nadir_deg: Optional[float]
    tip_cue_delay_min: Optional[float]
    seed: Optional[int]


def parse_mission_name(name: str) -> MissionMeta:
    """parse_mission_name(name) -> MissionMeta: Parse folder name into structured metadata."""
    mission_name = name.strip()

    re_const = re.compile(r"(?P<orbits>\d+)x(?P<sats>\d+)sat", re.IGNORECASE)
    re_tc = re.compile(
        r"^TC_(?P<const>\d+x\d+sat)_(?P<deg>-?\d+(?:\.\d+)?)deg_(?P<delay>-?\d+(?:\.\d+)?)min_(?P<seed>\d+)sd$",
        re.IGNORECASE,
    )
    re_c = re.compile(
        r"^C_(?P<const>\d+x\d+sat)_(?P<seed>\d+)sd$",
        re.IGNORECASE,
    )

    orbits: Optional[int] = None
    sats_per_orbit: Optional[int] = None
    off_nadir_deg: Optional[float] = None
    tip_cue_delay_min: Optional[float] = None
    seed: Optional[int] = None

    m = re_tc.match(mission_name)
    if m:
        m2 = re_const.search(m.group("const"))
        if m2:
            orbits = int(m2.group("orbits"))
            sats_per_orbit = int(m2.group("sats"))
        off_nadir_deg = float(m.group("deg"))
        tip_cue_delay_min = float(m.group("delay"))
        seed = int(m.group("seed"))
        return MissionMeta(mission_name, orbits, sats_per_orbit, off_nadir_deg, tip_cue_delay_min, seed)

    m = re_c.match(mission_name)
    if m:
        m2 = re_const.search(m.group("const"))
        if m2:
            orbits = int(m2.group("orbits"))
            sats_per_orbit = int(m2.group("sats"))
        seed = int(m.group("seed"))
        return MissionMeta(mission_name, orbits, sats_per_orbit, off_nadir_deg, tip_cue_delay_min, seed)

    # Fallback: best effort
    m2 = re_const.search(mission_name)
    if m2:
        orbits = int(m2.group("orbits"))
        sats_per_orbit = int(m2.group("sats"))

    m3 = re.search(r"(?P<deg>-?\d+(?:\.\d+)?)deg", mission_name, re.IGNORECASE)
    if m3:
        off_nadir_deg = float(m3.group("deg"))

    m4 = re.search(r"(?P<delay>-?\d+(?:\.\d+)?)min", mission_name, re.IGNORECASE)
    if m4:
        tip_cue_delay_min = float(m4.group("delay"))

    m5 = re.search(r"(?P<seed>\d+)sd$", mission_name, re.IGNORECASE)
    if m5:
        seed = int(m5.group("seed"))

    return MissionMeta(mission_name, orbits, sats_per_orbit, off_nadir_deg, tip_cue_delay_min, seed)


def _coerce_number(x: Any) -> Optional[float]:
    """_coerce_number(x) -> Optional[float]: Convert Excel cell value to float if possible."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def read_metrics_from_results_xlsx(xlsx_path: Path, metric_labels: Dict[str, str]) -> Dict[str, Optional[float]]:
    """read_metrics_from_results_xlsx(xlsx_path, metric_labels) -> dict[str,Optional[float]]: Extract metrics by label lookup."""
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb.worksheets[0]

    label_to_value: Dict[str, Optional[float]] = {}
    max_rows = min(ws.max_row or 0, 400)

    for r in range(1, max_rows + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is None:
            continue
        label = str(a).strip()
        if not label:
            continue
        if label not in label_to_value:
            label_to_value[label] = _coerce_number(b)

    return {out_key: label_to_value.get(excel_label, None) for out_key, excel_label in metric_labels.items()}


def find_results_excel(run_dir: Path) -> Optional[Path]:
    """find_results_excel(run_dir) -> Optional[Path]: Find results_*.xlsx inside a run folder."""
    candidates = sorted(run_dir.glob("results_*.xlsx"))
    if candidates:
        return candidates[0]
    candidates2 = sorted(run_dir.glob("*.xlsx"))
    for p in candidates2:
        if "result" in p.name.lower():
            return p
    return None


def _mean(values: List[float]) -> Optional[float]:
    """_mean(values) -> Optional[float]: Mean for a non-empty list, else None."""
    if not values:
        return None
    return sum(values) / float(len(values))


def _make_avg_mission_name(orbits: Any, sats_per_orbit: Any, off_nadir_deg: Any, tip_cue_delay_min: Any) -> str:
    """_make_avg_mission_name(orbits, sats_per_orbit, off_nadir_deg, tip_cue_delay_min) -> str: Build a stable averaged identifier."""
    def _fmt_num(x: Any) -> str:
        if x is None:
            return "NA"
        if isinstance(x, (int, float)) and float(x).is_integer():
            return str(int(float(x)))
        return str(x)

    const = f"{_fmt_num(orbits)}x{_fmt_num(sats_per_orbit)}sat"
    deg = f"{_fmt_num(off_nadir_deg)}deg"
    delay = f"{_fmt_num(tip_cue_delay_min)}min"
    return f"AVG_{const}_{deg}_{delay}"


def write_summary_excel(rows: List[Dict[str, Any]], out_path: Path) -> None:
    """write_summary_excel(rows, out_path) -> None: Write raw summary + averaged-over-seed summary."""
    wb = Workbook()

    base_cols = [
        "mission_name",
        "orbits",
        "sats_per_orbit",
        "off_nadir_deg",
        "tip_cue_delay_min",
        "seed",
        "run_folder",
        "results_xlsx",
    ]
    metric_cols = list(METRIC_LABELS.keys())
    cols = base_cols + metric_cols

    # -----------------------
    # Sheet 1: per-run rows
    # -----------------------
    ws = wb.active
    ws.title = "summary"
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, None) for c in cols])

    # -----------------------------------------
    # Sheet 2: average over seeds (grouped)
    # -----------------------------------------
    ws2 = wb.create_sheet("summary_avg_over_seed")

    # Group key = everything that defines the configuration EXCEPT seed
    group_cols = ["orbits", "sats_per_orbit", "off_nadir_deg", "tip_cue_delay_min"]

    groups: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = {}
    for r in rows:
        key = tuple(r.get(c, None) for c in group_cols)
        groups.setdefault(key, []).append(r)

    avg_cols = ["avg_mission_name"] + group_cols + ["n_seeds"] + metric_cols
    ws2.append(avg_cols)

    for key, group_rows in sorted(groups.items(), key=lambda kv: str(kv[0])):
        orbits, sats_per_orbit, off_nadir_deg, tip_cue_delay_min = key
        avg_name = _make_avg_mission_name(orbits, sats_per_orbit, off_nadir_deg, tip_cue_delay_min)

        out_row: Dict[str, Any] = {
            "avg_mission_name": avg_name,
            "orbits": orbits,
            "sats_per_orbit": sats_per_orbit,
            "off_nadir_deg": off_nadir_deg,
            "tip_cue_delay_min": tip_cue_delay_min,
            "n_seeds": len(group_rows),
        }

        for col in metric_cols:
            vals = [float(gr[col]) for gr in group_rows if isinstance(gr.get(col, None), (int, float))]
            out_row[col] = _mean(vals)

        ws2.append([out_row.get(c, None) for c in avg_cols])

    # Basic column width tuning
    for wsx in [ws, ws2]:
        for col_idx in range(1, wsx.max_column + 1):
            header = wsx.cell(row=1, column=col_idx).value
            header_s = str(header) if header is not None else ""
            width = max(12, min(55, len(header_s) + 2))
            wsx.column_dimensions[wsx.cell(row=1, column=col_idx).column_letter].width = width

    wb.save(str(out_path))


def main() -> None:
    """main() -> None: Scan 0_FINAL runs, extract metrics, write summary_results.xlsx."""
    script_dir = Path(__file__).resolve().parent
    final_dir = script_dir / FINAL_DIRNAME

    if not final_dir.is_dir():
        raise SystemExit(f"ERROR: Expected folder not found: {final_dir}")

    rows: List[Dict[str, Any]] = []

    for run_dir in sorted([p for p in final_dir.iterdir() if p.is_dir()]):
        meta = parse_mission_name(run_dir.name)
        xlsx_path = find_results_excel(run_dir)

        if xlsx_path is None or not xlsx_path.is_file():
            # Keep a row so missing runs are visible
            row_missing: Dict[str, Any] = {
                "mission_name": meta.mission_name,
                "orbits": meta.orbits,
                "sats_per_orbit": meta.sats_per_orbit,
                "off_nadir_deg": meta.off_nadir_deg,
                "tip_cue_delay_min": meta.tip_cue_delay_min,
                "seed": meta.seed,
                "run_folder": str(run_dir),
                "results_xlsx": None,
            }
            for k in METRIC_LABELS.keys():
                row_missing[k] = None
            rows.append(row_missing)
            continue

        metrics = read_metrics_from_results_xlsx(xlsx_path, METRIC_LABELS)

        row: Dict[str, Any] = {
            "mission_name": meta.mission_name,
            "orbits": meta.orbits,
            "sats_per_orbit": meta.sats_per_orbit,
            "off_nadir_deg": meta.off_nadir_deg,
            "tip_cue_delay_min": meta.tip_cue_delay_min,
            "seed": meta.seed,
            "run_folder": str(run_dir),
            "results_xlsx": str(xlsx_path),
        }
        row.update(metrics)
        rows.append(row)

    out_path = script_dir / OUTPUT_XLSX_NAME
    write_summary_excel(rows, out_path)
    print(f"OK: wrote {out_path} with {len(rows)} rows")


if __name__ == "__main__":
    main()