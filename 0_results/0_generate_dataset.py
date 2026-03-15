#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import random
import re
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from PIL import Image


GENERATED_ROOT_REL = os.environ.get("GENERATED_ROOT_REL", "generated_dataset").strip()
os.environ["GENERATED_ROOT_REL"] = GENERATED_ROOT_REL

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
CREATE_DATASET_DIR = PROJECT_ROOT / "dataset" / "create_dataset"

for path in (PROJECT_ROOT, CREATE_DATASET_DIR):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from settings import *


SIMULATION_SEED = int(globals().get("whale_seed", 17))
rng_location = random.Random(SIMULATION_SEED + 1)
rng_image = random.Random(SIMULATION_SEED + 2)
rng_wind = random.Random(SIMULATION_SEED + 3)
rng_patch_seed = random.Random(SIMULATION_SEED + 4)
rng_tip_detection = random.Random(SIMULATION_SEED + 5)
rng_rotation = random.Random(SIMULATION_SEED + 6)
rng_mirror = random.Random(SIMULATION_SEED + 7)

FINAL_RESULTS_DIR = "FINAL_RESULTS"
SOURCE_SHEET = "Img"
GEN_SHEET = "dataset_generaton"

DATASET_ROOT = PROJECT_ROOT / "dataset" / "whales_from_space"

# Choose the input locations here.
# Available folders:
# Auckland2006, Auckland2011, ignacio2017, Maui2015, Pelagos2016,
# Valdes2012, Valdes2014, Valdes2016, Witsand2009
INPUT_LOCATIONS = [
    "Pelagos2016",
    "Auckland2006",
]

# "random"   -> one sample per Excel row, location picked randomly from INPUT_LOCATIONS
# "distinct" -> one sample per Excel row for every location in INPUT_LOCATIONS,
#               saved inside the same run folder in:
#               satellite_images_<location> and supporting_dataset_<location>
# "all"      -> for every row, do both:
#               1) one combined random sample in satellite_images/supporting_dataset
#               2) one per-location sample in satellite_images_<location>/supporting_dataset_<location>
LOCATION_MODE = "all"

IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}

GLOBAL_ANNS_PATH = PROJECT_ROOT / "dataset" / "create_dataset" / "final_annotations.json"
WORKER_RUN_PATH = CREATE_DATASET_DIR / "worker_run.py"

READ_COLUMNS = [
    "detection_id",
    "cue_lat",
    "cue_lon",
    "cue_alt",
    "tgt_lat",
    "tgt_lon",
    "tgt_alt",
    "t_datetime",
    "dem_seed",
]

GEN_COLUMNS = [
    "row_number",
    "detection_id",
    "wind_speed",
    "location",
    "image_file",
    "patch_seed",
    "whale_present",
    "rotation_angle_deg",
    "mirror_bool",
    "patch_name",
    "offnadir_deg",
    "saved_image",
]

PATCH_MODE_MULTIPLE_ALLOW_PARTIAL = False
PATCH_WINDOW_SIZE = 64
PATCH_NOWHALE_MAX_FRACTION = 0.10
PATCH_WHALE_MIN_FRACTION = 0.99
PATCH_HALF_FRACTION_RANGE = (0.20, 0.80)
PATCH_MASK_ALPHA = 80

SHOW_PLOT = False
SATELLITE_DIRNAME = "satellite_images"
SUPPORTING_DIRNAME = "supporting_dataset"
SATELLITE_JSON_NAME = "annotations.json"
TARGET_IMAGE_SPLIT = "reflection_offnadir_glint_255"

PATCH_ATTEMPT_LIMIT = 1000
TIP_WHALE_PROBABILITY = float(tip_tpr)
ROTATION_CHOICES = [0, 90, 180, -90]

SUPPORTING_SPLITS = [
    "texture_nadir_255",
    "texture_nadir_npy",
    "texture_offnadir_255",
    "texture_offnadir_npy",
    "radiance_offnadir_glint_255",
    "radiance_offnadir_glint_npy",
    "radiance_offnadir_no_glint_255",
    "radiance_offnadir_no_glint_npy",
    "reflection_offnadir_glint_255",
    "reflection_offnadir_glint_npy",
    "reflection_offnadir_no_glint_255",
    "reflection_offnadir_no_glint_npy",
]

POSTPROCESS_CATEGORY_IDS_0_BASED = True
POSTPROCESS_REPAIR_BBOX = True
POSTPROCESSED_NAME = "annotations_postprocessed.json"


def result_generated_root(result_folder: Path) -> Path:
    """Return temporary worker output root inside the result folder."""
    return result_folder / SUPPORTING_DIRNAME / "_generated_worker"


def result_generated_root_rel(result_folder: Path) -> str:
    """Return worker output root as absolute path string."""
    return str(result_generated_root(result_folder).resolve())


def normalize_header(value: object) -> str:
    """Normalize Excel header text."""
    return str(value).strip().lower() if value is not None else ""


def sanitize_filename(value: str) -> str:
    """Make filesystem-safe filename stem."""
    text = re.sub(r'[<>:"/\\\\|?*]+', "_", str(value).strip())
    return text.strip(" .") or "unknown_detection"


def validate_location_mode(mode: str) -> str:
    """Validate and normalize the location selection mode."""
    mode_norm = str(mode).strip().lower()
    if mode_norm not in {"random", "distinct", "all"}:
        raise ValueError(f"LOCATION_MODE must be 'random', 'distinct', or 'all', got: {mode}")
    return mode_norm


def get_output_mode(job_mode: str) -> str:
    """Map job mode to output mode used by folders and sheets."""
    if job_mode == "combined":
        return "random"
    if job_mode == "distinct":
        return "distinct"
    raise ValueError(f"Unknown job_mode: {job_mode}")


def location_suffix(output_mode: str, location_name: str | None) -> str:
    """Return directory suffix for the current output mode."""
    if output_mode == "distinct":
        if not location_name:
            raise ValueError("location_name is required in distinct mode")
        return f"_{location_name}"
    return ""


def get_satellite_dir(result_folder: Path, output_mode: str, location_name: str | None) -> Path:
    """Return the output satellite directory for this mode/location."""
    return result_folder / f"{SATELLITE_DIRNAME}{location_suffix(output_mode, location_name)}"


def get_supporting_dir(result_folder: Path, output_mode: str, location_name: str | None) -> Path:
    """Return the output supporting directory for this mode/location."""
    return result_folder / f"{SUPPORTING_DIRNAME}{location_suffix(output_mode, location_name)}"


def parse_datetime_utc(value: object) -> datetime:
    """Convert Excel datetime-like value to timezone-aware UTC datetime."""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)

    text = str(value).strip()
    if not text:
        raise ValueError("Empty t_datetime value")

    text = text.replace(" ", "T")
    if text.endswith("Z"):
        text = text[:-1]
    text = re.sub(r"([+-]\d{2}:\d{2})[+-]\d{2}:\d{2}$", r"\1", text)

    dt = datetime.fromisoformat(text)
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def find_excel_file(folder: Path) -> Path | None:
    """Return first Excel file in a folder."""
    for file_path in sorted(folder.iterdir()):
        if file_path.is_file() and file_path.suffix.lower() == ".xlsx" and not file_path.name.startswith("~$"):
            return file_path
    return None


def make_excel_safe_sheet_name(sheet_name: str) -> str:
    """Return Excel-safe sheet name with invalid chars removed and max length 31."""
    cleaned = re.sub(r'[:\\/?*\[\]]+', "_", str(sheet_name).strip())
    cleaned = cleaned.strip("'")
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:31]


def get_generation_sheet_name(output_mode: str, location_name: str | None = None) -> str:
    """Return output sheet name for combined or distinct mode."""
    if output_mode == "distinct":
        if not location_name:
            raise ValueError("location_name is required in distinct mode")
        return make_excel_safe_sheet_name(f"{GEN_SHEET}_{location_name}")
    return make_excel_safe_sheet_name(GEN_SHEET)


def ensure_generation_sheet(excel_path: Path, output_mode: str, location_name: str | None = None) -> str:
    """Ensure the required generation sheet exists and has the expected headers."""
    sheet_name = get_generation_sheet_name(output_mode=output_mode, location_name=location_name)

    wb = load_workbook(excel_path)
    try:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        for idx, col_name in enumerate(GEN_COLUMNS, start=1):
            ws.cell(row=1, column=idx, value=col_name)

        wb.save(excel_path)
    finally:
        wb.close()

    return sheet_name


def upsert_generation_row(excel_path: Path, row_number: int, detection_id: str, wind_speed: float, location: str, image_file: str, patch_seed: int, whale_present: bool, rotation_angle_deg: int, mirror_bool: bool, patch_name: str, offnadir_deg: float | None, saved_image: str, output_mode: str) -> None:
    """Upsert one row in the correct generation sheet."""
    sheet_name = get_generation_sheet_name(
        output_mode=output_mode,
        location_name=location if output_mode == "distinct" else None,
    )

    wb = load_workbook(excel_path)
    try:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            for idx, col_name in enumerate(GEN_COLUMNS, start=1):
                ws.cell(row=1, column=idx, value=col_name)
        else:
            ws = wb[sheet_name]

        header_map = {
            normalize_header(ws.cell(row=1, column=col).value): col
            for col in range(1, ws.max_column + 1)
            if normalize_header(ws.cell(row=1, column=col).value)
        }

        target_excel_row = None
        for r in range(2, ws.max_row + 1):
            det_val = ws.cell(r, header_map["detection_id"]).value
            row_val = ws.cell(r, header_map["row_number"]).value
            loc_val = ws.cell(r, header_map["location"]).value
            if (
                str(det_val).strip() == str(detection_id).strip()
                and str(row_val).strip() == str(row_number)
                and str(loc_val).strip() == str(location).strip()
            ):
                target_excel_row = r
                break

        if target_excel_row is None:
            target_excel_row = ws.max_row + 1

        values = {
            "row_number": int(row_number),
            "detection_id": detection_id,
            "wind_speed": float(wind_speed),
            "location": location,
            "image_file": image_file,
            "patch_seed": int(patch_seed),
            "whale_present": bool(whale_present),
            "rotation_angle_deg": int(rotation_angle_deg),
            "mirror_bool": bool(mirror_bool),
            "patch_name": patch_name,
            "offnadir_deg": None if offnadir_deg is None else float(offnadir_deg),
            "saved_image": saved_image,
        }

        for key, value in values.items():
            ws.cell(target_excel_row, header_map[key], value=value)

        wb.save(excel_path)
    finally:
        wb.close()


def read_img_rows(excel_path: Path) -> list[tuple[int, str, float, float, float, float, float, float, datetime, int]]:
    """Read all Img rows into memory, then close workbook before later writes."""
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    parsed_rows: list[tuple[int, str, float, float, float, float, float, float, datetime, int]] = []

    try:
        if SOURCE_SHEET not in wb.sheetnames:
            return parsed_rows

        ws = wb[SOURCE_SHEET]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return parsed_rows

        header_map = {normalize_header(value): idx for idx, value in enumerate(header)}

        missing = [col for col in READ_COLUMNS if col not in header_map]
        if missing:
            raise KeyError(f"Missing required columns in '{SOURCE_SHEET}': {missing}")

        idx = [header_map[col] for col in READ_COLUMNS]

        for row_number, row in enumerate(rows, start=2):
            if row is None:
                continue

            extracted = [row[i] if i < len(row) else None for i in idx]

            if all(value is None or str(value).strip() == "" for value in extracted):
                continue

            detection_id = str(extracted[0]).strip() if extracted[0] is not None else ""
            if not detection_id:
                continue

            parsed_rows.append((
                row_number,
                detection_id,
                float(extracted[1]),
                float(extracted[2]),
                float(extracted[3]),
                float(extracted[4]),
                float(extracted[5]),
                float(extracted[6]),
                parse_datetime_utc(extracted[7]),
                int(extracted[8]),
            ))

        return parsed_rows
    finally:
        wb.close()


def iter_generation_jobs(base_dir: Path, dataset_paths: list[Path], location_mode: str):
    """Yield generation jobs per row for random, distinct, or all mode."""
    final_results = base_dir / FINAL_RESULTS_DIR
    if not final_results.is_dir():
        raise FileNotFoundError(f"Missing folder: {final_results}")

    for folder in sorted(final_results.iterdir()):
        if not folder.is_dir():
            continue

        excel_file = find_excel_file(folder)
        if excel_file is None:
            continue

        if location_mode in {"random", "all"}:
            ensure_generation_sheet(excel_file, output_mode="random")

        if location_mode in {"distinct", "all"}:
            for dataset_folder in dataset_paths:
                ensure_generation_sheet(
                    excel_file,
                    output_mode="distinct",
                    location_name=dataset_folder.name,
                )

        rows = read_img_rows(excel_file)

        for row in rows:
            if location_mode in {"random", "all"}:
                yield folder, excel_file, "combined", None, *row

            if location_mode in {"distinct", "all"}:
                for dataset_folder in dataset_paths:
                    yield folder, excel_file, "distinct", dataset_folder, *row


def list_images_recursive(folder: Path) -> list[Path]:
    """Return all image files recursively."""
    return sorted(
        path for path in folder.rglob("*")
        if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES
    )


def resolve_dataset_choices(dataset_root: Path, requested_names: list[str]) -> list[Path]:
    """Resolve requested dataset folders case-insensitively."""
    by_lower = {p.name.lower(): p for p in dataset_root.iterdir() if p.is_dir()}
    out = []
    for name in requested_names:
        key = name.lower()
        if key not in by_lower:
            raise FileNotFoundError(f"Dataset folder not found: {name}")
        out.append(by_lower[key])
    return out


def choose_dataset_folder(dataset_paths: list[Path], forced_dataset_folder: Path | None = None) -> Path:
    """Choose dataset folder randomly, or return the forced folder in distinct mode."""
    if forced_dataset_folder is not None:
        return forced_dataset_folder
    return rng_location.choice(dataset_paths)


def choose_random_image(dataset_folder: Path) -> Path:
    """Choose random image from selected dataset folder."""
    images = list_images_recursive(dataset_folder)
    if not images:
        raise FileNotFoundError(f"No images found in {dataset_folder}")
    return rng_image.choice(images)


def load_global_coco_meta(json_path: Path) -> dict:
    """Load top-level COCO metadata."""
    with json_path.open("r", encoding="utf-8") as f:
        coco = json.load(f)

    return {
        "info": coco.get("info", {}),
        "licenses": coco.get("licenses", []),
        "categories": coco.get("categories", []),
    }


def choose_patch_mode() -> tuple[bool, str]:
    """Sample whale vs ocean according to tip_tpr."""
    whale_present = rng_tip_detection.random() < TIP_WHALE_PROBABILITY
    mode_single = "full" if whale_present else "ocean"
    return whale_present, mode_single


def choose_rotation_and_mirror() -> tuple[int, bool]:
    """Choose randomized rotation and mirror reproducibly."""
    rotation_angle_deg = rng_rotation.choice(ROTATION_CHOICES)
    mirror_bool = bool(rng_mirror.randint(0, 1))
    return rotation_angle_deg, mirror_bool


def build_worker_common_cmd(img_file: str, patch_seed: int, dem_seed: int, sat_lat: float, sat_lon: float, sat_alt: float, tgt_lat: float, tgt_lon: float, tgt_alt: float, datetime_utc: datetime, sensor_chars: dict, wave_props: dict, meta_out: Path, run_idx: int, mode_single: str, rotation_angle_deg: int, mirror_bool: bool) -> list[str]:
    """Build worker_run.py common CLI arguments."""
    return [
        sys.executable,
        str(WORKER_RUN_PATH),
        "--img_file", img_file,
        "--patch_seed", str(int(patch_seed)),
        "--dem_seed", str(int(dem_seed)),
        "--show_plot", "1" if SHOW_PLOT else "0",
        "--sat_lat", str(float(sat_lat)),
        "--sat_lon", str(float(sat_lon)),
        "--sat_alt", str(float(sat_alt)),
        "--tgt_lat", str(float(tgt_lat)),
        "--tgt_lon", str(float(tgt_lon)),
        "--tgt_alt", str(float(tgt_alt)),
        "--datetime_utc", datetime_utc.isoformat(),
        "--mode_single", mode_single,
        "--mode_multiple_allow_partial", "1" if PATCH_MODE_MULTIPLE_ALLOW_PARTIAL else "0",
        "--window_size", str(int(PATCH_WINDOW_SIZE)),
        "--rotation_angle_deg", str(int(rotation_angle_deg)),
        "--mirror_bool", "1" if mirror_bool else "0",
        "--nowhale_max_fraction", str(float(PATCH_NOWHALE_MAX_FRACTION)),
        "--whale_min_fraction", str(float(PATCH_WHALE_MIN_FRACTION)),
        "--half_fraction_low", str(float(PATCH_HALF_FRACTION_RANGE[0])),
        "--half_fraction_high", str(float(PATCH_HALF_FRACTION_RANGE[1])),
        "--mask_alpha", str(int(PATCH_MASK_ALPHA)),
        "--sensor_json", json.dumps(sensor_chars),
        "--wave_json", json.dumps(wave_props),
        "--meta_out", str(meta_out.resolve()),
        "--run_idx", str(int(run_idx)),
    ]


def run_worker(stage: str, cmd_common: list[str], result_folder: Path, patch_name: str = "") -> None:
    """Run one worker stage with per-run output root inside the result folder."""
    env = os.environ.copy()
    env["GENERATED_ROOT_REL"] = result_generated_root_rel(result_folder)

    pythonpath_parts = [str(PROJECT_ROOT), str(CREATE_DATASET_DIR)]
    old_pythonpath = env.get("PYTHONPATH", "").strip()
    if old_pythonpath:
        pythonpath_parts.append(old_pythonpath)
    env["PYTHONPATH"] = os.pathsep.join(pythonpath_parts)

    cmd = list(cmd_common)
    cmd.extend(["--stage", stage])
    if patch_name:
        cmd.extend(["--patch_name", patch_name])

    subprocess.run(cmd, check=True, env=env, cwd=str(PROJECT_ROOT))


def run_one_pipeline(result_folder: Path, img_file: str, patch_seed: int, dem_seed: int, sat_lat: float, sat_lon: float, sat_alt: float, tgt_lat: float, tgt_lon: float, tgt_alt: float, datetime_utc: datetime, sensor_chars: dict, wave_props: dict, meta_out: Path, run_idx: int, mode_single: str, rotation_angle_deg: int, mirror_bool: bool) -> dict:
    """Run nadir then offnadir worker pipeline and return meta json."""
    if meta_out.exists():
        meta_out.unlink()

    cmd_common = build_worker_common_cmd(
        img_file=img_file,
        patch_seed=patch_seed,
        dem_seed=dem_seed,
        sat_lat=sat_lat,
        sat_lon=sat_lon,
        sat_alt=sat_alt,
        tgt_lat=tgt_lat,
        tgt_lon=tgt_lon,
        tgt_alt=tgt_alt,
        datetime_utc=datetime_utc,
        sensor_chars=sensor_chars,
        wave_props=wave_props,
        meta_out=meta_out,
        run_idx=run_idx,
        mode_single=mode_single,
        rotation_angle_deg=rotation_angle_deg,
        mirror_bool=mirror_bool,
    )

    run_worker(stage="nadir", cmd_common=cmd_common, result_folder=result_folder)

    if not meta_out.exists():
        raise FileNotFoundError(f"Nadir worker did not write meta file: {meta_out}")

    meta = json.loads(meta_out.read_text(encoding="utf-8"))
    patch_name = str(meta.get("patch_name", "")).strip()
    if not patch_name:
        raise RuntimeError("patch_name missing after nadir stage")

    run_worker(stage="offnadir", cmd_common=cmd_common, result_folder=result_folder, patch_name=patch_name)

    if not meta_out.exists():
        raise FileNotFoundError(f"Offnadir worker did not write meta file: {meta_out}")

    return json.loads(meta_out.read_text(encoding="utf-8"))


def get_generated_root_safe(result_folder: Path) -> Path:
    """Return the first existing worker output root candidate."""
    expected = result_generated_root(result_folder).resolve()
    env_value = os.environ.get("GENERATED_ROOT_REL", "").strip()

    candidates: list[Path] = [expected]

    if env_value:
        env_path = Path(env_value)
        if env_path.is_absolute():
            candidates.append(env_path.resolve())
        else:
            candidates.append((PROJECT_ROOT / env_path).resolve())
            candidates.append((CREATE_DATASET_DIR / env_path).resolve())
            candidates.append((CREATE_DATASET_DIR.parent / env_path).resolve())

    seen = set()
    unique_candidates: list[Path] = []
    for path in candidates:
        key = str(path).lower()
        if key not in seen:
            seen.add(key)
            unique_candidates.append(path)

    for path in unique_candidates:
        if path.exists():
            return path

    raise FileNotFoundError(
        "Worker output root not found. Checked: "
        + " | ".join(str(p) for p in unique_candidates)
    )


def matches_patch_output(path: Path, patch_name: str) -> bool:
    """Check whether output file belongs to patch_name."""
    stem = path.stem
    return stem == patch_name or stem.startswith(patch_name + "_")


def move_generated_file(src: Path, dst: Path) -> Path:
    """Move one generated file to destination, overwriting existing file."""
    dst.parent.mkdir(parents=True, exist_ok=True)
    if dst.exists():
        dst.unlink()
    shutil.move(str(src), str(dst))
    return dst


def move_supporting_outputs(result_folder: Path, detection_id: str, img_file: str, patch_name: str, output_mode: str, location_name: str | None) -> dict[str, list[Path]]:
    """Move generated outputs for one patch into the correct supporting folder."""
    generated_root = get_generated_root_safe(result_folder)
    supporting_root = get_supporting_dir(result_folder=result_folder, output_mode=output_mode, location_name=location_name)
    safe_id = sanitize_filename(detection_id)

    moved: dict[str, list[Path]] = {}
    found_any = False

    for split_name in SUPPORTING_SPLITS:
        moved[split_name] = []
        split_root = generated_root / split_name
        dst_dir = supporting_root / split_name

        if not split_root.exists():
            continue

        for src in sorted(split_root.rglob("*")):
            if not src.is_file():
                continue
            if not matches_patch_output(src, patch_name):
                continue

            dst = dst_dir / f"{safe_id}{src.suffix.lower()}"
            moved[split_name].append(move_generated_file(src, dst))
            found_any = True

    if not found_any:
        available = []
        for split_name in SUPPORTING_SPLITS:
            split_root = generated_root / split_name
            if split_root.exists():
                sample_files = [p.name for p in sorted(split_root.rglob("*")) if p.is_file()][:10]
                available.append(f"{split_name}: {sample_files}")

        raise FileNotFoundError(
            f"No generated files found for patch_name='{patch_name}' under worker root '{generated_root}'. "
            f"Available files by split: {available}"
        )

    return moved


def save_satellite_image(result_folder: Path, detection_id: str, moved_outputs: dict[str, list[Path]], output_mode: str, location_name: str | None) -> Path:
    """Copy the target split PNG into the correct satellite_images folder."""
    candidates = [p for p in moved_outputs.get(TARGET_IMAGE_SPLIT, []) if p.suffix.lower() == ".png"]

    if not candidates:
        available = {k: [p.name for p in v] for k, v in moved_outputs.items() if v}
        raise FileNotFoundError(
            f"No PNG output found for split '{TARGET_IMAGE_SPLIT}'. "
            f"Available moved outputs: {available}"
        )

    src = candidates[0]
    satellite_dir = get_satellite_dir(result_folder=result_folder, output_mode=output_mode, location_name=location_name)
    satellite_dir.mkdir(parents=True, exist_ok=True)

    dst = satellite_dir / f"{sanitize_filename(detection_id)}.png"
    if dst.exists():
        dst.unlink()
    shutil.copy2(src, dst)
    return dst


def ann_rows_to_coco(ann_rows: list[dict]) -> list[dict]:
    """Convert worker meta annotation rows to COCO annotations."""
    out = []
    for row in ann_rows:
        if not isinstance(row, dict):
            continue

        ann = {
            "category_id": row.get("category_id"),
            "bbox": row.get("bbox", []),
            "segmentation": row.get("segmentation", []),
            "area": row.get("area", 0.0),
            "iscrowd": row.get("iscrowd", 0),
        }

        other = row.get("other")
        if isinstance(other, dict):
            ann.update(other)

        out.append(ann)
    return out


def ann_rows_to_coco_nadir(ann_rows: list[dict]) -> list[dict]:
    """Convert worker meta nadir annotation rows to COCO annotations."""
    out = []
    for row in ann_rows:
        if not isinstance(row, dict):
            continue

        ann = {
            "category_id": row.get("category_id"),
            "bbox": row.get("bbox", []),
            "segmentation": row.get("segmentation", []),
            "area": row.get("area", 0.0),
            "iscrowd": row.get("iscrowd", 0),
        }

        other = row.get("other")
        if isinstance(other, dict):
            ann.update(other)

        out.append(ann)
    return out


def get_split_annotations(meta: dict, split_name: str) -> list[dict]:
    """Return correct annotations for a supporting split."""
    if split_name.startswith("texture_nadir_"):
        return ann_rows_to_coco_nadir(meta.get("anns_nadir", []))
    return ann_rows_to_coco(meta.get("anns_offnadir", []))


def load_or_init_coco(json_path: Path, coco_meta: dict) -> dict:
    """Load existing COCO json or initialize one."""
    if json_path.exists():
        with json_path.open("r", encoding="utf-8") as f:
            return json.load(f)

    return {
        "info": coco_meta["info"],
        "licenses": coco_meta["licenses"],
        "categories": coco_meta["categories"],
        "images": [],
        "annotations": [],
    }


def next_image_id(coco: dict) -> int:
    """Return next free COCO image id."""
    ids = [int(img.get("id", 0)) for img in coco.get("images", [])]
    return max(ids, default=0) + 1


def next_annotation_id(coco: dict) -> int:
    """Return next free COCO annotation id."""
    ids = [int(ann.get("id", 0)) for ann in coco.get("annotations", [])]
    return max(ids, default=0) + 1


def remove_existing_image_entries(coco: dict, file_name: str) -> None:
    """Remove old image and linked annotations for same file name."""
    image_ids = {int(img["id"]) for img in coco.get("images", []) if img.get("file_name") == file_name}
    if not image_ids:
        return

    coco["images"] = [img for img in coco.get("images", []) if int(img.get("id", -1)) not in image_ids]
    coco["annotations"] = [ann for ann in coco.get("annotations", []) if int(ann.get("image_id", -1)) not in image_ids]


def get_image_size_for_annotation(path: Path) -> tuple[int, int]:
    """Return width, height for png or npy image-like arrays."""
    suffix = path.suffix.lower()

    if suffix == ".png":
        with Image.open(path) as img:
            w, h = img.size
        return int(w), int(h)

    if suffix == ".npy":
        arr = np.load(path)
        if arr.ndim < 2:
            raise ValueError(f"NPY file does not look like an image: {path}")
        h, w = arr.shape[:2]
        return int(w), int(h)

    raise ValueError(f"Unsupported annotation image type: {path}")


def append_split_annotations(split_dir: Path, image_path: Path, anns: list[dict], coco_meta: dict, detection_id: str, image_file: str, row_number: int, patch_seed: int, wind_speed: float, location: str, whale_present: bool, patch_name: str, offnadir_deg: float | None, rotation_angle_deg: int | float, mirror_bool: bool) -> None:
    """Write annotations.json inside one supporting split folder."""
    json_path = split_dir / "annotations.json"
    coco = load_or_init_coco(json_path=json_path, coco_meta=coco_meta)

    img_w, img_h = get_image_size_for_annotation(image_path)

    rel_file_name = image_path.name
    remove_existing_image_entries(coco=coco, file_name=rel_file_name)

    image_id = next_image_id(coco)
    ann_id = next_annotation_id(coco)

    coco["images"].append({
        "id": int(image_id),
        "file_name": rel_file_name,
        "width": int(img_w),
        "height": int(img_h),
        "detection_id": detection_id,
        "source_image_file": image_file,
        "row_number": int(row_number),
        "patch_seed": int(patch_seed),
        "wind_speed": float(wind_speed),
        "location": location,
        "whale_present": bool(whale_present),
        "patch_name": patch_name,
        "offnadir_deg": offnadir_deg,
        "rotation_angle_deg": float(rotation_angle_deg),
        "mirror_bool": bool(mirror_bool),
    })

    for ann in anns:
        ann_copy = dict(ann)
        ann_copy["id"] = int(ann_id)
        ann_copy["image_id"] = int(image_id)
        coco["annotations"].append(ann_copy)
        ann_id += 1

    with json_path.open("w", encoding="utf-8") as f:
        json.dump(coco, f, indent=2)


def append_satellite_annotations(result_folder: Path, saved_image_path: Path, meta: dict, detection_id: str, image_file: str, row_number: int, patch_seed: int, wind_speed: float, location: str, whale_present: bool, coco_meta: dict, output_mode: str, location_name: str | None) -> None:
    """Append one satellite image and its offnadir annotations to the correct satellite folder."""
    satellite_dir = get_satellite_dir(result_folder=result_folder, output_mode=output_mode, location_name=location_name)
    satellite_dir.mkdir(parents=True, exist_ok=True)

    json_path = satellite_dir / SATELLITE_JSON_NAME
    coco = load_or_init_coco(json_path=json_path, coco_meta=coco_meta)

    with Image.open(saved_image_path) as img:
        img_w, img_h = img.size

    rel_file_name = saved_image_path.name
    remove_existing_image_entries(coco=coco, file_name=rel_file_name)

    image_id = next_image_id(coco)
    ann_id = next_annotation_id(coco)

    coco["images"].append({
        "id": int(image_id),
        "file_name": rel_file_name,
        "width": int(img_w),
        "height": int(img_h),
        "detection_id": detection_id,
        "source_image_file": image_file,
        "row_number": int(row_number),
        "patch_seed": int(patch_seed),
        "wind_speed": float(wind_speed),
        "location": location,
        "whale_present": bool(whale_present),
        "patch_name": meta.get("patch_name", ""),
        "label_simple": meta.get("label_simple", ""),
        "offnadir_deg": meta.get("offnadir_deg", None),
        "rotation_angle_deg": meta.get("rotation_angle_deg", 0.0),
        "mirror_bool": meta.get("mirror_bool", False),
    })

    anns = ann_rows_to_coco(meta.get("anns_offnadir", []))
    for ann in anns:
        ann_copy = dict(ann)
        ann_copy["id"] = int(ann_id)
        ann_copy["image_id"] = int(image_id)
        coco["annotations"].append(ann_copy)
        ann_id += 1

    with json_path.open("w", encoding="utf-8") as f:
        json.dump(coco, f, indent=2)


def append_supporting_annotations(result_folder: Path, moved_outputs: dict[str, list[Path]], meta: dict, detection_id: str, image_file: str, row_number: int, patch_seed: int, wind_speed: float, location: str, whale_present: bool, coco_meta: dict, output_mode: str, location_name: str | None) -> None:
    """Write per-split annotations.json inside the correct supporting folder."""
    supporting_dir = get_supporting_dir(result_folder=result_folder, output_mode=output_mode, location_name=location_name)
    supporting_dir.mkdir(parents=True, exist_ok=True)

    patch_name = str(meta.get("patch_name", ""))
    offnadir_deg = meta.get("offnadir_deg", None)
    rotation_angle_deg = meta.get("rotation_angle_deg", 0.0)
    mirror_bool = meta.get("mirror_bool", False)

    for split_name, files in moved_outputs.items():
        if not files:
            continue

        anns = get_split_annotations(meta=meta, split_name=split_name)
        split_dir = supporting_dir / split_name
        split_dir.mkdir(parents=True, exist_ok=True)

        for image_path in files:
            append_split_annotations(
                split_dir=split_dir,
                image_path=image_path,
                anns=anns,
                coco_meta=coco_meta,
                detection_id=detection_id,
                image_file=image_file,
                row_number=row_number,
                patch_seed=patch_seed,
                wind_speed=wind_speed,
                location=location,
                whale_present=whale_present,
                patch_name=patch_name,
                offnadir_deg=offnadir_deg,
                rotation_angle_deg=rotation_angle_deg,
                mirror_bool=mirror_bool,
            )


def rollback_worker_patch_outputs(result_folder: Path, img_file: str, patch_name: str) -> None:
    """Delete failed patch outputs from the worker root by recursive search."""
    try:
        generated_root = get_generated_root_safe(result_folder)
    except Exception:
        return

    for split_name in SUPPORTING_SPLITS:
        split_root = generated_root / split_name
        if not split_root.exists():
            continue

        for src in list(split_root.rglob("*")):
            if not src.is_file():
                continue
            if not matches_patch_output(src, patch_name):
                continue
            try:
                src.unlink()
            except Exception:
                pass


def cleanup_empty_generated_dirs(result_folder: Path, img_file: str) -> None:
    """No-op; full worker root is cleaned up later."""
    return


def cleanup_worker_root(result_folder: Path) -> None:
    """Remove temporary per-run worker output root completely."""
    tmp_root = result_generated_root(result_folder)
    if tmp_root.exists():
        shutil.rmtree(tmp_root, ignore_errors=True)


def load_json(path: Path) -> dict:
    """Read JSON utf-8."""
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: dict) -> None:
    """Write JSON utf-8 pretty."""
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def remap_categories_0_based(coco: dict) -> dict:
    """Remap category ids to 0..K-1 and update annotations."""
    coco2 = dict(coco)
    cats = [dict(c) for c in coco.get("categories", [])]
    anns = [dict(a) for a in coco.get("annotations", [])]

    ids = set()
    for c in cats:
        if "id" in c:
            ids.add(c["id"])
    for a in anns:
        if "category_id" in a:
            ids.add(a["category_id"])

    if not ids:
        coco2["categories"] = cats
        coco2["annotations"] = anns
        return coco2

    old_ids = sorted(ids)
    mapping = {old: new for new, old in enumerate(old_ids)}

    for c in cats:
        if "id" in c:
            c["id"] = mapping[c["id"]]

    for a in anns:
        if "category_id" in a:
            a["category_id"] = mapping[a["category_id"]]

    coco2["categories"] = cats
    coco2["annotations"] = anns
    return coco2


def bbox_from_segmentation(seg: object) -> list[float] | None:
    """Compute [x,y,w,h] from COCO polygon segmentation."""
    if not isinstance(seg, list) or not seg:
        return None

    xs: list[float] = []
    ys: list[float] = []

    for poly in seg:
        if not isinstance(poly, list) or len(poly) < 6 or len(poly) % 2 != 0:
            continue
        for i in range(0, len(poly), 2):
            try:
                xs.append(float(poly[i]))
                ys.append(float(poly[i + 1]))
            except Exception:
                return None

    if not xs or not ys:
        return None

    x0 = min(xs)
    y0 = min(ys)
    x1 = max(xs)
    y1 = max(ys)
    w = x1 - x0
    h = y1 - y0
    if w <= 0 or h <= 0:
        return None
    return [x0, y0, w, h]


def is_valid_bbox(b: object) -> bool:
    """True if bbox is [x,y,w,h] with w,h>0."""
    if not isinstance(b, (list, tuple)) or len(b) != 4:
        return False
    try:
        w = float(b[2])
        h = float(b[3])
    except Exception:
        return False
    return w > 0 and h > 0


def repair_annotations_bboxes(coco: dict) -> tuple[dict, dict]:
    """Fix missing bbox from seg or drop invalid annotations."""
    coco2 = dict(coco)
    anns = [dict(a) for a in coco.get("annotations", [])]

    fixed = 0
    dropped = 0
    kept: list[dict] = []

    for a in anns:
        if is_valid_bbox(a.get("bbox")):
            kept.append(a)
            continue

        b = bbox_from_segmentation(a.get("segmentation"))
        if b is not None:
            a["bbox"] = b
            fixed += 1
            kept.append(a)
            continue

        dropped += 1

    coco2["annotations"] = kept
    return coco2, {"bbox_fixed_from_segmentation": fixed, "bbox_dropped_no_bbox_no_seg": dropped}


def postprocess_one_annotation_file(json_path: Path) -> None:
    """Read annotations.json and write annotations_postprocessed.json."""
    if not json_path.exists():
        return

    coco = load_json(json_path)

    if POSTPROCESS_REPAIR_BBOX:
        coco, stats = repair_annotations_bboxes(coco)
    else:
        stats = {"bbox_fixed_from_segmentation": 0, "bbox_dropped_no_bbox_no_seg": 0}

    if POSTPROCESS_CATEGORY_IDS_0_BASED:
        coco = remap_categories_0_based(coco)

    out_path = json_path.with_name(POSTPROCESSED_NAME)
    save_json(out_path, coco)

    print(
        f"POSTPROCESSED: {out_path} | "
        f"bbox_fixed={stats['bbox_fixed_from_segmentation']} | "
        f"bbox_dropped={stats['bbox_dropped_no_bbox_no_seg']}"
    )


def postprocess_result_folder(result_folder: Path) -> None:
    """Postprocess all annotations.json files inside one result folder."""
    for child in sorted(result_folder.iterdir()):
        if not child.is_dir():
            continue

        if child.name == SATELLITE_DIRNAME or child.name.startswith(SATELLITE_DIRNAME + "_"):
            sat_json = child / SATELLITE_JSON_NAME
            postprocess_one_annotation_file(sat_json)

        if child.name == SUPPORTING_DIRNAME or child.name.startswith(SUPPORTING_DIRNAME + "_"):
            for split_name in SUPPORTING_SPLITS:
                split_json = child / split_name / "annotations.json"
                postprocess_one_annotation_file(split_json)


def run_pipeline_with_retries(result_folder: Path, dataset_choices: list[Path], dem_seed: int, cue_lat: float, cue_lon: float, cue_alt: float, tgt_lat: float, tgt_lon: float, tgt_alt: float, t_datetime: datetime, sensor_chars: dict, wave_props: dict, meta_out: Path, run_idx: int, whale_present: bool, rotation_angle_deg: int, mirror_bool: bool, forced_dataset_folder: Path | None = None) -> tuple[dict, str, int, str]:
    """Retry with new source images until one valid patch succeeds."""
    mode_single = "full" if whale_present else "ocean"
    last_error: Exception | None = None

    for attempt_idx in range(PATCH_ATTEMPT_LIMIT):
        dataset_folder = choose_dataset_folder(dataset_paths=dataset_choices, forced_dataset_folder=forced_dataset_folder)
        img_path = choose_random_image(dataset_folder)
        image_file = img_path.relative_to(DATASET_ROOT).as_posix()
        patch_seed = rng_patch_seed.randrange(0, 2**31 - 1)

        try:
            meta = run_one_pipeline(
                result_folder=result_folder,
                img_file=image_file,
                patch_seed=patch_seed,
                dem_seed=dem_seed,
                sat_lat=cue_lat,
                sat_lon=cue_lon,
                sat_alt=cue_alt,
                tgt_lat=tgt_lat,
                tgt_lon=tgt_lon,
                tgt_alt=tgt_alt,
                datetime_utc=t_datetime,
                sensor_chars=sensor_chars,
                wave_props=wave_props,
                meta_out=meta_out,
                run_idx=run_idx,
                mode_single=mode_single,
                rotation_angle_deg=rotation_angle_deg,
                mirror_bool=mirror_bool,
            )
            return meta, image_file, patch_seed, dataset_folder.name

        except Exception as exc:
            last_error = exc
            try:
                if meta_out.exists():
                    meta_tmp = json.loads(meta_out.read_text(encoding="utf-8"))
                    patch_name = str(meta_tmp.get("patch_name", "")).strip()
                    if patch_name:
                        rollback_worker_patch_outputs(result_folder=result_folder, img_file=image_file, patch_name=patch_name)
            except Exception:
                pass

            print(f"[retry {attempt_idx + 1}/{PATCH_ATTEMPT_LIMIT}] failed for {image_file} mode={mode_single}: {exc}")

    raise RuntimeError(f"Failed to generate a valid '{mode_single}' sample after {PATCH_ATTEMPT_LIMIT} attempts") from last_error


def main() -> None:
    if not GLOBAL_ANNS_PATH.exists():
        raise FileNotFoundError(f"Annotation file not found: {GLOBAL_ANNS_PATH}")
    if not WORKER_RUN_PATH.exists():
        raise FileNotFoundError(f"worker_run.py not found: {WORKER_RUN_PATH}")

    location_mode = validate_location_mode(LOCATION_MODE)
    dataset_choices = resolve_dataset_choices(dataset_root=DATASET_ROOT, requested_names=INPUT_LOCATIONS)
    coco_meta = load_global_coco_meta(json_path=GLOBAL_ANNS_PATH)

    run_idx = 0
    current_folder: Path | None = None

    for (
        result_folder,
        excel_path,
        job_mode,
        forced_dataset_folder,
        row_number,
        detection_id,
        cue_lat,
        cue_lon,
        cue_alt,
        tgt_lat,
        tgt_lon,
        tgt_alt,
        t_datetime,
        dem_seed,
    ) in iter_generation_jobs(SCRIPT_DIR, dataset_choices, location_mode):

        if current_folder is None:
            current_folder = result_folder
        elif result_folder != current_folder:
            postprocess_result_folder(current_folder)
            cleanup_worker_root(current_folder)
            current_folder = result_folder

        output_mode = get_output_mode(job_mode)

        whale_present, _mode_single = choose_patch_mode()
        rotation_angle_deg, mirror_bool = choose_rotation_and_mirror()
        wind_speed = round(rng_wind.uniform(2.0, 12.0), 3)

        sensor_chars = dict(sensor_characteristics)
        wave_props = dict(wave_properties)
        wave_props["wind_speed"] = float(wind_speed)

        meta_dir = result_folder / SUPPORTING_DIRNAME / "_meta"
        meta_dir.mkdir(parents=True, exist_ok=True)
        meta_out = meta_dir / f"run_{run_idx:06d}.json"

        meta, image_file, patch_seed, location = run_pipeline_with_retries(
            result_folder=result_folder,
            dataset_choices=dataset_choices,
            dem_seed=dem_seed,
            cue_lat=cue_lat,
            cue_lon=cue_lon,
            cue_alt=cue_alt,
            tgt_lat=tgt_lat,
            tgt_lon=tgt_lon,
            tgt_alt=tgt_alt,
            t_datetime=t_datetime,
            sensor_chars=sensor_chars,
            wave_props=wave_props,
            meta_out=meta_out,
            run_idx=run_idx,
            whale_present=whale_present,
            rotation_angle_deg=rotation_angle_deg,
            mirror_bool=mirror_bool,
            forced_dataset_folder=forced_dataset_folder,
        )

        patch_name = str(meta.get("patch_name", "")).strip()
        if not patch_name:
            raise RuntimeError("patch_name missing in final meta")

        moved_outputs = move_supporting_outputs(
            result_folder=result_folder,
            detection_id=detection_id,
            img_file=image_file,
            patch_name=patch_name,
            output_mode=output_mode,
            location_name=location if output_mode == "distinct" else None,
        )
        cleanup_empty_generated_dirs(result_folder, image_file)

        saved_image_path = save_satellite_image(
            result_folder=result_folder,
            detection_id=detection_id,
            moved_outputs=moved_outputs,
            output_mode=output_mode,
            location_name=location if output_mode == "distinct" else None,
        )

        append_satellite_annotations(
            result_folder=result_folder,
            saved_image_path=saved_image_path,
            meta=meta,
            detection_id=detection_id,
            image_file=image_file,
            row_number=row_number,
            patch_seed=patch_seed,
            wind_speed=wind_speed,
            location=location,
            whale_present=whale_present,
            coco_meta=coco_meta,
            output_mode=output_mode,
            location_name=location if output_mode == "distinct" else None,
        )

        append_supporting_annotations(
            result_folder=result_folder,
            moved_outputs=moved_outputs,
            meta=meta,
            detection_id=detection_id,
            image_file=image_file,
            row_number=row_number,
            patch_seed=patch_seed,
            wind_speed=wind_speed,
            location=location,
            whale_present=whale_present,
            coco_meta=coco_meta,
            output_mode=output_mode,
            location_name=location if output_mode == "distinct" else None,
        )

        upsert_generation_row(
            excel_path=excel_path,
            row_number=row_number,
            detection_id=detection_id,
            wind_speed=wind_speed,
            location=location,
            image_file=image_file,
            patch_seed=patch_seed,
            whale_present=whale_present,
            rotation_angle_deg=rotation_angle_deg,
            mirror_bool=mirror_bool,
            patch_name=patch_name,
            offnadir_deg=meta.get("offnadir_deg", None),
            saved_image=saved_image_path.name,
            output_mode=output_mode,
        )

        print(
            result_folder.name,
            f"row={row_number}",
            detection_id,
            image_file,
            f"job_mode={job_mode}",
            f"location={location}",
            f"patch_seed={patch_seed}",
            f"wind={wind_speed}",
            f"whale_present={whale_present}",
            f"rotation={rotation_angle_deg}",
            f"mirror={mirror_bool}",
            f"patch_name={patch_name}",
            f"offnadir_deg={meta.get('offnadir_deg', None)}",
            saved_image_path.name,
        )

        run_idx += 1

    if current_folder is not None:
        postprocess_result_folder(current_folder)
        cleanup_worker_root(current_folder)


if __name__ == "__main__":
    main()