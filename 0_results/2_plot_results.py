#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import load_workbook


# =========================
# CONFIG
# =========================

SUMMARY_XLSX_NAME = "summary_results.xlsx"
SHEET_RAW = "summary"  # per-seed rows

# Experiment A: time-delay sweep at 40deg requirement
REQ_OFFNADIR_FOR_TIMEDELAY_DEG = 40.0
OUT_DIR_TIMEDELAY = Path("plots") / "timedelay"
X_DELAY = "tip_cue_delay_min"
TIMEDELAY_METRICS = [
    "avg_viewing_time_s",
    "avg_latency_observation_s",
    "avg_latency_confirmation_s",
    "confirmations_per_day_all_sats",
    "confirmations_per_satellite_per_day",
]

# Experiment B: off-nadir sweep at fixed delay
FIXED_DELAY_MIN_FOR_OFFNADIR = 5.0
OUT_DIR_OFFNADIR = Path("plots") / "offnadir"
X_TARGET_OFFNADIR = "off_nadir_deg"  # requirement / target (x-axis)

INCLUDE_60_DEG = False  # <-- set True to include 60deg points for ALL metrics

Y_VIEWING = "avg_viewing_time_s"
Y_GSD = "avg_gsd_m"
Y_LAT_OBS = "avg_latency_observation_s"
Y_LAT_CONF = "avg_latency_confirmation_s"
Y_CONF_ALL = "confirmations_per_day_all_sats"
Y_CONF_PER_SAT = "confirmations_per_satellite_per_day"

LABELS = {
    X_DELAY: "Time delay tip→cue (min)",
    X_TARGET_OFFNADIR: "Target off-nadir angle (deg)",
    Y_VIEWING: "Average viewing time (s)",
    Y_GSD: "Average GSD (m)",
    Y_LAT_OBS: "Average latency, observation (s)",
    Y_LAT_CONF: "Average latency, confirmation (s)",
    Y_CONF_ALL: "Confirmations per day (all sats)",
    Y_CONF_PER_SAT: "Confirmations per satellite per day",
}

# =========================


def _as_float(x: Any) -> Optional[float]:
    """_as_float(x) -> Optional[float]: Convert to float if possible."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _as_int(x: Any) -> Optional[int]:
    """_as_int(x) -> Optional[int]: Convert to int if possible."""
    if x is None:
        return None
    if isinstance(x, int):
        return int(x)
    if isinstance(x, float) and float(x).is_integer():
        return int(x)
    s = str(x).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _close(a: Optional[float], b: float, tol: float = 1e-6) -> bool:
    """_close(a, b, tol) -> bool: Float closeness check."""
    if a is None:
        return False
    return abs(a - b) <= tol


def _safe_name(s: str) -> str:
    """_safe_name(s) -> str: Make filename-safe string."""
    return "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in s)


def read_sheet_as_dicts(xlsx_path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    """read_sheet_as_dicts(xlsx_path, sheet_name) -> list[dict[str,Any]]: Read worksheet into list of row dicts."""
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"ERROR: sheet '{sheet_name}' not found in {xlsx_path}.")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []

    for r in rows[1:]:
        d: Dict[str, Any] = {}
        empty = True
        for k, v in zip(header, r):
            if k == "":
                continue
            d[k] = v
            if v is not None and str(v).strip() != "":
                empty = False
        if not empty:
            out.append(d)
    return out


def _group_key_constellation(row: Dict[str, Any]) -> Tuple[Optional[int], Optional[int]]:
    """_group_key_constellation(row) -> (orbits, sats_per_orbit)."""
    return (_as_int(row.get("orbits")), _as_int(row.get("sats_per_orbit")))


def _sorted_xy(pairs: List[Tuple[float, float]]) -> Tuple[List[float], List[float]]:
    """_sorted_xy(pairs) -> (xs, ys)."""
    pairs.sort(key=lambda t: t[0])
    return [p[0] for p in pairs], [p[1] for p in pairs]


def plot_per_seed_lines(
    rows: List[Dict[str, Any]],
    x_col: str,
    y_col: str,
    out_png: Path,
    xlim_0_60: bool,
) -> None:
    """plot_per_seed_lines(rows, x_col, y_col, out_png, xlim_0_60) -> None: One line per seed."""
    by_seed: Dict[int, List[Tuple[float, float]]] = {}

    for r in rows:
        seed = _as_int(r.get("seed"))
        x = _as_float(r.get(x_col))
        y = _as_float(r.get(y_col))
        if seed is None or x is None or y is None:
            continue
        by_seed.setdefault(seed, []).append((x, y))

    if not by_seed:
        return

    plt.figure()
    for seed in sorted(by_seed.keys()):
        xs, ys = _sorted_xy(by_seed[seed])
        plt.plot(xs, ys, marker="o", label=f"{seed}sd")

    plt.xlabel(LABELS.get(x_col, x_col))
    plt.ylabel(LABELS.get(y_col, y_col))
    plt.grid(True)
    plt.legend()
    if xlim_0_60:
        plt.xlim(0, 60)
    plt.tight_layout()

    out_png.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(out_png, dpi=200)
    plt.close()


def plot_average_line(
    rows: List[Dict[str, Any]],
    x_col: str,
    y_col: str,
    out_png: Path,
    xlim_0_60: bool,
) -> None:
    """plot_average_line(rows, x_col, y_col, out_png, xlim_0_60) -> None: Mean over seeds per x value."""
    by_x: Dict[float, List[float]] = {}

    for r in rows:
        x = _as_float(r.get(x_col))
        y = _as_float(r.get(y_col))
        if x is None or y is None:
            continue
        by_x.setdefault(x, []).append(y)

    if not by_x:
        return

    xs = sorted(by_x.keys())
    ys = [sum(by_x[x]) / float(len(by_x[x])) for x in xs]

    plt.figure()
    plt.plot(xs, ys, marker="o")
    plt.xlabel(LABELS.get(x_col, x_col))
    plt.ylabel(LABELS.get(y_col, y_col))
    plt.grid(True)
    if xlim_0_60:
        plt.xlim(0, 60)
    plt.tight_layout()

    out_png.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(out_png, dpi=200)
    plt.close()


def run_experiment_timedelay(all_rows: List[Dict[str, Any]], out_base: Path) -> None:
    """run_experiment_timedelay(all_rows, out_base) -> None: Plots for delay sweep at 40deg requirement."""
    rows = [
        r for r in all_rows
        if _close(_as_float(r.get("off_nadir_deg")), REQ_OFFNADIR_FOR_TIMEDELAY_DEG)
    ]
    if not rows:
        print("WARN: timedelay experiment: no rows found for off_nadir_deg==40")
        return

    groups: Dict[Tuple[Optional[int], Optional[int]], List[Dict[str, Any]]] = {}
    for r in rows:
        groups.setdefault(_group_key_constellation(r), []).append(r)

    for (orbits, sats), grp_rows in groups.items():
        tag = f"{orbits}x{sats}sat" if orbits is not None and sats is not None else "unknown"
        tag = _safe_name(tag)

        for y_col in TIMEDELAY_METRICS:
            plot_per_seed_lines(
                grp_rows,
                x_col=X_DELAY,
                y_col=y_col,
                out_png=out_base / f"req40_{tag}_{_safe_name(y_col)}_over_delay_seeds.png",
                xlim_0_60=False,
            )
            plot_average_line(
                grp_rows,
                x_col=X_DELAY,
                y_col=y_col,
                out_png=out_base / f"req40_{tag}_{_safe_name(y_col)}_over_delay_avg.png",
                xlim_0_60=False,
            )


def run_experiment_offnadir(all_rows: List[Dict[str, Any]], out_base: Path) -> None:
    """run_experiment_offnadir(all_rows, out_base) -> None: Off-nadir sweep at fixed 5min delay.
    X-axis is target off_nadir_deg. Optionally exclude 60deg points for selected metrics, while keeping x-axis 0..60.
    """
    rows = [
        r for r in all_rows
        if _close(_as_float(r.get("tip_cue_delay_min")), FIXED_DELAY_MIN_FOR_OFFNADIR)
    ]
    if not rows:
        print("WARN: offnadir experiment: no rows found for delay==5")
        return

    groups: Dict[Tuple[Optional[int], Optional[int]], List[Dict[str, Any]]] = {}
    for r in rows:
        groups.setdefault(_group_key_constellation(r), []).append(r)

    exclude_60_metrics = {Y_VIEWING, Y_LAT_OBS, Y_LAT_CONF, Y_GSD}
    metrics = [Y_VIEWING, Y_LAT_OBS, Y_LAT_CONF, Y_CONF_ALL, Y_CONF_PER_SAT, Y_GSD]

    for (orbits, sats), grp_rows in groups.items():
        tag = f"{orbits}x{sats}sat" if orbits is not None and sats is not None else "unknown"
        tag = _safe_name(tag)

        for y_col in metrics:
            if (not INCLUDE_60_DEG) and (y_col in exclude_60_metrics):
                use_rows = [
                    r for r in grp_rows
                    if not _close(_as_float(r.get(X_TARGET_OFFNADIR)), 60.0)
                ]
            else:
                use_rows = grp_rows

            if not use_rows:
                continue

            stub = f"{_safe_name(y_col)}_vs_target_offnadir"

            plot_per_seed_lines(
                use_rows,
                x_col=X_TARGET_OFFNADIR,
                y_col=y_col,
                out_png=out_base / f"delay5_{tag}_{stub}_seeds.png",
                xlim_0_60=True,
            )
            plot_average_line(
                use_rows,
                x_col=X_TARGET_OFFNADIR,
                y_col=y_col,
                out_png=out_base / f"delay5_{tag}_{stub}_avg.png",
                xlim_0_60=True,
            )


def main() -> None:
    """main() -> None: Generate timedelay and offnadir plots."""
    script_dir = Path(__file__).resolve().parent
    xlsx_path = script_dir / SUMMARY_XLSX_NAME

    if not xlsx_path.is_file():
        raise SystemExit(f"ERROR: {xlsx_path} not found. Run summarize_results.py first.")

    all_rows = read_sheet_as_dicts(xlsx_path, SHEET_RAW)
    if not all_rows:
        raise SystemExit("ERROR: No rows found in summary sheet.")

    out_timedelay = script_dir / OUT_DIR_TIMEDELAY
    out_offnadir = script_dir / OUT_DIR_OFFNADIR
    out_timedelay.mkdir(parents=True, exist_ok=True)
    out_offnadir.mkdir(parents=True, exist_ok=True)

    run_experiment_timedelay(all_rows, out_timedelay)
    run_experiment_offnadir(all_rows, out_offnadir)

    print(f"OK: wrote timedelay plots to {out_timedelay}")
    print(f"OK: wrote offnadir plots to {out_offnadir}")


if __name__ == "__main__":
    main()