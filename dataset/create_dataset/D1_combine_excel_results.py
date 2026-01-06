from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


FIELDS = [
    "result_name",
    "detection_id",
    "offnadir_deg",
    "offnadir_deg_round5",
    "cue_lat",
    "cue_lon",
    "cue_alt",
    "tgt_lat",
    "tgt_lon",
    "tgt_alt",
    "t_datetime",
]


def find_repo_root(start: Path) -> Path:
    """find_repo_root(start) -> Path: Walk up from start to find project root containing '0_results/0_FINAL'."""
    p = start.resolve()
    for parent in [p, *p.parents]:
        if (parent / "0_results" / "0_FINAL").is_dir():
            return parent
    raise FileNotFoundError("Could not find repo root containing '0_results/0_FINAL' when walking up from script CWD.")


def norm(s: Any) -> str:
    """norm(s) -> str: Normalize header cell to trimmed string; empty -> ''."""
    return str(s).strip() if s is not None else ""


def norm_id(v: Any) -> Any:
    """norm_id(v) -> Any: Normalize Excel-ish IDs so '1', 1, 1.0 match the same key."""
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else v
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        try:
            f = float(s)
            return int(f) if f.is_integer() else f
        except Exception:
            return s
    return v


def to_float(v: Any) -> float | None:
    """to_float(v) -> float|None: Convert a value to float if possible; otherwise None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return float(int(v))
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def round_to_nearest_5(v: Any) -> int | None:
    """round_to_nearest_5(v) -> int|None: Round numeric value to nearest 5 (0,5,10,...) using half-up behavior."""
    x = to_float(v)
    if x is None:
        return None
    # half-up rounding to nearest integer of (x/5), then *5
    q = x / 5.0
    rounded_q = int(q + 0.5) if q >= 0 else int(q - 0.5)
    return int(rounded_q * 5)


def iter_sheet_rows(xlsx_path: Path, sheet_candidates: tuple[str, ...]) -> tuple[list[str], list[dict[str, Any]]]:
    """iter_sheet_rows(xlsx_path,sheet_candidates) -> tuple[list[str], list[dict[str,Any]]]: Read rows into dicts keyed by header."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = None
        for name in sheet_candidates:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            raise KeyError(f"Sheet not found. Tried {sheet_candidates}. Sheets: {wb.sheetnames}")

        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return ([], [])

        headers = [norm(h) for h in header_row]
        header_map = {h: i for i, h in enumerate(headers) if h}

        out: list[dict[str, Any]] = []
        for r in rows:
            if r is None:
                continue
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in r):
                continue
            row_dict: dict[str, Any] = {}
            for h, idx in header_map.items():
                row_dict[h] = r[idx] if idx < len(r) else None
            out.append(row_dict)

        return (headers, out)
    finally:
        wb.close()


def iter_img_rows(xlsx_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    """iter_img_rows(xlsx_path) -> tuple[list[str], list[dict[str,Any]]]: Read all rows from sheet ' Img' or 'Img'."""
    return iter_sheet_rows(xlsx_path, (" Img", "Img"))


def iter_combined_offnadir_map(xlsx_path: Path) -> dict[Any, Any]:
    """iter_combined_offnadir_map(xlsx_path) -> dict[Any,Any]: Map detection_id -> offnadir_deg from sheet 'Combined'."""
    _, combined_rows = iter_sheet_rows(xlsx_path, ("Combined", " combined", "COMBINED"))

    det_cols = ("detection_id", "detectionId", "det_id", "id", "ID", "idx", "index")
    out: dict[Any, Any] = {}
    for i, row in enumerate(combined_rows, start=1):
        det_id = None
        for c in det_cols:
            if c in row and row[c] is not None:
                det_id = row[c]
                break
        if det_id is None:
            det_id = i
        key = norm_id(det_id)
        if key is None:
            continue
        out[key] = row.get("offnadir_deg", None)
    return out


def choose_results_xlsx(folder: Path) -> Path | None:
    """choose_results_xlsx(folder) -> Path|None: Pick the results Excel in a folder."""
    candidates = sorted(folder.glob("results_*.xlsx"))
    if candidates:
        return candidates[0]
    any_xlsx = sorted(folder.glob("*.xlsx"))
    return any_xlsx[0] if any_xlsx else None


def write_output_xlsx(out_path: Path, rows: list[list[Any]]) -> None:
    """write_output_xlsx(out_path,rows) -> None: Write rows (including header row) to a new xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "all_results"
    for r in rows:
        ws.append(r)

    for col_idx, _name in enumerate(rows[0], start=1):
        ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)

    ws.freeze_panes = "A2"
    wb.save(out_path)
    wb.close()


def main() -> int:
    """main() -> int: Aggregate all rows from 'Img' of each TC_1x1* results xlsx into one output xlsx (adds offnadir columns)."""
    cwd = Path.cwd()
    root = find_repo_root(cwd)
    final_dir = root / "0_results" / "0_FINAL"

    tc_dirs = sorted([p for p in final_dir.iterdir() if p.is_dir() and p.name.startswith("TC_1x1")])
    if not tc_dirs:
        print(f"No folders starting with 'TC_1x1' found in: {final_dir}")
        return 1

    output_rows: list[list[Any]] = [FIELDS]
    missing_col_counts: dict[str, int] = {k: 0 for k in FIELDS if k not in ("result_name", "detection_id")}
    total_written = 0
    total_skipped_files = 0

    for d in tc_dirs:
        xlsx = choose_results_xlsx(d)
        if xlsx is None:
            total_skipped_files += 1
            continue

        try:
            _, img_rows = iter_img_rows(xlsx)
            offnadir_map = iter_combined_offnadir_map(xlsx)
        except Exception as e:
            total_skipped_files += 1
            print(f"SKIP {d.name} ({xlsx.name}): {e}")
            continue

        det_cols = ("detection_id", "detectionId", "det_id", "id", "ID", "idx", "index")

        for i, row_dict in enumerate(img_rows, start=1):
            det_id = None
            for cand in det_cols:
                if cand in row_dict and row_dict[cand] is not None:
                    det_id = row_dict[cand]
                    break
            if det_id is None:
                det_id = i

            det_key = norm_id(det_id)
            offnadir_deg = offnadir_map.get(det_key, offnadir_map.get(i, None))
            offnadir_deg_round5 = round_to_nearest_5(offnadir_deg)

            out_row: list[Any] = []
            out_row.append(d.name)  # result_name
            out_row.append(det_id)  # detection_id
            out_row.append(offnadir_deg)
            out_row.append(offnadir_deg_round5)

            for k in FIELDS[4:]:
                v = row_dict.get(k, None)
                if k != "t_datetime" and isinstance(v, str) and v.strip() == "":
                    v = None
                if k != "t_datetime" and (k in row_dict) is False:
                    missing_col_counts[k] += 1
                out_row.append(v)

            output_rows.append(out_row)
            total_written += 1

    out_path = cwd / "combined_results.xlsx"
    write_output_xlsx(out_path, output_rows)

    print(f"Wrote {total_written} rows to: {out_path}")
    if total_skipped_files:
        print(f"Skipped {total_skipped_files} folder(s) with no readable xlsx/'Img'/'Combined' sheet.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
