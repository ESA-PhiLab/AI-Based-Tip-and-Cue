# main_create.py
import json
import os
import subprocess
import sys
import time
from pathlib import Path
import shutil

import openpyxl

from obtain_data import pick_random_pose


# =========================
# Path handling (match create_patch.py style)
# =========================
main_path = Path(__file__).resolve().parents[2]
os.chdir(main_path)

# Script directory: dataset/create_dataset
SCRIPT_DIR = Path(__file__).resolve().parent


def load_json(path: Path) -> dict:
    """load_json(path) -> dict: Read JSON utf-8."""
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def _rmtree_force(path: Path, retries: int = 5, delay: float = 0.5) -> None:
    """_rmtree_force(path,retries,delay) -> None: Windows-safe recursive delete with retries."""
    import time
    import stat

    def onerror(func, p, exc_info):
        try:
            os.chmod(p, stat.S_IWRITE)
            func(p)
        except Exception:
            pass

    for attempt in range(retries):
        try:
            shutil.rmtree(path, onerror=onerror)
            return
        except PermissionError:
            if attempt == retries - 1:
                raise
            time.sleep(delay)


def cleanup_previous_outputs() -> None:
    """cleanup_previous_outputs() -> None: Remove old dataset outputs before a fresh run."""
    base = Path("dataset") / "create_dataset"

    for d in ["nadir", "offnadir", "sunglint"]:
        p = base / d
        if p.exists():
            _rmtree_force(p)
            print(f"Deleted directory: {p}")

    overview = base / "dataset_overview.xlsx"
    if overview.exists():
        overview.unlink()
        print(f"Deleted file: {overview}")

    meta = base / "_meta"
    if meta.exists():
        _rmtree_force(meta)
        print(f"Deleted directory: {meta}")




def ensure_workbook(xlsx_path: Path) -> openpyxl.Workbook:
    """ensure_workbook(xlsx_path) -> Workbook: Create/load workbook with required sheets and headers."""
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()

    for name in ["settings", "patch_param", "offnadir_param", "annotations"]:
        if name not in wb.sheetnames:
            wb.create_sheet(title=name)

    if "Sheet" in wb.sheetnames:
        ws0 = wb["Sheet"]
        if ws0.max_row == 1 and ws0.max_column == 1 and ws0.cell(1, 1).value is None and len(wb.sheetnames) > 1:
            wb.remove(ws0)

    # --- settings header MUST be on row 1 ---
    ws_settings = wb["settings"]
    header = [ws_settings.cell(1, 1).value, ws_settings.cell(1, 2).value]
    if header != ["key", "value"]:
        # wipe sheet and re-add header
        ws_settings.delete_rows(1, ws_settings.max_row)
        ws_settings.append(["key", "value"])
        ws_settings.freeze_panes = "A2"

    # keep the others as you had them (unchanged)
    ws_patch = wb["patch_param"]
    if ws_patch.max_row == 1 and ws_patch.cell(1, 1).value is None:
        ws_patch.append([
            "i", "img_file", "result_name", "detection_id",
            "pick_img_seed_i", "crop_patch_seed_i",
            "patch_name", "label_simple",
            "top_left_x", "top_left_y",
            "offset_x", "offset_y",
            "fracs_json",
        ])
        ws_patch.freeze_panes = "A2"

    ws_off = wb["offnadir_param"]
    if ws_off.max_row == 1 and ws_off.cell(1, 1).value is None:
        ws_off.append([
            "i", "img_file",
            "dem_seed_i", "pick_pose_seed_i",
            "sat_lat", "sat_lon", "sat_alt",
            "tgt_lat", "tgt_lon", "tgt_alt",
            "datetime_utc",
        ])
        ws_off.freeze_panes = "A2"

    ws_ann = wb["annotations"]
    if ws_ann.max_row == 1 and ws_ann.cell(1, 1).value is None:
        ws_ann.append([
            "i", "img_file", "patch_name",
            "image_id", "annotation_id", "category_id",
            "bbox_json", "segmentation_json",
            "area", "iscrowd", "other_keys_json",
        ])
        ws_ann.freeze_panes = "A2"

    return wb


def write_settings_once(ws_settings, patch_parameters: dict, render_resolution: int) -> None:
    """write_settings_once(ws_settings,patch_parameters,render_resolution) -> None: Write general settings once in generate_patch order."""
    # header guaranteed on row 1 by ensure_workbook()
    for r in range(2, ws_settings.max_row + 1):
        if ws_settings.cell(r, 1).value is not None:
            return  # already written

    # EXACT order as generate_patch signature
    ordered_keys = [
        "mode_single",
        "mode_multiple_allow_partial",
        "window_size",
        "img_file",                 # conceptual input (fixed in main)
        "nowhale_max_fraction",
        "whale_min_fraction",
        "half_fraction_range",
        "crop_black_border",
        "crop_threshold",
        "max_tries",
        "mask_alpha",
        "plot_patch",
    ]

    # values coming from patch_parameters or fixed defaults
    for k in ordered_keys:
        if k == "img_file":
            continue  # not a generate_patch parameter you want here
        if k in patch_parameters:
            v = patch_parameters[k]
        else:
            continue

        if isinstance(v, (list, dict, tuple)):
            v = json.dumps(v)
        ws_settings.append([k, v])

    # off-nadir global setting (not part of generate_patch, but global)
    ws_settings.append(["render_resolution", int(render_resolution)])




def clear_sheet_keep_header(ws) -> None:
    """clear_sheet_keep_header(ws) -> None: Delete all rows below header."""
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)




def ann_row_from_dict(i_val: int,
                      img_file: str,
                      patch_name: str,
                      ann: dict) -> list[object]:
    """ann_row_from_dict(i_val,img_file,patch_name,ann) -> list[object]: Build an Excel row from a normalized annotation dict."""
    bbox = ann.get("bbox", None)
    seg = ann.get("segmentation", None)
    other = ann.get("other", None)

    bbox_json = json.dumps(bbox) if bbox is not None else ""
    seg_json = json.dumps(seg) if seg is not None else ""
    other_json = json.dumps(other) if other is not None else ""

    return [
        i_val,
        img_file,
        patch_name,
        ann.get("image_id", None),
        ann.get("annotation_id", None),
        ann.get("category_id", None),
        bbox_json,
        seg_json,
        ann.get("area", None),
        ann.get("iscrowd", None),
        other_json,
    ]


def run_one(i: int,
            img_file: str,
            crop_patch_seed_i: int,
            dem_seed_i: int,
            show_plot: bool,
            render_resolution: int,
            sat_lat: float, sat_lon: float, sat_alt: float,
            tgt_lat: float, tgt_lon: float, tgt_alt: float,
            datetime_utc: str,
            patch_parameters: dict,
            meta_out: Path) -> None:
    """run_one(...) -> None: Spawn one worker; it writes meta_out JSON."""
    import faulthandler
    faulthandler.enable(all_threads=True)

    shutil.rmtree(Path.home() / "AppData/Local/Temp/drjit", ignore_errors=True)

    env = os.environ.copy()

    base = Path(r"C:\drjit_temp") / f"worker_{i}"
    tmp = base / "tmp"
    tmp.mkdir(parents=True, exist_ok=True)

    env["TEMP"] = str(tmp)
    env["TMP"] = str(tmp)

    shutil.rmtree(tmp / "drjit", ignore_errors=True)

    meta_out_abs = meta_out.resolve()
    meta_out_abs.parent.mkdir(parents=True, exist_ok=True)

    cmd = [
        sys.executable, str(SCRIPT_DIR / "worker_run.py"),
        "--img_file", img_file,
        "--patch_seed", str(crop_patch_seed_i),
        "--dem_seed", str(dem_seed_i),
        "--show_plot", "1" if show_plot else "0",
        "--render_resolution", str(render_resolution),

        "--sat_lat", str(sat_lat),
        "--sat_lon", str(sat_lon),
        "--sat_alt", str(sat_alt),

        "--tgt_lat", str(tgt_lat),
        "--tgt_lon", str(tgt_lon),
        "--tgt_alt", str(tgt_alt),

        "--datetime_utc", str(datetime_utc),

        "--mode_single", str(patch_parameters["mode_single"]),
        "--mode_multiple_allow_partial", "1" if bool(patch_parameters["mode_multiple_allow_partial"]) else "0",
        "--window_size", str(int(patch_parameters["window_size"])),
        "--nowhale_max_fraction", str(float(patch_parameters["nowhale_max_fraction"])),
        "--whale_min_fraction", str(float(patch_parameters["whale_min_fraction"])),
        "--half_fraction_low", str(float(patch_parameters["half_fraction_range"][0])),
        "--half_fraction_high", str(float(patch_parameters["half_fraction_range"][1])),
        "--mask_alpha", str(int(patch_parameters["mask_alpha"])),

        "--meta_out", str(meta_out_abs),
    ]

    try:
        subprocess.run(cmd, check=True, stdout=sys.stdout, stderr=sys.stderr, env=env)
    finally:
        shutil.rmtree(base, ignore_errors=True)


def main() -> None:
    """main() -> None: Run N workers and log settings/patch/offnadir params, plus patch annotations."""

    cleanup_previous_outputs()

    img_file = "Pelagos2016/PelagosIm4_FW_WV3_PS_20160619_B2.PNG"
    render_resolution = 124

    pick_img_seed = 12
    crop_patch_seed = 42
    dem_seed = 1
    pick_pose_seed = 17

    show_plot = False

    # mode_single options:
    #   "full"      -> only full whales
    #   "half"      -> only half whales
    #   "ocean"     -> only ocean (no whales)
    #   "full_half" -> full OR half whales
    #   "all"       -> anything
    #
    # mode_multiple_allow_partial:
    #   True  -> if multiple whales, allow other partial whales in the patch
    #   False -> forbid any whale in (nowhale_max_fraction, whale_min_fraction)


    patch_parameters = {
        "mode_single": "ocean",
        "mode_multiple_allow_partial": False,
        "window_size": 64,
        "nowhale_max_fraction": 0.10,
        "whale_min_fraction": 0.99,
        "half_fraction_range": (0.20, 0.80),
        "mask_alpha": 80,
    }

    # Do it as before: relative to script folder
    poses_xlsx = SCRIPT_DIR / "combined_results.xlsx"
    overview_xlsx = SCRIPT_DIR / "dataset_overview.xlsx"

    if not poses_xlsx.is_file():
        raise FileNotFoundError(f"Missing poses file: {poses_xlsx}")

    wb = ensure_workbook(overview_xlsx)
    ws_settings = wb["settings"]
    ws_patch = wb["patch_param"]
    ws_off = wb["offnadir_param"]
    ws_ann = wb["annotations"]

    # Correctly add settings once
    write_settings_once(ws_settings, patch_parameters, render_resolution)

    # Meta folder next to script
    meta_dir = (SCRIPT_DIR / "_meta").resolve()
    meta_dir.mkdir(parents=True, exist_ok=True)

    n_runs = 5
    for i in range(n_runs):
        pick_img_seed_i = pick_img_seed + i
        crop_patch_seed_i = crop_patch_seed + i
        dem_seed_i = dem_seed + i
        pick_pose_seed_i = pick_pose_seed + i

        result_name, detection_id, sat_lat, sat_lon, sat_alt, tgt_lat, tgt_lon, tgt_alt, datetime_utc = pick_random_pose(
            poses_xlsx, pick_pose_seed=pick_pose_seed_i
        )

        meta_out = meta_dir / f"run_{i:04d}.json"
        if meta_out.exists():
            meta_out.unlink()

        print(f"\n ====================== Start new process {i} ====================== \n")
        print(f"Pose: sat=({sat_lat},{sat_lon},{sat_alt}) tgt=({tgt_lat},{tgt_lon},{tgt_alt}) dt={datetime_utc}")

        run_one(
            i=i,
            img_file=img_file,
            crop_patch_seed_i=crop_patch_seed_i,
            dem_seed_i=dem_seed_i,
            show_plot=show_plot,
            render_resolution=render_resolution,
            sat_lat=sat_lat, sat_lon=sat_lon, sat_alt=sat_alt,
            tgt_lat=tgt_lat, tgt_lon=tgt_lon, tgt_alt=tgt_alt,
            datetime_utc=datetime_utc,
            patch_parameters=patch_parameters,
            meta_out=meta_out,
        )

        if not meta_out.exists():
            raise FileNotFoundError(f"Worker did not write meta_out: {meta_out}")

        meta = json.loads(meta_out.read_text(encoding="utf-8"))

        top_left = meta.get("top_left", [None, None])
        offset_xy = meta.get("offset_xy", [None, None])
        fracs = meta.get("fracs", [])
        fracs_json = json.dumps(fracs) if isinstance(fracs, list) else str(fracs)

        patch_name = meta.get("patch_name", "")
        label_simple = meta.get("label_simple", "")

        ws_patch.append([
            i,
            img_file,
            result_name,
            detection_id,
            pick_img_seed_i,
            crop_patch_seed_i,
            patch_name,
            label_simple,
            top_left[0] if isinstance(top_left, list) and len(top_left) == 2 else None,
            top_left[1] if isinstance(top_left, list) and len(top_left) == 2 else None,
            offset_xy[0] if isinstance(offset_xy, list) and len(offset_xy) == 2 else None,
            offset_xy[1] if isinstance(offset_xy, list) and len(offset_xy) == 2 else None,
            fracs_json,
        ])

        ws_off.append([
            i,
            img_file,
            dem_seed_i,
            pick_pose_seed_i,
            sat_lat,
            sat_lon,
            sat_alt,
            tgt_lat,
            tgt_lon,
            tgt_alt,
            datetime_utc,
        ])

        anns_patch = meta.get("anns_patch", [])
        if isinstance(anns_patch, list) and anns_patch:
            for ann in anns_patch:
                if isinstance(ann, dict):
                    ws_ann.append(ann_row_from_dict(i, img_file, patch_name, ann))
        else:
            ws_ann.append([i, img_file, patch_name, None, None, None, "", "", None, None, ""])

        wb.save(overview_xlsx)
        time.sleep(0.1)

    wb.save(overview_xlsx)
    wb.close()


if __name__ == "__main__":
    main()
