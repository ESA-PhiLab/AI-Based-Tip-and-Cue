# read_and_write_data.py
from __future__ import annotations

import json
import os
import random
import shutil
from pathlib import Path
from typing import Any
import numpy as np


import openpyxl

import random
import re
from pathlib import Path
from typing import Optional, Tuple, Any

from pathlib import Path
from typing import Iterable, Sequence


from typing import Any, Optional
from pathlib import Path
import random



# =========================
# Windows-safe delete helpers
# =========================
def _rmtree_force(path: Path, retries: int = 5, delay: float = 0.5) -> None:
    """_rmtree_force(path,retries,delay) -> None: Windows-safe recursive delete with retries."""
    import time
    import stat

    def onerror(func, p, exc_info):
        try:
            os.chmod(p, stat.S_IWRITE)
            func(p)
        except Exception:
            pass

    for attempt in range(retries):
        try:
            shutil.rmtree(path, onerror=onerror)
            return
        except PermissionError:
            if attempt == retries - 1:
                raise
            time.sleep(delay)


def cleanup_previous_outputs(base: Path) -> None:
    """cleanup_previous_outputs(base) -> None: Remove old dataset outputs before a fresh run."""
    for d in [
    "patch_raw_255",
    "patch_raw_rot_255",
    "texture_nadir_255",
    "radiance_nadir_255",
    "radiance_nadir_npy",
    "reflection_nadir_255",
    "reflection_nadir_npy",
    "texture_offnadir_255",
    "radiance_offnadir_255",
    "radiance_offnadir_npy",
    "reflection_offnadir_255",
    "reflection_offnadir_npy",
]:
        p = base / d
        if p.exists():
            _rmtree_force(p)
            print(f"Deleted directory: {p}")

    overview = base / "dataset_overview.xlsx"
    if overview.exists():
        overview.unlink()
        print(f"Deleted file: {overview}")

    meta = base / "_meta"
    if meta.exists():
        _rmtree_force(meta)
        print(f"Deleted directory: {meta}")


def cleanup_meta_only(base: Path) -> None:
    """cleanup_meta_only(base) -> None: Remove only _meta folder after run."""
    meta = base / "_meta"
    if meta.exists():
        _rmtree_force(meta)
        print(f"Deleted directory: {meta}")


# =========================
# Excel: pose selection
# =========================
def _load_excel_cache(xlsx_path: Path) -> tuple[list[str], list[list[Any]]]:
    """_load_excel_cache(xlsx_path) -> (list[str],list[list[Any]]): Load header + all data rows from first sheet."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip() for h in next(it)]
        rows = [list(r) for r in it if r and any(v is not None for v in r)]
        return header, rows
    finally:
        wb.close()




_OFFNADIR_RE = re.compile(r"(?P<ang>\d+(?:\.\d+)?)\s*deg", re.IGNORECASE)


def _extract_offnadir_angle_from_result_name(result_name: Any) -> Optional[float]:
    """_extract_offnadir_angle_from_result_name(result_name) -> Optional[float]: Extract angle in degrees from '...10deg...' style names; returns float or None."""
    if result_name is None:
        return None
    m = _OFFNADIR_RE.search(str(result_name))
    return float(m.group("ang")) if m else None


def _normalize_offnadir_angle(offnadir_angle: Any) -> float:
    """_normalize_offnadir_angle(offnadir_angle) -> float: Normalize input (e.g., 10, '10', '10deg', 10.0) into a float degrees value."""
    if offnadir_angle is None:
        raise ValueError("offnadir_angle is None; only call _normalize_offnadir_angle when an angle is provided.")

    if isinstance(offnadir_angle, (int, float)):
        return float(offnadir_angle)

    s = str(offnadir_angle).strip().lower()
    m = _OFFNADIR_RE.search(s)
    if m:
        return float(m.group("ang"))

    # allow plain numeric strings like "10" or "10.0"
    try:
        return float(s)
    except ValueError as e:
        raise ValueError(f"Could not parse offnadir_angle={offnadir_angle!r}. Use e.g. 10, 30, '10deg'.") from e



def to_float(v: Any) -> float | None:
    """to_float(v) -> float|None: Convert value to float if possible; otherwise None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return float(int(v))
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def round_to_nearest_5(v: Any) -> int | None:
    """round_to_nearest_5(v) -> int|None: Round numeric value to nearest 5 (0,5,10,...) using half-up behavior."""
    x = to_float(v)
    if x is None:
        return None
    q = x / 5.0
    rq = int(q + 0.5) if q >= 0 else int(q - 0.5)
    return int(rq * 5)


def pick_random_pose(
    xlsx_path: Path,
    pick_pose_seed: int,
    offnadir_angle: Optional[object] = None,
    selection_method: str = "exact",        # exact or mission
) -> tuple:
    """pick_random_pose(xlsx_path,pick_pose_seed,offnadir_angle=None,selection_method='mission') -> tuple:
    Pick one random row. 'mission' filters via result_name (legacy), 'exact' via offnadir_deg_round5."""
    random.seed(pick_pose_seed)

    header, rows = _load_excel_cache(xlsx_path)
    if not rows:
        raise ValueError("No data rows found in Excel file.")

    if offnadir_angle is not None:
        target_ang = round_to_nearest_5(offnadir_angle)
        if target_ang is None:
            raise ValueError(f"offnadir_angle={offnadir_angle!r} is not numeric.")

        filtered = []
        present_angles: set[int] = set()

        if selection_method == "mission":
            # legacy behavior: derive angle from result_name
            for row in rows:
                d_tmp = dict(zip(header, row))
                ang = _extract_offnadir_angle_from_result_name(d_tmp.get("result_name"))
                if ang is not None:
                    present_angles.add(ang)
                if ang is not None and ang == target_ang:
                    filtered.append(row)

            if not filtered:
                available = ", ".join(str(a) for a in sorted(present_angles)) or "(none found in result_name)"
                raise ValueError(
                    f"No rows match offnadir_angle={target_ang}deg (derived from result_name). "
                    f"Available angles: {available}."
                )

        elif selection_method == "exact":
            # new behavior: use explicit column
            if "offnadir_deg_round5" not in header:
                raise KeyError("Required column 'offnadir_deg_round5' not found in Excel header.")

            idx_ang = header.index("offnadir_deg_round5")
            for row in rows:
                ang = round_to_nearest_5(row[idx_ang] if idx_ang < len(row) else None)
                if ang is not None:
                    present_angles.add(ang)
                if ang is not None and ang == target_ang:
                    filtered.append(row)

            if not filtered:
                available = ", ".join(str(a) for a in sorted(present_angles)) or "(none found in offnadir_deg_round5)"
                raise ValueError(
                    f"No rows match offnadir_angle={target_ang}deg (using offnadir_deg_round5). "
                    f"Available angles: {available}."
                )

        else:
            raise ValueError("selection_method must be 'mission' or 'exact'.")

        rows = filtered

    row = random.choice(rows)
    d = dict(zip(header, row))

    result_name = d["result_name"]
    detection_id = d["detection_id"]
    sat_lat = float(d["cue_lat"])
    sat_lon = float(d["cue_lon"])
    sat_alt = float(d["cue_alt"])
    tgt_lat = float(d["tgt_lat"])
    tgt_lon = float(d["tgt_lon"])
    tgt_alt = float(d["tgt_alt"])
    datetime_utc = str(d["t_datetime"])

    return result_name, detection_id, sat_lat, sat_lon, sat_alt, tgt_lat, tgt_lon, tgt_alt, datetime_utc





def count_images_in_subfolders(dataset_root: str | Path, allowed_ext: Sequence[str] = (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp")) -> int:
    """count_images_in_subfolders(dataset_root,allowed_ext) -> int: Count images only inside first-level subfolders of dataset_root."""
    root = Path(dataset_root)
    if not root.is_dir():
        raise FileNotFoundError(f"Dataset root not found: {root}")

    allowed = {e.lower() for e in allowed_ext}
    total = 0

    for sub in root.iterdir():
        if not sub.is_dir():
            continue
        for f in sub.iterdir():
            if f.is_file() and f.suffix.lower() in allowed:
                total += 1

    return total


# =========================
# Excel: dataset_overview workbook
# =========================
def ensure_workbook(xlsx_path: Path) -> openpyxl.Workbook:
    """ensure_workbook(xlsx_path) -> Workbook: Create/load workbook with required sheets and headers."""
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()

    for name in ["settings", "patch_param", "offnadir_param", "annotations_nadir", "annotations_offnadir", "rad_refl_values"]:
        if name not in wb.sheetnames:
            wb.create_sheet(title=name)


    if "Sheet" in wb.sheetnames:
        ws0 = wb["Sheet"]
        if ws0.max_row == 1 and ws0.max_column == 1 and ws0.cell(1, 1).value is None and len(wb.sheetnames) > 1:
            wb.remove(ws0)

    # --- settings header MUST be on row 1 ---
    ws_settings = wb["settings"]
    header = [ws_settings.cell(1, 1).value, ws_settings.cell(1, 2).value]
    if header != ["key", "value"]:
        ws_settings.delete_rows(1, ws_settings.max_row)
        ws_settings.append(["key", "value"])
        ws_settings.freeze_panes = "A2"

    ws_patch = wb["patch_param"]
    if ws_patch.max_row == 1 and ws_patch.cell(1, 1).value is None:
        ws_patch.append([
            "i", "img_file", "result_name", "detection_id",
            "pick_img_seed_i", "crop_patch_seed_i",
            "patch_name",
            "top_left_x", "top_left_y",
            "offset_x", "offset_y", "rotation_angle_deg", "mirror_bool",
            "fracs_json", "label_simple", "category_id"
        ])
        ws_patch.freeze_panes = "A2"

    ws_off = wb["offnadir_param"]
    if ws_off.max_row == 1 and ws_off.cell(1, 1).value is None:
        ws_off.append([
            "i", "img_file",
            "dem_seed_i", "pick_pose_seed_i",
            "sat_lat", "sat_lon", "sat_alt",
            "tgt_lat", "tgt_lon", "tgt_alt",
            "datetime_utc", "wind_speed", "offnadir_deg",
        ])

        ws_off.freeze_panes = "A2"

    ws_rr = wb["rad_refl_values"]
    if ws_rr.max_row == 1 and ws_rr.cell(1, 1).value is None:
        ws_rr.append([
            "i", "img_file", "patch_name",
            "offnadir_deg",
            "radiance_min_nadir", "radiance_max_nadir", "radiance_mean_nadir",
            "reflectance_min_nadir", "reflectance_max_nadir", "reflectance_mean_nadir",
            "radiance_min_offnadir", "radiance_max_offnadir", "radiance_mean_offnadir",
            "reflectance_min_offnadir", "reflectance_max_offnadir", "reflectance_mean_offnadir",
        ])
        ws_rr.freeze_panes = "A2"

    def _init_ann_sheet(ws) -> None:
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append([
                "i", "img_file", "patch_name",
                "image_id", "annotation_id", "category_id",
                "bbox_json", "segmentation_json",
                "area", "iscrowd", "other_keys_json",
            ])
            ws.freeze_panes = "A2"




    ws_ann_nadir = wb["annotations_nadir"]
    ws_ann_off = wb["annotations_offnadir"]
    _init_ann_sheet(ws_ann_nadir)
    _init_ann_sheet(ws_ann_off)


    return wb


def write_settings_once(ws_settings, patch_parameters: dict, sensor_characteristics: dict) -> None:
    """write_settings_once(ws_settings,patch_parameters,render_resolution) -> None: Write general settings once in generate_patch order."""
    for r in range(2, ws_settings.max_row + 1):
        if ws_settings.cell(r, 1).value is not None:
            return  # already written

    # EXACT order as generate_patch signature
    ordered_keys = [
        "mode_single",
        "mode_multiple_allow_partial",
        "window_size",
        "img_file",                 # conceptual input (fixed in main)
        "nowhale_max_fraction",
        "whale_min_fraction",
        "half_fraction_range",
        "crop_black_border",
        "crop_threshold",
        "max_tries",
        "mask_alpha",
        "plot_patch",
    ]

    for k in ordered_keys:
        if k == "img_file":
            continue
        if k not in patch_parameters:
            continue

        v = patch_parameters[k]
        if isinstance(v, (list, dict, tuple)):
            v = json.dumps(v)
        ws_settings.append([k, v])

    # off-nadir global setting (not part of generate_patch, but global)
    ws_settings.append(["resolution", sensor_characteristics['resolution']])
    ws_settings.append(["specular_weight", sensor_characteristics['specular_weight']])


def ann_row_from_dict(i_val: int, img_file: str, patch_name: str, ann: dict) -> list[object]:
    """ann_row_from_dict(i_val,img_file,patch_name,ann) -> list[object]: Build an Excel row from a normalized annotation dict."""
    bbox = ann.get("bbox", None)
    seg = ann.get("segmentation", None)

    other = ann.get("other", None)

    bbox_json = json.dumps(bbox) if bbox is not None else ""
    seg_json = json.dumps(seg) if (seg is not None and seg != [] and seg != [[]]) else ""
    other_json = json.dumps(other) if other is not None else ""

    return [
        i_val,
        img_file,
        patch_name,
        ann.get("image_id", None),
        ann.get("annotation_id", None),
        ann.get("category_id", None),
        bbox_json,
        seg_json,
        ann.get("area", None),
        ann.get("iscrowd", None),
        other_json,
    ]

def open_overview_book(xlsx_path: Path,
                       patch_parameters: dict,
                       sensor_characteristics: dict) -> tuple[openpyxl.Workbook, Any, Any, Any, Any, Any]:
    """open_overview_book(xlsx_path,patch_parameters,sensor_characteristics) -> (wb,ws_settings,ws_patch,ws_off,ws_ann): Open workbook + ensure headers + write settings once."""
    wb = ensure_workbook(xlsx_path)
    ws_settings = wb["settings"]
    ws_patch = wb["patch_param"]
    ws_off = wb["offnadir_param"]
    ws_ann_nadir = wb["annotations_nadir"]
    ws_ann_off = wb["annotations_offnadir"]
    ws_radrefl = wb["rad_refl_values"]

    write_settings_once(ws_settings, patch_parameters, sensor_characteristics)
    return wb, ws_settings, ws_patch, ws_off, ws_ann_nadir, ws_ann_off, ws_radrefl


def append_run_rows(ws_patch,
                    ws_off,
                    ws_ann_nadir,
                    ws_ann_off,
                    ws_radrefl,
                    i: int,
                    img_file: str,
                    result_name: str,
                    detection_id: str,
                    pick_img_seed_i: int,
                    crop_patch_seed_i: int,
                    dem_seed_i: int,
                    pick_pose_seed_i: int,
                    sat_lat: float, sat_lon: float, sat_alt: float,
                    tgt_lat: float, tgt_lon: float, tgt_alt: float,
                    datetime_utc: str,
                    wind_speed: float,
                    meta: dict) -> None:


    """append_run_rows(...) -> None: Append patch/offnadir/annotation rows for one run."""
    top_left = meta.get("top_left", [None, None])
    offset_xy = meta.get("offset_xy", [None, None])
    fracs = meta.get("fracs", [])
    fracs_json = json.dumps(fracs) if isinstance(fracs, list) else str(fracs)
    fracs_json = float(fracs_json.strip("[]"))

    patch_name = meta.get("patch_name", "")
    label_simple = meta.get("label_simple", "")
    category_id = meta.get("category_id", None)
    offnadir_deg = meta.get("offnadir_deg", None)
    rotation_angle_deg = meta.get("rotation_angle_deg", None)
    mirror_bool = meta.get("mirror_bool", None)


    if offnadir_deg is not None:
        try:
            offnadir_deg = float(offnadir_deg)
        except Exception:
            offnadir_deg = None

    if rotation_angle_deg is not None:
        try:
            rotation_angle_deg = float(rotation_angle_deg)
        except Exception:
            rotation_angle_deg = None

    if mirror_bool is not None:
        try:
            mirror_bool = bool(mirror_bool)
        except Exception:
            mirror_bool = None

    ws_patch.append([
        i,
        img_file,
        result_name,
        detection_id,
        pick_img_seed_i,
        crop_patch_seed_i,
        patch_name,
        top_left[0] if isinstance(top_left, list) and len(top_left) == 2 else None,
        top_left[1] if isinstance(top_left, list) and len(top_left) == 2 else None,
        offset_xy[0] if isinstance(offset_xy, list) and len(offset_xy) == 2 else None,
        offset_xy[1] if isinstance(offset_xy, list) and len(offset_xy) == 2 else None,
        rotation_angle_deg,
        mirror_bool,
        fracs_json,
        label_simple,
        category_id
    ])

    ws_off.append([
        i,
        img_file,
        dem_seed_i,
        pick_pose_seed_i,
        sat_lat,
        sat_lon,
        sat_alt,
        tgt_lat,
        tgt_lon,
        tgt_alt,
        datetime_utc,
        float(wind_speed),
        offnadir_deg
    ])

    anns_nadir = meta.get("anns_nadir", [])
    if isinstance(anns_nadir, list) and anns_nadir:
        for ann in anns_nadir:
            if isinstance(ann, dict):
                ws_ann_nadir.append(ann_row_from_dict(i, img_file, patch_name, ann))
    else:
        ws_ann_nadir.append([i, img_file, patch_name, None, None, None, "", "", None, None, ""])

    anns_off = meta.get("anns_offnadir", [])
    if isinstance(anns_off, list) and anns_off:
        for ann in anns_off:
            if isinstance(ann, dict):
                ws_ann_off.append(ann_row_from_dict(i, img_file, patch_name, ann))
    else:
        ws_ann_off.append([i, img_file, patch_name, None, None, None, "", "", None, None, ""])

    def _get_stats(meta_key: str) -> tuple[float | None, float | None, float | None]:
        """_get_stats(meta_key) -> (min,max,mean): Read stats dict; convert NaN to None."""
        v = meta.get(meta_key, None)
        if not isinstance(v, dict):
            return None, None, None

        def _clean(x: object) -> float | None:
            if x is None:
                return None
            try:
                y = float(x)
            except Exception:
                return None
            return None if (not np.isfinite(y)) else y

        return _clean(v.get("min")), _clean(v.get("max")), _clean(v.get("mean"))

    rmin_n, rmax_n, rmean_n = _get_stats("rad_stats_nadir")
    fmin_n, fmax_n, fmean_n = _get_stats("refl_stats_nadir")

    rmin_o, rmax_o, rmean_o = _get_stats("rad_stats_offnadir")
    fmin_o, fmax_o, fmean_o = _get_stats("refl_stats_offnadir")

    # One row per image/patch
    if any(v is not None for v in (
            rmin_n, rmax_n, rmean_n, fmin_n, fmax_n, fmean_n,
            rmin_o, rmax_o, rmean_o, fmin_o, fmax_o, fmean_o,
            offnadir_deg
    )):
        ws_radrefl.append([
            i, img_file, patch_name,
            offnadir_deg,
            rmin_n, rmax_n, rmean_n,
            fmin_n, fmax_n, fmean_n,
            rmin_o, rmax_o, rmean_o,
            fmin_o, fmax_o, fmean_o,
        ])


